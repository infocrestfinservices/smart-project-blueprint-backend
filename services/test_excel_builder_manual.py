"""
test_excel_builder_manual.py

Test for the final orchestration layer, per the requested spec: mock a template
definition, create a temporary workbook, call build_excel_workbook(...), reload, and
verify the workbook was created, values were written, and validation/metadata are
preserved with a valid output path.

build_excel_workbook now lives in services/excel_builder.py (alongside the existing
build_excel), so this test imports it from there.

Run from backend/:
    python services/test_excel_builder_manual.py
"""

import math
import os
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

BACKEND_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND_DIR))

from services.excel_builder import build_excel_workbook  # noqa: E402

fails = []


def ok(cond, label):
    print(f"   {'OK ' if cond else 'FAIL'}  {label}")
    if not cond:
        fails.append(label)


ASSUMPTIONS = {
    "name_of_unit": "Demo Manufacturing Pvt Ltd", "constitution": "Private Limited",
    "line_of_activity": "Precision components", "industry_type": "Manufacturing",
    "term_loan_amount": 4_000_000, "promoters_capital": 2_000_000,
    "interest_rate_term_loan": 0.11, "term_loan_tenure_months": 60,
    "moratorium_months": 0, "interest_rate_wc": 0.12, "income_tax_rate": 0.25,
    "installed_capacity": 300_000,
    "capacity_utilisation_y1_y5": [0.6, 0.7, 0.75, 0.8, 0.85],
    "monthly_seasonality_weights": [1] * 12,
    "selling_price_y1": 100, "selling_price_escalation": 0.05,
    "cost1_per_unit_y1": 40, "cost1_escalation": 0.05,
    "cost2_per_unit_y1": 15, "cost2_escalation": 0.05,
    "other_variable_cost_y1": 5, "other_variable_escalation": 0.05,
    "wages_monthly_y1": 150_000, "wages_escalation": 0.08,
    "factory_overheads_monthly_y1": 80_000, "factory_oh_escalation": 0.06,
    "repairs_maintenance_monthly_y1": 20_000, "rm_escalation": 0.06,
    "admin_expenses_monthly_y1": 40_000, "admin_escalation": 0.06,
    "selling_distribution": 0.02,
    "land_cost": 500_000, "building_cost": 2_000_000, "building_dep_rate": 0.10,
    "plant_machinery_cost": 4_000_000, "plant_machinery_dep_rate": 0.15,
    "furniture_other_cost": 500_000, "furniture_dep_rate": 0.15,
    "raw_material_holding_days": 30, "finished_goods_holding_days": 15,
    "receivables_days": 45, "payables_days": 30,
    "min_cash_balance": 100_000, "wc_margin_pct": 0.25,
    "discount_rate": 0.12,
}

# 1. mock template definition
TEMPLATE_DEFINITION = {
    "Dashboard": {"Revenue": ["B5", "C5", "D5", "E5", "F5"], "IRR": "B7", "NPV": "B8"},
    "ProfitLoss": {"pat": ["B5", "C5", "D5", "E5", "F5"]},
}


def make_template(path):
    """2. a temporary workbook with the sheets/cells the definition targets."""
    wb = Workbook()
    dash = wb.active
    dash.title = "Dashboard"
    for c in ("B5", "C5", "D5", "E5", "F5", "B7", "B8"):
        dash[c] = 0
    pl = wb.create_sheet("ProfitLoss")
    for c in ("B5", "C5", "D5", "E5", "F5"):
        pl[c] = 0
    wb.save(path)


def main():
    tmp = tempfile.mkdtemp(prefix="excelbuilder_")
    template_path = os.path.join(tmp, "template.xlsx")
    output_path = os.path.join(tmp, "output.xlsx")
    make_template(template_path)

    print("=" * 74)
    print("EXCEL BUILDER — final orchestration layer")
    print("=" * 74)

    # 3. call the orchestrator
    result = build_excel_workbook(ASSUMPTIONS, template_path, output_path, TEMPLATE_DEFINITION)

    print("\nReturn value:")
    print(f"   status={result['status']!r}  output_path set={bool(result.get('output_path'))}")
    print(f"   metadata={result['metadata']}")
    print(f"   validation.passed={result['validation']['passed']}  "
          f"errors={len(result['validation']['errors'])}")

    print("\n5. Verifications:")
    # workbook created / returned output path exists
    ok(result["status"] == "success", "returned status == 'success'")
    ok(result["output_path"] == output_path and os.path.isfile(output_path),
       "returned output_path exists on disk")

    # 4. reload workbook
    wb = load_workbook(output_path)
    ok(wb is not None, "workbook created and reloads")
    ok(wb.sheetnames == ["Dashboard", "ProfitLoss"], f"sheets: {wb.sheetnames}")

    # values written (Year-1 revenue is deterministic: 300000 * 0.6 * 100 = 18,000,000)
    ok(wb["Dashboard"]["B5"].value == 18_000_000, "values written (Dashboard!B5 == 18,000,000)")
    ok(all(wb["Dashboard"][c].value not in (0, None) for c in ("C5", "D5", "E5", "F5", "B7", "B8")),
       "all mapped Dashboard cells populated (no longer 0)")
    ok(all(wb["ProfitLoss"][c].value not in (0, None) for c in ("B5", "C5", "D5", "E5", "F5")),
       "all mapped ProfitLoss cells populated")

    # validation preserved / metadata preserved
    ok(result["validation"]["passed"] is True, "validation preserved (passed == True)")
    ok(result["metadata"].get("engine_version") == "1.0"
       and result["metadata"].get("status") == "success", "metadata preserved")

    print("\nMalformed inputs raise ValueError:")
    for label, fn in [
        ("assumptions not a dict",
         lambda: build_excel_workbook("x", template_path, output_path, TEMPLATE_DEFINITION)),
        ("workbook does not exist",
         lambda: build_excel_workbook(ASSUMPTIONS, os.path.join(tmp, "nope.xlsx"),
                                      output_path, TEMPLATE_DEFINITION)),
        ("template_definition not a dict",
         lambda: build_excel_workbook(ASSUMPTIONS, template_path, output_path, None)),
    ]:
        try:
            fn(); ok(False, f"{label} -> should have raised")
        except ValueError as e:
            ok(True, f"{label} -> ValueError: {str(e).split(':',1)[1].strip()[:34]}")

    print("\n" + "=" * 74)
    if fails:
        print(f"FAILED — {len(fails)} check(s): {fails}")
        raise AssertionError("build_excel_workbook did not behave as specified")
    print("PASSED — workbook created, values written, validation & metadata preserved,")
    print("output path exists, malformed inputs rejected.")
    print("=" * 74)
    return 0


if __name__ == "__main__":
    sys.exit(main())
