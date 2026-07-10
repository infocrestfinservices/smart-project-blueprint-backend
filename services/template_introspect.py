"""
template_introspect.py

Auto-extracts a template's *input schema* — the labeled assumption/input cells a
user should fill — from its sample workbook. This is the raw first pass; the
resulting <template_id>.schema.json is then lightly hand-curated (trim noise,
fix labels, group). The template-fill engine writes user answers back into these
exact cells (see template_fill_service.py).

An "input" is either (a) a cell the designer colour-coded as a hard-coded input
(blue font — the near-universal financial-model convention, e.g. this workbook's
legend "Blue = hard-coded input"), or, failing any colour coding, (b) a numeric
CONSTANT (not a formula) on an assumption/input sheet. Colour detection is what
lets us read a BLANK template whose input cells are still empty. We attach the
row's text label plus the unit/basis text beside it, and infer a type from the
cell's number format.

Run as a script to (re)generate raw schemas for every template_fill template:
    python -m services.template_introspect          # from backend/
"""

import json
import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Sheet-name hints that usually hold user inputs (not computed outputs).
_INPUT_SHEET_HINTS = (
    "assumption", "input", "cover", "cost of project", "general", "costing",
    "market data", "startup", "project profile", "sources", "construction budget",
    "data backed", "notes", "profile",
)

_MAX_COLS = 30            # ignore monthly/blown-out sheets beyond this width
_MAX_ROWS = 160
_MAX_FIELDS = 140         # cap raw fields; curation trims further


def _is_input_sheet(name: str) -> bool:
    n = (name or "").lower()
    return any(h in n for h in _INPUT_SHEET_HINTS)


def _infer_type(cell) -> str:
    fmt = (cell.number_format or "").lower()
    if "%" in fmt:
        return "percent"
    if "yyyy" in fmt or "mmm" in fmt or "dd" in fmt:
        return "date"
    return "number"


def _row_label(ws, row: int, value_col: int) -> str:
    """First text cell to the LEFT of the value on this row; fall back to the
    nearest text label above in column A/B."""
    for c in range(1, value_col):
        v = ws.cell(row=row, column=c).value
        if isinstance(v, str) and len(v.strip()) > 1 and not v.startswith("="):
            return v.strip()
    for r in range(row - 1, max(row - 6, 0), -1):
        for c in (1, 2):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and len(v.strip()) > 1 and not v.startswith("="):
                return v.strip()
    return ""


def _column_header(ws, row: int, col: int) -> str:
    """A short header for a value column (used to disambiguate multi-value rows),
    scanning up to 8 rows above for a text cell in the same column."""
    for r in range(row - 1, max(row - 9, 0), -1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and len(v.strip()) > 1 and not v.startswith("="):
            return v.strip()[:24]
    return ""


# Font colours designers use for "hard-coded input, change me" cells. Pure blue
# (FF0000FF) is the classic convention; hyperlink blue (FF0563C1) is NOT an input
# and is deliberately excluded.
_INPUT_FONT_COLORS = {"FF0000FF"}


def _font_color(cell) -> str:
    c = cell.font.color if cell.font else None
    if c is not None and getattr(c, "type", None) == "rgb" and isinstance(c.rgb, str):
        return c.rgb.upper()
    return ""


def _is_blue_input(cell) -> bool:
    """A colour-coded input: blue font and not itself a formula."""
    if isinstance(cell.value, str) and cell.value.startswith("="):
        return False
    return _font_color(cell) in _INPUT_FONT_COLORS


def _infer_type_blue(cell) -> str:
    """As _infer_type, but a pre-filled string placeholder (e.g. a cover-page
    '[ To be filled ]') marks a free-text input."""
    if isinstance(cell.value, str) and not cell.value.startswith("="):
        return "text"
    return _infer_type(cell)


def _context_right(ws, row: int, col: int, span: int = 4) -> str:
    """Non-formula text just to the right of a value (unit, basis, source) — folded
    into a single hint string that grounds the AI when it fills the cell."""
    out = []
    for c in range(col + 1, col + 1 + span):
        v = ws.cell(row=row, column=c).value
        if isinstance(v, str) and v.strip() and not v.startswith("="):
            out.append(v.strip())
    return " · ".join(out)[:120]


def _extract_blue_schema(wb, template_id: str, currency: str, path: str) -> dict:
    """Schema from colour-coded (blue-font) input cells across ALL sheets. Works on
    a blank template because it keys off styling, not on cells already holding
    numbers. Repeated labels (e.g. a monthly ramp row) are disambiguated by the
    column header above them (M1…M60)."""
    from collections import Counter

    groups = []
    field_count = 0
    for ws in wb.worksheets:
        if ws.max_row > 5000:
            continue
        blue = [c for row in ws.iter_rows() for c in row if _is_blue_input(c)]
        if not blue:
            continue
        base = {c.coordinate: (_row_label(ws, c.row, c.column) or ws.title) for c in blue}
        dupes = Counter(base.values())
        fields = []
        for cell in blue:
            if field_count >= _MAX_FIELDS:
                break
            label = base[cell.coordinate]
            if dupes[label] > 1:
                hdr = _column_header(ws, cell.row, cell.column)
                label = f"{label} — {hdr}" if hdr else f"{label} ({cell.coordinate})"
            fields.append({
                "cell": cell.coordinate,
                "label": label[:80],
                "hint": _context_right(ws, cell.row, cell.column),
                "default": None if isinstance(cell.value, str) else cell.value,
                "type": _infer_type_blue(cell),
            })
            field_count += 1
        if fields:
            groups.append({"title": ws.title, "sheet": ws.title, "fields": fields})
        if field_count >= _MAX_FIELDS:
            break

    return {
        "template_id": template_id,
        "file": os.path.basename(path),
        "currency": currency,
        "truncated": field_count >= _MAX_FIELDS,
        "groups": groups,
    }


def extract_schema(path: str, template_id: str, currency: str = "") -> dict:
    """Return a raw input schema dict for a template workbook.

    Prefer colour-coded (blue-font) inputs — the only signal that survives a blank
    template. Fall back to the legacy "numeric constant on an input sheet" scan for
    pre-filled sample workbooks that carry no input colouring.
    """
    wb = load_workbook(path, data_only=False, keep_vba=path.endswith(".xlsm"))

    blue = _extract_blue_schema(wb, template_id, currency, path)
    if any(g["fields"] for g in blue["groups"]):
        wb.close()
        return blue

    groups = []
    field_count = 0

    for ws in wb.worksheets:
        if not _is_input_sheet(ws.title):
            continue
        if ws.max_column > _MAX_COLS or ws.max_row > 5000:
            continue

        fields = []
        for r in range(1, min(ws.max_row, _MAX_ROWS) + 1):
            row_vals = []
            for c in range(1, min(ws.max_column, _MAX_COLS) + 1):
                cell = ws.cell(row=r, column=c)
                v = cell.value
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    row_vals.append((c, cell))
            if not row_vals:
                continue
            label = _row_label(ws, r, row_vals[0][0])
            if not label:
                continue
            multi = len(row_vals) > 1
            for c, cell in row_vals:
                if field_count >= _MAX_FIELDS:
                    break
                flabel = label
                if multi:
                    hdr = _column_header(ws, r, c)
                    flabel = f"{label} — {hdr}" if hdr else f"{label} ({get_column_letter(c)})"
                fields.append({
                    "cell": f"{get_column_letter(c)}{r}",
                    "label": flabel[:70],
                    "default": cell.value,
                    "type": _infer_type(cell),
                })
                field_count += 1
        if fields:
            groups.append({"title": ws.title, "sheet": ws.title, "fields": fields})
        if field_count >= _MAX_FIELDS:
            break

    wb.close()
    return {
        "template_id": template_id,
        "file": os.path.basename(path),
        "currency": currency,
        "truncated": field_count >= _MAX_FIELDS,
        "groups": groups,
    }


def load_schema(purpose_key: str, template_id: str):
    """Load the curated schema JSON for a template, or None if not generated."""
    from template_config import schema_path
    p = schema_path(purpose_key, template_id)
    if not os.path.isfile(p):
        return None
    with open(p, "r", encoding="utf-8") as fh:
        return json.load(fh)


def _regen_all():
    """CLI: regenerate raw schema JSON for every template_fill template."""
    from template_config import TEMPLATES, template_path, schema_path

    for purpose_key, templates in TEMPLATES.items():
        for t in templates:
            if t.get("engine") != "template_fill":
                continue
            path = template_path(purpose_key, t["id"])
            if not path:
                print(f"  SKIP {purpose_key}/{t['id']}: file missing")
                continue
            schema = extract_schema(path, t["id"], t.get("currency", ""))
            out = schema_path(purpose_key, t["id"])
            with open(out, "w", encoding="utf-8") as fh:
                json.dump(schema, fh, indent=2, ensure_ascii=False)
            nfields = sum(len(g["fields"]) for g in schema["groups"])
            flag = " (TRUNCATED)" if schema["truncated"] else ""
            print(f"  {purpose_key}/{t['id']}: {len(schema['groups'])} groups, {nfields} fields{flag} -> {os.path.basename(out)}")


if __name__ == "__main__":
    _regen_all()
