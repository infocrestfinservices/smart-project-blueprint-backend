"""
sample_blueprint_service.py

Each purpose can ship a *sample* .xlsx report under backend/samples/. That sample
is the BLUEPRINT for the purpose: it defines the sheets, column names, layout,
number formats, calculation logic (formulas) and charts we want the generated
report to mirror.

We never copy the sample's data. We extract its STRUCTURE as a compact text
description and feed that into the model prompt; the model then rebuilds the same
structure from the user's own data.

Mapping is by canonical purpose key (see purpose_config). Drop a file named after
the key (e.g. feasibility_study.xlsx) into backend/samples/ and it is picked up
automatically. No sample -> the pipeline falls back to its config-driven default.
"""

import os

from openpyxl import load_workbook

from purpose_config import FEASIBILITY, CMA, IRR, GENERIC

SAMPLES_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "samples"
)

# purpose key -> sample filename
_SAMPLE_FILES = {
    FEASIBILITY: "feasibility_study.xlsx",
    CMA: "cma_data.xlsx",
    IRR: "irr_analysis.xlsx",
    GENERIC: "generic.xlsx",
}

# Keep the blueprint compact so it doesn't blow up the prompt.
_MAX_ROWS_PER_SHEET = 14
_MAX_COLS = 12
_MAX_CELL_CHARS = 40
_MAX_TOTAL_CHARS = 6000

# purpose_key -> (file mtime, blueprint_text | None). mtime keying means a freshly
# dropped/edited sample is picked up without restarting the server.
_cache: dict = {}


def sample_path(purpose_key: str):
    """Absolute path to the purpose's sample file, or None if it doesn't exist."""
    fname = _SAMPLE_FILES.get(purpose_key)
    if not fname:
        return None
    path = os.path.join(SAMPLES_DIR, fname)
    return path if os.path.isfile(path) else None


def has_sample(purpose_key: str) -> bool:
    return sample_path(purpose_key) is not None


def _cell_text(value) -> str:
    if value is None:
        return ""
    s = str(value).replace("\n", " ").strip()
    return (s[: _MAX_CELL_CHARS - 1] + "…") if len(s) > _MAX_CELL_CHARS else s


def _describe_sheet(ws) -> str:
    lines = [f'--- Sheet: "{ws.title}" (rows={ws.max_row}, cols={ws.max_column}) ---']
    formats = set()
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, _MAX_ROWS_PER_SHEET)):
        cells = []
        row_no = None
        for cell in row[:_MAX_COLS]:
            row_no = cell.row
            cells.append(_cell_text(cell.value))
            if cell.number_format and cell.number_format != "General":
                formats.add(cell.number_format)
        if any(cells):
            lines.append(f"R{row_no}: " + " | ".join(cells))

    if ws.max_row > _MAX_ROWS_PER_SHEET:
        lines.append(f"... (+{ws.max_row - _MAX_ROWS_PER_SHEET} more rows)")
    if formats:
        lines.append("Number formats: " + " ; ".join(sorted(formats)[:6]))
    try:
        charts = getattr(ws, "_charts", None) or []
        if charts:
            types = ", ".join(sorted({c.__class__.__name__ for c in charts}))
            lines.append(f"Charts: {len(charts)} ({types})")
    except Exception:
        pass
    return "\n".join(lines)


def build_blueprint_text(purpose_key: str):
    """Return a compact text description of the purpose's sample workbook, or None.

    Cached per file mtime so each sample is parsed once until it changes.
    """
    path = sample_path(purpose_key)
    if not path:
        return None

    mtime = os.path.getmtime(path)
    cached = _cache.get(purpose_key)
    if cached and cached[0] == mtime:
        return cached[1]

    text = None
    try:
        wb = load_workbook(path, data_only=False, read_only=False)
        parts = [f"Workbook with {len(wb.sheetnames)} sheet(s): {wb.sheetnames}"]
        for ws in wb.worksheets:
            parts.append(_describe_sheet(ws))
        wb.close()
        text = "\n\n".join(parts)
        if len(text) > _MAX_TOTAL_CHARS:
            text = text[:_MAX_TOTAL_CHARS] + "\n... (blueprint truncated)"
    except Exception:
        # A malformed sample must never abort generation — fall back to config.
        text = None

    _cache[purpose_key] = (mtime, text)
    return text
