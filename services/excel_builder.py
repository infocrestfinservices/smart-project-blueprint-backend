"""
excel_builder.py

Two responsibilities live here:

1. build_excel(model, purpose_key, project) -> bytes  [EXISTING, UNCHANGED]
   Turns the structured financial-model JSON into a professional, CA-style .xlsx
   workbook (openpyxl). Sheets are driven entirely by the AI output; a final
   Dashboard sheet renders native Excel charts defined in purpose_config. Returns the
   workbook as bytes. Used by routers/generation_router.

2. build_excel_workbook(assumptions, template_path, output_path, template_definition)
   -> dict  [ORCHESTRATION LAYER]
   The final orchestration layer: builds one complete Excel workbook from a business
   assumptions dict by passing outputs from one existing module into the next. It does
   NO financial calculation, contains NO business logic, and uses openpyxl NOT at all
   (all Excel work happens inside the write step's module). Its pipeline module imports
   are done INSIDE the function so that importing this module for build_excel (the app
   startup path via generation_router) has exactly the same import footprint as before.
"""

import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

from purpose_config import get_config

NAVY = "10254A"
NAVY_LT = "1F3A6B"
GOLD = "B08D3F"
IVORY = "F3EEE2"
WHITE = "FFFFFF"
GREY = "6B7280"

THIN = Side(style="thin", color="D9D2C2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _is_num(v):
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        try:
            float(v.replace(",", "").strip())
            return True
        except (ValueError, AttributeError):
            return False
    return False


def _num(v):
    if isinstance(v, (int, float)):
        return v
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, AttributeError):
        return v


def _safe_sheet_name(name, used):
    clean = "".join(c for c in str(name) if c not in '[]:*?/\\')[:31] or "Sheet"
    base, i = clean, 1
    while clean in used:
        suffix = f" ({i})"
        clean = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(clean)
    return clean


def _style_title(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, ncols))
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26


def _build_sheet(wb, sheet, used_names):
    name = _safe_sheet_name(sheet.get("name", "Sheet"), used_names)
    ws = wb.create_sheet(title=name)
    columns = sheet.get("columns") or []
    rows = sheet.get("rows") or []
    ncols = max(len(columns), max((len(r) for r in rows), default=0), 1)

    _style_title(ws, sheet.get("name", name), ncols)

    header_row = 2
    for j in range(ncols):
        c = ws.cell(row=header_row, column=j + 1, value=(columns[j] if j < len(columns) else ""))
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=NAVY_LT)
        c.alignment = Alignment(horizontal="center" if j else "left", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[header_row].height = 22

    first_data_row = header_row + 1
    r = first_data_row
    existing_total = False
    for row in rows:
        label = str(row[0]).lower() if row else ""
        is_total = any(k in label for k in ("total", "net profit", "pat", "grand"))
        if "total" in label:
            existing_total = True
        for j in range(ncols):
            val = row[j] if j < len(row) else ""
            cell = ws.cell(row=r, column=j + 1)
            if j > 0 and _is_num(val):
                cell.value = _num(val)
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.value = val
                cell.alignment = Alignment(horizontal="left" if j == 0 else "right", wrap_text=(j == 0))
            cell.border = BORDER
            cell.font = Font(bold=is_total, size=10, color=NAVY if j == 0 else "000000")
            if is_total:
                cell.fill = PatternFill("solid", fgColor=IVORY)
        ws.row_dimensions[r].height = 18
        r += 1
    last_data_row = r - 1

    # Auto total row with SUM formulas when requested and not already present.
    if sheet.get("total_row") and not existing_total and last_data_row >= first_data_row:
        ws.cell(row=r, column=1, value="Total").font = Font(bold=True, color=NAVY)
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=IVORY)
        for j in range(1, ncols):
            col_letter = get_column_letter(j + 1)
            # only sum columns that have numeric data
            has_num = any(_is_num((rw[j] if j < len(rw) else "")) for rw in rows)
            cell = ws.cell(row=r, column=j + 1)
            if has_num:
                cell.value = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
                cell.number_format = "#,##0"
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=IVORY)
            cell.alignment = Alignment(horizontal="right")
            cell.border = BORDER

    # Column widths
    for j in range(ncols):
        col_letter = get_column_letter(j + 1)
        max_len = len(str(columns[j])) if j < len(columns) else 8
        for row in rows:
            if j < len(row):
                max_len = max(max_len, len(str(row[j])))
        ws.column_dimensions[col_letter].width = min(46, max(11, (max_len + 2) * (1.25 if j == 0 else 1.0)))

    ws.sheet_view.showGridLines = False
    return {
        "ws": ws, "name": name, "header_row": header_row,
        "first_data_row": first_data_row, "last_data_row": last_data_row,
        "col_index": {str(columns[j]).strip(): j + 1 for j in range(len(columns))},
    }


def _add_chart(dash, meta, spec, anchor):
    ws = meta["ws"]
    xcol = meta["col_index"].get(spec["x"])
    series_cols = [meta["col_index"].get(s) for s in spec["series"]]
    series_cols = [c for c in series_cols if c]
    if not series_cols or meta["last_data_row"] < meta["first_data_row"]:
        return False

    ctype = spec.get("type", "bar")
    if ctype == "pie":
        chart = PieChart()
    elif ctype == "line":
        chart = LineChart()
    else:
        chart = BarChart()
        chart.type = "col"
    chart.title = spec["title"]
    chart.height = 7.5
    chart.width = 14
    chart.style = 10

    if ctype == "pie":
        data = Reference(ws, min_col=series_cols[0], min_row=meta["first_data_row"], max_row=meta["last_data_row"])
        chart.add_data(data, titles_from_data=False)
    else:
        for col in series_cols:
            data = Reference(ws, min_col=col, min_row=meta["header_row"], max_row=meta["last_data_row"])
            chart.add_data(data, titles_from_data=True)
    if xcol:
        cats = Reference(ws, min_col=xcol, min_row=meta["first_data_row"], max_row=meta["last_data_row"])
        chart.set_categories(cats)

    dash.add_chart(chart, anchor)
    return True


def build_excel(model: dict, purpose_key: str, project: dict) -> bytes:
    config = get_config(purpose_key)
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet
    used = set()

    # Dashboard first (KPIs at top), charts added after data sheets exist.
    dash = wb.create_sheet(title=_safe_sheet_name("Dashboard", used))
    dash.sheet_view.showGridLines = False
    dash.merge_cells("A1:H1")
    t = dash["A1"]
    t.value = f"{config['label']} — {project.get('title') or 'Financial Model'}"
    t.font = Font(size=16, bold=True, color=WHITE)
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    dash.row_dimensions[1].height = 32

    kpis = model.get("kpis") or []
    col = 1
    for kpi in kpis[:6]:
        lc = dash.cell(row=3, column=col, value=str(kpi.get("label", "")))
        lc.font = Font(size=9, bold=True, color=GREY)
        vc = dash.cell(row=4, column=col, value=str(kpi.get("value", "")))
        vc.font = Font(size=15, bold=True, color=NAVY)
        dash.column_dimensions[get_column_letter(col)].width = 18
        col += 1

    # Data sheets
    metas = {}
    for sheet in model.get("sheets", []):
        meta = _build_sheet(wb, sheet, used)
        metas[meta["name"]] = meta
        # also index by the AI's original requested name
        metas[str(sheet.get("name"))] = meta

    # Charts on the dashboard
    anchors = ["A7", "J7", "A24", "J24", "A41", "J41"]
    ai = 0
    for spec in config.get("charts", []):
        meta = metas.get(spec["sheet"])
        if meta and ai < len(anchors):
            if _add_chart(dash, meta, spec, anchors[ai]):
                ai += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── ORCHESTRATION LAYER ─────────────────────────────────────────────────────────
def build_excel_workbook(
    assumptions: dict,
    template_path: str,
    output_path: str,
    template_definition: dict,
) -> dict:
    """Build one complete Excel workbook from a business assumptions dict by
    orchestrating existing modules only. Performs no calculation, no business logic,
    and touches openpyxl only inside the write step's module.

    Pipeline (outputs pass straight from one stage to the next):
        1. run_financial_engine(assumptions)
        2. build_financial_model(engine_output)
        3. build_excel_mapping(financial_model)                 -> worksheet_mapping
        4. build_template_cell_mapping(worksheet_mapping, template_definition)
        5. write_excel_mapping(template_path, output_path, ...)

    Returns {status, output_path, validation, metadata}. Raises ValueError for
    malformed input (non-dict assumptions, missing workbook, non-dict template
    definition); downstream module errors propagate unchanged.
    """
    fn = "build_excel_workbook"

    # -- input validation (only these three; downstream validates the rest) --
    if not isinstance(assumptions, dict):
        raise ValueError(f"{fn}: assumptions must be a dict, got {type(assumptions).__name__}")
    if not isinstance(template_path, str) or not os.path.isfile(template_path):
        raise ValueError(f"{fn}: workbook does not exist: {template_path!r}")
    if not isinstance(template_definition, dict):
        raise ValueError(f"{fn}: template_definition must be a dict, "
                         f"got {type(template_definition).__name__}")

    # Pipeline module imports are done here (lazily) so that importing this module for
    # build_excel keeps exactly its original, lighter import footprint.
    from financial_engine.engine.financial_engine_runner import run_financial_engine
    from financial_engine.models.financial_model_builder import build_financial_model
    from excel_writer.financial_model_mapper import build_excel_mapping
    from excel_writer.template_cell_mapper import build_template_cell_mapping
    from excel_writer.excel_writer import write_excel_mapping

    # -- pipeline (no calculation, no cell mapping here; just pass outputs along) --
    engine_output = run_financial_engine(assumptions)
    financial_model = build_financial_model(engine_output)
    worksheet_mapping = build_excel_mapping(financial_model)
    resolved_cell_mapping = build_template_cell_mapping(worksheet_mapping, template_definition)

    # write_excel_mapping resolves field->cell from (worksheet_mapping, cell_mapping);
    # step 4 already produced {sheet:{cell:value}}, so feed it through with a
    # cell-identity map (each cell is both the value-key and its own A1 target).
    identity_cell_map = {
        sheet: {cell: cell for cell in cells}
        for sheet, cells in resolved_cell_mapping.items()
    }
    write_excel_mapping(
        workbook_path=template_path,
        output_path=output_path,
        worksheet_mapping=resolved_cell_mapping,
        cell_mapping=identity_cell_map,
    )

    return {
        "status": "success",
        "output_path": output_path,
        "validation": financial_model["validation"],
        "metadata": financial_model["metadata"],
    }
