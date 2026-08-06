"""
financial_summary_service.py

Reads the ONE set of numbers the whole report must agree on straight out of the
recalculated workbook, so the Word report can never quote a figure the Excel model
contradicts.

Template-AGNOSTIC by design. The CMA workbooks carry Annual_Summary / DSCR / Ratios,
but the grant, VC and angel workbooks have entirely different sheets ("Annual P&L",
"Returns & Ratios", "Socio-Economic Impact" …). Hard-coding sheet+row therefore
produced an EMPTY summary for those templates — which is why their Word reports came
out with no charts and no tables. So instead of fixed addresses this scans the
workbook for LABELLED rows ("Net Sales", "EBITDA", "Profit after tax", "DSCR" …) and
reads the annual figures beside them.

Returns 5-year series + headline cards; anything not found is simply absent.
"""

from __future__ import annotations

import logging
import re
from io import BytesIO

from openpyxl import load_workbook

logger = logging.getLogger("financial_summary")

YEARS = 5
# The CMA statements put Year 1..5 in columns C..G.
_YEAR_COLS = ["C", "D", "E", "F", "G"]

# concept -> ordered label patterns (first match wins, so put the precise ones first)
_CONCEPTS = {
    "Net Sales / Revenue": [r"net\s*sales", r"\brevenue\b", r"total\s+revenue", r"turnover",
                            r"gross\s+merchandise", r"\bsales\b"],
    # "\bebitda\b" alone also matches "EBITDA margin" — which is a RATIO row, and
    # matching it first put a percentage where the money series belongs.
    "EBITDA": [r"\bebitda\b(?!\s*margin)"],
    "EBIT (Operating Profit)": [r"\bebit\b(?!da)", r"operating\s+profit"],
    "Profit Before Tax": [r"profit\s+before\s+tax", r"\bpbt\b", r"pre-?tax\s+profit"],
    "Profit After Tax": [r"profit\s+after\s+tax", r"\bpat\b", r"net\s+profit(?!\s*margin)",
                         r"net\s+income"],
    "Cash Accrual": [r"cash\s+accrual", r"net\s+cash\s+flow", r"cash\s+profit"],
}

_RATIO_CONCEPTS = {
    "DSCR": [r"\bdscr\b", r"debt\s+service\s+coverage"],
    "Current Ratio": [r"current\s+ratio"],
    "Debt–Equity Ratio": [r"debt[\s\-–]*equity", r"\bd\s*/\s*e\b"],
    "EBITDA Margin": [r"ebitda\s+margin"],
    "Net Profit Margin": [r"net\s+profit\s+margin", r"net\s+margin", r"pat\s+margin"],
}

# sheets that hold ANNUAL figures; monthly/60-column sheets are skipped so we never
# mistake twelve monthly cells for five yearly ones.
# "Dashboard" is skipped too: it holds scattered chart-helper cells, not clean annual
# series, and matched stray rows ahead of the real statements.
_SKIP_SHEET = re.compile(r"month|assumption|cover|source|instruction|index|dashboard", re.I)


def _num(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip().lower()


def _row_label(ws, row, max_col=4):
    """The first text cell on the row (labels sit in column A or B)."""
    for c in range(1, max_col + 1):
        v = ws.cell(row=row, column=c).value
        if isinstance(v, str) and len(v.strip()) > 1 and not v.startswith("="):
            return v
    return ""


def _row_numbers(ws, row, start_col, limit=YEARS):
    """Up to `limit` annual numbers to the right of the label."""
    out = []
    for c in range(start_col, min(ws.max_column, start_col + 40) + 1):
        n = _num(ws.cell(row=row, column=c).value)
        if n is not None:
            out.append(n)
            if len(out) >= limit:
                break
    return out


def _scan(wb, concepts):
    """{concept: [5 annual values]} found by label anywhere in the workbook."""
    found = {}
    for ws in wb.worksheets:
        if _SKIP_SHEET.search(ws.title) or ws.max_row > 400:
            continue
        # a 60-month sheet has far more numeric columns than an annual one
        if ws.max_column > 24:
            continue
        for row in range(1, min(ws.max_row, 200) + 1):
            label = _norm(_row_label(ws, row))
            if not label:
                continue
            for concept, patterns in concepts.items():
                if concept in found:
                    continue
                if any(re.search(p, label) for p in patterns):
                    vals = _row_numbers(ws, row, 2)
                    # A real series, not a stray cell, and not a mostly-zero row. A
                    # grant-funded project has no term loan, so its DSCR row is zeros
                    # (bar a stray first year); surfacing "0.07×" reads as broken
                    # rather than "not applicable", so such rows are left out.
                    nonzero = [v for v in vals if v not in (0, None)]
                    if len(vals) >= 2 and len(nonzero) * 2 > len(vals):
                        found[concept] = (vals + [None] * YEARS)[:YEARS]
    return found


def _first(series):
    return next((v for v in series if v is not None), None)


def _last(series):
    return next((v for v in reversed(series) if v is not None), None)


def _avg(series):
    vals = [v for v in series if v is not None]
    return sum(vals) / len(vals) if vals else None


# The statutory / analytical statements the workbook already computes. Rendering them
# in the Word report is what makes it a full appraisal document rather than a summary,
# and every figure is read from the recalculated workbook so the two always agree.
_STATEMENT_SHEETS = [
    ("Annual_Summary", "Consolidated Five-Year Summary",
     "Production, sales, cost structure and profitability rolled up by year"),
    ("Form_II_Operating", "Operating Statement (Form II)",
     "Projected profitability in the format required for credit appraisal"),
    ("Form_III_BalanceSheet", "Projected Balance Sheet (Form III)",
     "Assets and liabilities at the close of each year"),
    ("Form_IV_CA_CL", "Current Assets & Current Liabilities (Form IV)",
     "The working-capital position year by year"),
    ("Form_V_MPBF", "Maximum Permissible Bank Finance (Form V)",
     "Working-capital gap and the limit the model supports"),
    ("Form_VI_FundFlow", "Fund Flow Statement (Form VI)",
     "Sources and uses of funds through the projection"),
    ("Repayment", "Term Loan Repayment Schedule",
     "Opening balance, interest, principal and closing balance"),
    ("DSCR", "Debt Service Coverage",
     "Cash available for debt service against the obligation"),
    ("Depreciation", "Depreciation Schedule", "Block-wise depreciation"),
    ("Ratios", "Financial Viability Ratios",
     "Liquidity, leverage, profitability and coverage, year by year"),
]


def extract_wc_seed(recalc_bytes: bytes) -> dict:
    """Year-1 current assets & liabilities from Form IV, returned as "WC & CC-OD Limit"
    input-cell values. Seeds the standalone MPBF calculator so it opens showing THIS
    project's real position (identical to Form V) while its cells stay plain blue inputs
    the user can overwrite with their own actual CA/CL figures."""
    try:
        wb = load_workbook(BytesIO(recalc_bytes), data_only=True)
    except Exception:
        return {}
    if "Form_IV_CA_CL" not in wb.sheetnames:
        return {}
    f = wb["Form_IV_CA_CL"]

    def _n(cell):
        v = f[cell].value
        return round(v) if isinstance(v, (int, float)) else 0

    return {
        "WC & CC-OD Limit!C5": _n("C10") + _n("C11"),   # inventory (RM + FG)
        "WC & CC-OD Limit!C6": _n("C12"),               # debtors / receivables
        "WC & CC-OD Limit!C8": _n("C13"),               # cash & bank
        "WC & CC-OD Limit!C12": _n("C16"),              # creditors / payables
        "WC & CC-OD Limit!C13": _n("C17"),              # other current liabilities
    }


def extract_key_assumptions(recalc_bytes: bytes) -> list:
    """The input assumptions the whole model is built on, as (label, value) pairs —
    a CA reads these first to judge whether the projections are reasonable."""
    try:
        wb = load_workbook(BytesIO(recalc_bytes), data_only=True)
    except Exception:
        return []
    if "Assumptions" not in wb.sheetnames:
        return []
    ws = wb["Assumptions"]
    out = []
    for row in range(1, min(ws.max_row, 70) + 1):
        label = ws.cell(row=row, column=2).value
        if not isinstance(label, str) or len(label.strip()) < 2 or label.startswith("="):
            continue
        label = label.strip()
        val = ws.cell(row=row, column=3).value
        if val is None or (isinstance(val, str) and not val.strip()):
            # section headers (A. FINANCING, B. CAPACITY …) keep the structure readable
            if label[:2].rstrip(".").isalpha() and label.isupper() or label.endswith(":"):
                out.append((label, None))
            continue
        fmt = (ws.cell(row=row, column=3).number_format or "")
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            if "%" in fmt:
                shown = f"{val * 100:.2f}%"
            elif abs(val) >= 1000:
                shown = f"{val:,.0f}"
            else:
                shown = f"{val:,.2f}".rstrip("0").rstrip(".")
        else:
            shown = str(val)
        out.append((label, shown))
    # drop a trailing header with nothing under it
    while out and out[-1][1] is None:
        out.pop()
    return out


# A sheet's own column-header row, which is not part of the data.
_HEADER_LABEL = re.compile(r"^(particulars?|ratio|indicator|item|description|head)s?$", re.I)
_YEAR_HEADER = re.compile(r"^(yr|year)\s*\d", re.I)


def _sheet_table(ws, max_rows=60):
    """A sheet's labelled rows × five year columns, as display rows.

    Returns [(label, [v1..v5], is_heading)]. A row whose label has no numbers beside it
    is treated as a sub-heading (the CMA sheets use those to group lines)."""
    out = []
    for row in range(1, min(ws.max_row, max_rows) + 1):
        label = None
        for col in (2, 1):
            v = ws.cell(row=row, column=col).value
            if isinstance(v, str) and len(v.strip()) > 1 and not v.startswith("="):
                label = v.strip()
                break
        if not label:
            continue
        # The Repayment sheet carries a month-by-month schedule below its annual table.
        # That section's columns are MONTHS, so reading on would pull M2..M6 into the
        # report as if they were Year 1..Year 5. The banner marks where years stop.
        if label.upper().startswith("MONTH-BY-MONTH"):
            break
        vals = [_num(ws[f"{c}{row}"].value) for c in _YEAR_COLS]
        if all(v is None for v in vals):
            # skip the sheet's own title/subtitle rows, keep genuine group headings
            if row <= 3 or len(label) > 70:
                continue
            # ...and skip the sheet's COLUMN-header row ("Particulars | Year 1 | …"),
            # which otherwise came through as an empty group heading above the data.
            if _HEADER_LABEL.match(label):
                continue
            if any(isinstance(ws[f"{c}{row}"].value, str) and
                   _YEAR_HEADER.match(str(ws[f"{c}{row}"].value).strip())
                   for c in _YEAR_COLS):
                continue
            out.append((label, [None] * 5, True))
        else:
            out.append((label, vals, False))
    # drop a trailing heading with nothing under it
    while out and out[-1][2]:
        out.pop()
    return out


def extract_statement_tables(recalc_bytes: bytes) -> list:
    """[{key, title, subtitle, rows}] for each statutory statement present."""
    try:
        wb = load_workbook(BytesIO(recalc_bytes), data_only=True)
    except Exception:
        return []
    tables = []
    for sheet, title, subtitle in _STATEMENT_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        rows = _sheet_table(wb[sheet])
        if len([r for r in rows if not r[2]]) >= 2:      # at least two real data rows
            tables.append({"key": sheet, "title": title, "subtitle": subtitle, "rows": rows})
    return tables


def extract_market_segments(recalc_bytes: bytes) -> list:
    """The target-market revenue split read out of the recalculated workbook, so the
    Word report shows the same segments and the same rupees the workbook does.

    Two layouts, because the two template families are shaped differently:

    * Manufacturing (the frozen Bank Loan workbook) keeps its year-STACKED "Sales"
      sheet — one 12-row block per year, the five segment rows six rows in, the year
      total in column N.
    * The industry templates now carry everything on a single "Revenue Build-Up" sheet
      whose years run ACROSS: the segment rows are fixed at 29-33 and each year's total
      is 13 columns further right (N, AA, AN, BA, BN).

    The shares live on Assumptions D58:D62 in both.
    """
    try:
        wb = load_workbook(BytesIO(recalc_bytes), data_only=True)
    except Exception:
        return []
    a = wb["Assumptions"] if "Assumptions" in wb.sheetnames else None

    if "Revenue Build-Up" in wb.sheetnames:
        s = wb["Revenue Build-Up"]
        SEG_ROW0, STRIDE, TOTAL_COL = 29, 13, 14
        rows = [(SEG_ROW0 + i, TOTAL_COL, TOTAL_COL + 4 * STRIDE) for i in range(5)]
    elif "Sales" in wb.sheetnames:
        s = wb["Sales"]
        SALES_BLOCK, SEG_OFFSET, TOTAL_COL = 12, 6, 14
        y1, y5 = 4 + SEG_OFFSET, 4 + 4 * SALES_BLOCK + SEG_OFFSET
        rows = [(y1 + i, TOTAL_COL, TOTAL_COL) for i in range(5)]
    else:
        return []

    out = []
    for i, (row, c1, c5) in enumerate(rows):
        name = s.cell(row=row, column=1).value
        if not isinstance(name, str) or not name.strip():
            continue
        r5 = row if "Revenue Build-Up" in wb.sheetnames else row + 4 * 12
        v1 = _num(s.cell(row=row, column=c1).value)
        v5 = _num(s.cell(row=r5, column=c5).value)
        share = _num(a.cell(row=58 + i, column=4).value) if a else None
        if v1 is None and v5 is None and share is None:
            continue
        out.append({"name": name.strip(), "share": share, "y1": v1, "y5": v5})
    return out


def extract_financial_summary(recalc_bytes: bytes) -> dict:
    """Pull the shared 5-year figures out of ANY recalculated workbook."""
    try:
        wb = load_workbook(BytesIO(recalc_bytes), data_only=True)
    except Exception:
        logger.warning("financial_summary: could not open recalculated workbook", exc_info=True)
        return {}

    series = _scan(wb, _CONCEPTS)
    ratios = _scan(wb, _RATIO_CONCEPTS)

    if not series and not ratios:
        logger.warning("financial_summary: no labelled figures found in %s", wb.sheetnames)
        return {}

    revenue = series.get("Net Sales / Revenue", [None] * YEARS)
    ebitda = series.get("EBITDA", [None] * YEARS)
    pat = series.get("Profit After Tax", [None] * YEARS)
    dscr = ratios.get("DSCR", [None] * YEARS)

    return {
        "years": [f"Year {i + 1}" for i in range(YEARS)],
        "series": {k: v for k, v in series.items()},
        "ratios": {k: v for k, v in ratios.items()},
        "cards": {
            "revenue_y1": _first(revenue),
            "revenue_y5": _last(revenue),
            "ebitda_y5": _last(ebitda),
            "pat_y5": _last(pat),
            "avg_dscr": _avg(dscr),
            "net_margin_y5": _last(ratios.get("Net Profit Margin", [None] * YEARS)),
        },
    }
