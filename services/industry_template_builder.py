"""
industry_template_builder.py

Generates an industry-specific CMA workbook from the base template + an industry
OPERATING MODEL, so every service/trade industry gets a workbook whose CALCULATIONS
(not just labels) match how that industry actually works — without hand-building a
file per industry.

The base template is manufacturing (capacity → production → per-unit cost). For a
volume-price industry the transformation is mechanical and identical in shape; only
the driver names and the gross margin change, both of which live in the operating
model (financial_engine/industry_calc/operating_models.py):

  * Production sheet  → the industry's VOLUME driver (transactions / covers /
    room-nights / subscribers / patients) instead of "units produced".
  * Sales sheet       → revenue = volume × the industry's PRICE (average bill / AOV /
    ARR / ARPU) instead of unit selling price.
  * Expenses sheet    → COST OF SALES = Net Sales × (1 − gross margin) across all 60
    monthly cells (margin-based), and the second per-unit production input zeroed —
    instead of the manufacturing per-unit raw-material build-up.
  * Assumptions       → the driver cells relabelled to the industry's vocabulary, and
    C25 turned into a gross-margin percentage.

Everything downstream (Profit, Balance Sheet, Ratios, DSCR, Forms II–VI,
Annual_Summary) consumes Net Sales and the expense lines and is reused unchanged.
Manufacturing (and the capacity family) keeps the base template untouched — Bank
Loan is never routed here.

This is the engine-model-driven bridge: one generator, driven by data, produces
every volume-price industry's workbook.
"""

from __future__ import annotations

import os
import shutil

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

BACKEND_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE_TEMPLATE = os.path.join(BACKEND_DIR, "templates", "bank_loan", "CMA_Dashboard_Premium.xlsx")


def _sset(ws, coord, val):
    """Set a cell, resolving a merged range to its top-left anchor (writing to a
    non-anchor merged cell raises)."""
    for rng in ws.merged_cells.ranges:
        if coord in rng:
            ws.cell(row=rng.min_row, column=rng.min_col).value = val
            return
    ws[coord] = val


def build_industry_template(model, out_path: str, base_path: str = BASE_TEMPLATE) -> str:
    """Write a volume-price industry workbook for `model` (an OperatingModel) to
    out_path and return it. Raises ValueError for a capacity-family model (those use
    the base template unchanged)."""
    if getattr(model, "family", "") != "volume_price":
        raise ValueError(f"{model.key}: only volume_price industries get a generated "
                         f"template; capacity industries use the base template.")

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    shutil.copy(base_path, out_path)
    wb = load_workbook(out_path)

    vol = model.volume_label or "Units of activity"
    price = model.price_label or "Average price / unit"
    cost = model.cost_label or "Cost of sales"

    # ── Assumptions: driver semantics (col B = label here) ───────────────────────
    a = wb["Assumptions"]
    for coord, val in [
        ("B15", "B.  SALES VOLUME"),
        ("B16", f"{vol} (Year 1)"),
        ("B17", "Sales Growth"),
        ("B18", "   → growth index Year 1…5 (Year 1 = 1.00)"),
        ("B22", "D.  PRICE & MARGIN  (Year-1 base, escalated yearly)"),
        ("B23", f"{price} (Year 1)"),
        ("B24", "   annual price / value growth"),
        ("B25", f"Gross margin on sales (%)  →  {cost} = Sales × (1 − margin)"),
        ("B26", "   (cost of sales is margin-based, not per-unit)"),
        ("B27", "   (not used in this industry)"),
        ("B29", "Other cost per unit of activity (packaging / fulfilment)"),
        ("B34", "Operating / facility overheads / month (Year 1)"),
    ]:
        _sset(a, coord, val)
    a["C25"].number_format = "0.0%"

    # ── Production sheet: the industry's volume driver (labels in col A) ──────────
    # The two sheet TITLES come from INDUSTRY_MAP (columns 9/10) so they follow the
    # selected industry like every other label — "Production Plan" for a factory,
    # "Purchase / Inventory Plan" for a shop, "Extraction Plan" for a mine.
    # Production and Sales are YEAR-STACKED (a 12-month block per year) and their
    # titles and row labels are INDUSTRY_MAP lookups, so nothing per-industry is
    # hard-coded into them here — only the prose subtitle.
    _sset(wb["Production"], "A2",
          f"Driven by annual {vol.lower()} × growth index × monthly phasing (Assumptions).")
    _sset(wb["Sales"], "A2",
          f"Revenue = {vol.lower()} × {price.lower()} (escalated yearly), split across the "
          f"target-market segments.")

    # ── Expenses sheet: cost of sales = Net Sales × (1 − gross margin) ────────────
    # Expenses is still 60 months across, but Sales is year-stacked, so each of its
    # month columns maps to (month column, the revenue row of that year's block).
    e = wb["Expenses"]
    _sset(e, "A2", "Cost of sales is margin-based; operating costs are period costs.")
    SALES_BLOCK, REV_OFFSET = 12, 4
    OLD_MONTHS = {0: range(2, 14), 1: range(15, 27), 2: range(28, 40),
                  3: range(41, 53), 4: range(54, 66)}
    for c in range(2, e.max_column + 1):
        v = e.cell(row=5, column=c).value
        if not isinstance(v, str) or ("Sales!" not in v and "Production!" not in v):
            continue
        for y, cols in OLD_MONTHS.items():
            if c in cols:
                col = get_column_letter(c)
                rev_row = 4 + y * SALES_BLOCK + REV_OFFSET
                e[f"{col}5"] = (f"=Sales!{get_column_letter(2 + (c - cols.start))}{rev_row}"
                                f"*(1-Assumptions!$C$25)")
                e[f"{col}6"] = 0
                break
    _sset(e, "A5", cost)

    wb.save(out_path)
    return out_path
