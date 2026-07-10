"""
template_analysis.py

Template-DRIVEN financial analysis: given any filled/recalculated workbook, work out
WHERE its headline metrics and consistency checks live — by reading the sheet, not by
hard-coding cell addresses for one specific workbook.

Two mechanisms, in priority order:
  1. Configurable mapping — a sidecar `<id>.analysis.json` (or an "analysis" block in
     the template's schema). If present it is authoritative, so any template can be
     wired precisely by editing one JSON file.
  2. Auto-detection — `analyze_template()` scans labels ("DSCR", "Current Ratio",
     "IRR", "NPV", "ROCE", …) and their value columns, plus balance/"check = 0" cells,
     and produces that mapping automatically. Runs at upload time so a freshly added
     template Just Works.

From the mapping we then, on a recalculated workbook:
  * extract_kpis()  -> the exact computed metric values (for Word/Excel/PDF — one
    source of truth, so every output agrees), and
  * run_checks()    -> Sources=Uses / balance-sheet / "= 0" validations with pass/fail.

Only metrics the template actually contains are reported (no template has every
metric — this one has DSCR + ratios but no IRR/NPV, so those are simply absent).
"""

import io
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from services.recalc_service import _fmt

# label pattern -> (canonical name, value type). Order matters (first match wins).
_METRICS = [
    (re.compile(r"\bdscr\b", re.I), "DSCR", "ratio_x"),
    (re.compile(r"current ratio", re.I), "Current Ratio", "ratio_x"),
    (re.compile(r"quick ratio|acid[- ]?test", re.I), "Quick Ratio", "ratio_x"),
    (re.compile(r"debt[\s–—-]*equity", re.I), "Debt–Equity", "ratio_x"),
    (re.compile(r"\btol\b.*\btnw\b", re.I), "TOL / TNW", "ratio_x"),
    (re.compile(r"interest coverage|interest cover\b", re.I), "Interest Coverage", "ratio_x"),
    (re.compile(r"ebitda margin", re.I), "EBITDA Margin", "percent"),
    (re.compile(r"(net profit|pat|net) margin", re.I), "Net Profit Margin", "percent"),
    (re.compile(r"gross margin", re.I), "Gross Margin", "percent"),
    (re.compile(r"return on capital|roce", re.I), "ROCE", "percent"),
    (re.compile(r"return on equity|\broe\b", re.I), "ROE", "percent"),
    (re.compile(r"return on investment|\broi\b", re.I), "ROI", "percent"),
    (re.compile(r"\birr\b|internal rate", re.I), "IRR", "percent"),
    (re.compile(r"\bnpv\b|net present value", re.I), "NPV", "currency"),
    (re.compile(r"payback", re.I), "Payback", "years"),
    (re.compile(r"break[\s-]?even", re.I), "Break-even", "percent"),
]
_YEAR = re.compile(r"^\s*year\s*\d+\s*$", re.I)
_AVG = re.compile(r"average|avg", re.I)
# a cell that expresses a consistency check (result should be ~0)
_CHECK = re.compile(r"\bcheck\b|difference|must (be|equal)\s*zero|=\s*0\s*\)?|"
                    r"sources\s*[-−–]\s*uses|balance(d)?\b.*zero", re.I)


def _cell_type(metric_type: str, number_format: str) -> str:
    """Refine the value type using the cell's number format where possible."""
    if number_format and "%" in number_format:
        return "percent"
    return metric_type


def _sheet_score(sheet: str) -> int:
    """Prefer headline sheets (dashboard/ratios/returns/summary) over intermediate
    working sheets (monthly model / P&L) when the same metric appears in several."""
    s = (sheet or "").lower()
    for i, kw in enumerate(("dashboard", "investor", "returns", "ratios",
                            "viability", "summary", "dscr")):
        if kw in s:
            return 100 - i
    if "cover" in s or "index" in s:
        return 50
    if any(k in s for k in ("monthly", "model", "p&l", "pnl", "operating", "sheet")):
        return 10
    return 30


def _dedupe_kpis(kpis: list) -> list:
    """Keep one value per canonical metric — the occurrence on the highest-ranked
    sheet — so headline KPIs are clean (one DSCR, one IRR, …)."""
    best = {}
    for k in kpis:
        m = k.get("metric", k["label"])
        if m not in best or _sheet_score(k["sheet"]) > _sheet_score(best[m]["sheet"]):
            best[m] = k
    # preserve first-seen order of the winners
    order, out = [], []
    for k in kpis:
        m = k.get("metric", k["label"])
        if best[m] is k and m not in order:
            order.append(m)
            out.append(k)
    return out


def _header_columns(ws):
    """Locate the value columns of a table: returns (label_col, [year_cols], avg_col)
    from the first row that carries >=2 'Year N' headers, else (None, [], None)."""
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12)):
        year_cols, avg_col = [], None
        for c in row:
            if isinstance(c.value, str):
                if _YEAR.match(c.value):
                    year_cols.append(c.column)
                elif _AVG.search(c.value):
                    avg_col = c.column
        if len(year_cols) >= 2:
            label_col = min(year_cols) - 1
            return label_col, sorted(year_cols), avg_col
    return None, [], None


def _first_value_right(ws, row, start_col):
    """The first non-empty cell to the right of a label (for single-value tables)."""
    for col in range(start_col + 1, min(ws.max_column, start_col + 8) + 1):
        if ws.cell(row=row, column=col).value is not None:
            return ws.cell(row=row, column=col)
    return None


def analyze_template(path: str) -> dict:
    """Auto-detect a template's KPI cells and consistency-check cells. Pure structure
    analysis on the formula workbook — safe to run at upload time."""
    wb = load_workbook(path, data_only=False,
                       keep_vba=path.endswith(".xlsm"), read_only=False)
    kpis, checks, seen = [], [], set()
    try:
        for ws in wb.worksheets:
            label_col, year_cols, avg_col = _header_columns(ws)
            n_years = len(year_cols)
            for row in ws.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    text = cell.value.strip()
                    if not text or text.startswith("="):
                        continue

                    # ---- consistency-check cell? ----
                    if _CHECK.search(text) and "ratio" not in text.lower():
                        vcell = _first_value_right(ws, cell.row, cell.column)
                        if vcell is not None:
                            checks.append({
                                "name": re.sub(r"\s+", " ", text)[:80],
                                "sheet": ws.title, "cell": vcell.coordinate,
                                "tolerance": 1.0, "expect": 0.0,
                            })
                        continue

                    # ---- KPI metric label? (only in the label column of a table) ----
                    if label_col and cell.column == label_col:
                        for pat, canon, mtype in _METRICS:
                            if pat.search(text):
                                if canon == "DSCR" and avg_col is not None:
                                    vcol, tag = avg_col, "Average DSCR"
                                elif year_cols:
                                    vcol, tag = year_cols[-1], f"{canon} (Yr {n_years})"
                                else:
                                    vcol = None
                                    tag = canon
                                if vcol is None:
                                    break
                                coord = f"{get_column_letter(vcol)}{cell.row}"
                                key = (ws.title, coord)
                                if key in seen:
                                    break
                                seen.add(key)
                                vtype = _cell_type(mtype, ws.cell(row=cell.row, column=vcol).number_format)
                                kpis.append({"label": tag, "metric": canon, "sheet": ws.title,
                                             "cell": coord, "type": vtype})
                                break
                    else:
                        # ---- single-value metric (dashboard-style, no year table) ----
                        for pat, canon, mtype in _METRICS:
                            if pat.search(text) and label_col is None:
                                vcell = _first_value_right(ws, cell.row, cell.column)
                                if vcell is not None and isinstance(vcell.value, str) \
                                        and vcell.value.startswith("="):
                                    key = (ws.title, vcell.coordinate)
                                    if key not in seen:
                                        seen.add(key)
                                        kpis.append({"label": canon, "metric": canon,
                                                     "sheet": ws.title, "cell": vcell.coordinate,
                                                     "type": _cell_type(mtype, vcell.number_format)})
                                break
    finally:
        wb.close()
    return {"kpis": _dedupe_kpis(kpis), "checks": checks}


def load_analysis(purpose_key: str, template_id: str) -> dict:
    """The KPI/consistency mapping for a template. Priority: the schema's inline
    "analysis" block (legacy/hand-wired) > the sidecar <id>.analysis.json (auto or
    edited) > empty. Any of these can be edited by hand to wire a template exactly."""
    import json
    import os
    from template_config import analysis_path
    from services.template_introspect import load_schema

    schema = load_schema(purpose_key, template_id) or {}
    if schema.get("analysis"):
        return schema["analysis"]
    p = analysis_path(purpose_key, template_id)
    if p and os.path.isfile(p):
        try:
            with open(p, "r", encoding="utf-8") as fh:
                return json.load(fh)
        except (ValueError, OSError):
            return {}
    return {}


def _read(wb, sheet, cell):
    if sheet not in wb.sheetnames:
        return None
    try:
        return wb[sheet][cell].value
    except Exception:
        return None


def extract_kpis(recalc_bytes: bytes, analysis: dict) -> list:
    """Formatted headline KPIs read from a recalculated workbook, using the mapping.
    Skips any cell that is empty/non-numeric (metric absent for this run)."""
    specs = (analysis or {}).get("kpis") or []
    if not specs:
        return []
    wb = load_workbook(io.BytesIO(recalc_bytes), data_only=True)
    try:
        out = []
        for sp in specs:
            v = _read(wb, sp.get("sheet"), sp.get("cell"))
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                out.append({"label": sp["label"], "value": _fmt(v, sp.get("type", "number"))})
        return out
    finally:
        wb.close()


def run_checks(recalc_bytes: bytes, analysis: dict) -> list:
    """Evaluate the template's consistency checks on a recalculated workbook.
    Returns [{name, value, ok, tolerance}] — ok means |value − expect| <= tolerance."""
    specs = (analysis or {}).get("checks") or []
    if not specs:
        return []
    wb = load_workbook(io.BytesIO(recalc_bytes), data_only=True)
    try:
        out = []
        for sp in specs:
            v = _read(wb, sp.get("sheet"), sp.get("cell"))
            expect = sp.get("expect", 0.0)
            tol = sp.get("tolerance", 1.0)
            ok = isinstance(v, (int, float)) and abs(v - expect) <= tol
            out.append({"name": sp.get("name", "check"), "sheet": sp.get("sheet"),
                        "cell": sp.get("cell"), "value": v, "ok": bool(ok),
                        "tolerance": tol})
        return out
    finally:
        wb.close()
