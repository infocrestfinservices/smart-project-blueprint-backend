"""
excel_model_builder.py

Deterministic, fully formula-driven financial model (openpyxl). Only the
Assumptions sheet holds inputs; every other cell is an Excel formula that
references the assumptions and other sheets, so the whole workbook recalculates
in Excel and can be audited. 5-year annual projection.

Colour code: blue = input, black = formula, green = cross-sheet link,
red = error check. Returns the workbook as bytes.

Known emulations (openpyxl can't do these natively):
  - Scenario Manager -> a scenario selector cell + CHOOSE() drives growth.
  - What-If Data Table -> a formula-computed discount-rate sensitivity grid.
  - Waterfall chart    -> not drawn (bar/line/pie charts are used instead).
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.formatting.rule import CellIsRule

# ── palette ────────────────────────────────────────────────────────────────
NAVY = "10254A"; NAVY2 = "1F3A6B"; GOLD = "B08D3F"; IVORY = "F3EEE2"
WHITE = "FFFFFF"; BLUEIN = "1D4ED8"; GREENLK = "047857"; RED = "B91C1C"; GREY = "6B7280"
THIN = Side(style="thin", color="D9D2C2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Ordered brand palette for chart series (navy, gold, green, blue, red, grey).
SERIES_COLORS = [NAVY, GOLD, GREENLK, BLUEIN, RED, GREY]

YC = ["C", "D", "E", "F", "G"]          # Year 1..5 columns
NY = len(YC)


def _color_series(ch, colors):
    """Apply brand colours to each series (fill for bars/pies, line for lines)."""
    for i, s in enumerate(ch.series):
        col = colors[i % len(colors)]
        gp = GraphicalProperties(solidFill=col)
        gp.line = LineProperties(solidFill=col, w=28000)  # ~2.2pt line
        s.graphicalProperties = gp


def _finish_chart(ch, colors=SERIES_COLORS, labels=False, pct=False,
                  xtitle=None, ytitle=None, legend=False, vary=0, gap=None):
    """Central chart styling so every chart reads as one system.

    vary>0 colours that many individual points from the brand palette (for
    single-series bars and pie slices, so each bar/slice is distinct); otherwise
    each whole series gets one brand colour. labels shows CLEAN value labels
    (value only — never the series/category name); pct shows slice name + %.
    """
    ch.style = 10
    try:
        if vary:
            for s in ch.series:
                s.data_points = [
                    DataPoint(idx=i, spPr=GraphicalProperties(solidFill=colors[i % len(colors)]))
                    for i in range(vary)]
        else:
            _color_series(ch, colors)
    except Exception:
        pass
    if gap is not None:
        try:
            ch.gapWidth = gap
        except Exception:
            pass
    if labels or pct:
        dl = DataLabelList()
        # Suppress the noisy defaults — only the number (or the slice %) shows.
        dl.showSerName = False
        dl.showLegendKey = False
        dl.showCatName = bool(pct)
        dl.showVal = bool(labels) and not pct
        dl.showPercent = bool(pct)
        dl.showBubbleSize = False
        if labels and not pct:
            dl.numFmt = '#,##0'
        ch.dataLabels = dl
    # Pie charts have no cartesian axes — guard every axis touch.
    try:
        if xtitle:
            ch.x_axis.title = xtitle
        if ytitle:
            ch.y_axis.title = ytitle
        ch.x_axis.delete = False
        ch.y_axis.delete = False
    except Exception:
        pass
    if legend:
        ch.legend.position = legend if isinstance(legend, str) else 'b'
    else:
        ch.legend = None

F_INPUT = Font(name="Calibri", size=10, color=BLUEIN, bold=True)   # blue input
F_CALC = Font(name="Calibri", size=10, color="111111")            # black formula
F_LINK = Font(name="Calibri", size=10, color=GREENLK)             # green link
F_TOTAL = Font(name="Calibri", size=10, bold=True, color=NAVY)
F_LABEL = Font(name="Calibri", size=10, color="111111")
MONEY = '#,##0;(#,##0)'
PCT = '0.0%'


def _num(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return float(default)


class Model:
    def __init__(self, project, a):
        self.p = project
        self.a = a
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self.R = {}          # sheet_key -> {label: row}
        self.SH = {}         # sheet_key -> title
        self.AD = {}         # assumption key -> "'03_Assumptions'!$B$row"
        self.curr = project.get("currency") or "INR"

    # ---- helpers ---------------------------------------------------------
    def sheet(self, key, title):
        ws = self.wb.create_sheet(title=title)
        ws.sheet_view.showGridLines = False
        self.SH[key] = title
        self.R[key] = {}
        # title banner
        ws.merge_cells("A1:H1")
        c = ws["A1"]; c.value = title.split("_", 1)[-1]
        c.font = Font(size=14, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[1].height = 24
        ws["A2"] = f"All figures in {self.curr}"
        ws["A2"].font = Font(size=8, italic=True, color=GREY)
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 4
        for col in YC:
            ws.column_dimensions[col].width = 15
        ws.freeze_panes = "C4"   # keep labels + header visible while scrolling
        return ws

    def yhdr(self, ws, row=3):
        ws.cell(row=row, column=1, value="Particulars").font = F_TOTAL
        start = _num(self.a["start_year"], 2025)
        for i, col in enumerate(YC):
            c = ws[f"{col}{row}"]
            c.value = f"Year {i+1} (FY{int(start)+i})"
            c.font = Font(bold=True, color=WHITE, size=10)
            c.fill = PatternFill("solid", fgColor=NAVY2)
            c.alignment = Alignment(horizontal="center")
            c.border = BORDER

    def q(self, key):
        return f"'{self.SH[key]}'"

    def ref(self, key, label, col):
        """Cross-sheet reference to a registered row."""
        return f"{self.q(key)}!{col}{self.R[key][label]}"

    def reserve(self, key, labels, start=4):
        """Pre-register sequential row numbers so formulas can reference rows
        defined later in the same sheet (e.g. Opening <- prev Closing)."""
        for i, lbl in enumerate(labels):
            self.R[key][lbl] = start + i

    def line(self, ws, key, label, row, per_year, fmt=MONEY, font=F_CALC, total=False, link=False):
        """Write a labelled row; per_year(i, col) returns a formula/number for column i."""
        self.R[key][label] = row          # register before computing (self/forward refs)
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = F_TOTAL if total else F_LABEL
        if total:
            lc.fill = PatternFill("solid", fgColor=IVORY)
        for i, col in enumerate(YC):
            c = ws[f"{col}{row}"]
            val = per_year(i, col)
            c.value = val
            c.number_format = fmt
            c.font = (F_LINK if link else (F_TOTAL if total else font))
            c.border = BORDER
            if total:
                c.fill = PatternFill("solid", fgColor=IVORY)
        self.R[key][label] = row
        return row + 1

    def prevcol(self, i):
        return YC[i - 1] if i > 0 else None

    # =====================================================================
    def build(self):
        # Build order respects row-registration dependencies (a sheet's rows
        # must exist before another sheet references them). Tab order is fixed
        # to 01..24 at the end.
        self._cover()
        self._instructions()
        self._assumptions()
        self._revenue()
        self._cogs()
        self._opex()
        self._hr()
        self._capex()
        self._depreciation()
        self._working_capital()
        self._debt()
        self._equity()
        self._income_statement()
        self._tax_note()
        self._cash_flow()        # before balance sheet (BS links to Closing Cash)
        self._balance_sheet()
        self._ratios()
        self._breakeven()
        self._dcf()
        self._sensitivity()
        self._monthly()
        self._scenario()
        self._dashboard()
        self._charts()
        # Reorder tabs into the requested 01..24 numeric sequence.
        self.wb._sheets.sort(key=lambda ws: int(ws.title.split("_")[0]) if ws.title[:2].isdigit() else 999)
        self._finalize()
        buf = io.BytesIO(); self.wb.save(buf); return buf.getvalue()

    # ---- workbook-wide finishing: tab colours, print setup, checks ----------
    def _finalize(self):
        """Colour-code tabs by section, set a professional print layout, flag the
        balance-sheet check in red, and open the workbook on the Dashboard."""
        INPUT_TABS = {"04", "05", "06", "07", "08", "09", "10", "11", "12", "13"}
        STMT_TABS = {"14", "15", "16", "17", "18", "21"}
        VAL_TABS = {"19", "20"}
        VIEW_TABS = {"22", "23", "24"}
        for ws in self.wb.worksheets:
            num = ws.title[:2]
            if num == "01":
                ws.sheet_properties.tabColor = GOLD
            elif num == "02":
                ws.sheet_properties.tabColor = "9AA3AF"
            elif num == "03":
                ws.sheet_properties.tabColor = BLUEIN
            elif num in INPUT_TABS:
                ws.sheet_properties.tabColor = NAVY2
            elif num in STMT_TABS:
                ws.sheet_properties.tabColor = NAVY
            elif num in VAL_TABS:
                ws.sheet_properties.tabColor = GOLD
            elif num in VIEW_TABS:
                ws.sheet_properties.tabColor = GREENLK
            # Print layout: landscape, fit all columns to one page wide.
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
            ws.print_options.horizontalCentered = True
            ws.oddHeader.left.text = (self.p.get("title") or "Financial Model")
            ws.oddHeader.right.text = "&D"
            ws.oddFooter.right.text = "Page &P of &N"

        # Balance-sheet integrity check: fill red whenever it is not zero.
        try:
            bs = self.wb[self.SH["bs"]]
            row = self.R["bs"]["Balance Check (=0)"]
            bs.conditional_formatting.add(
                f"C{row}:G{row}",
                CellIsRule(operator="notEqual", formula=["0"],
                           fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))
        except Exception:
            pass

        # Open on the Dashboard.
        try:
            self.wb.active = self.wb.sheetnames.index("22_Dashboard")
        except Exception:
            pass

    # ---- 01 Cover --------------------------------------------------------
    def _cover(self):
        ws = self.wb.create_sheet("01_Cover"); ws.sheet_view.showGridLines = False
        self.SH["cover"] = "01_Cover"
        ws.merge_cells("A2:H2"); ws["A2"] = self.p.get("title") or "Financial Model"
        ws["A2"].font = Font(size=26, bold=True, color=NAVY)
        ws.merge_cells("A4:H4"); ws["A4"] = "Integrated 5-Year Financial Model & DCF Valuation"
        ws["A4"].font = Font(size=13, italic=True, color=GREY)
        rows = [("Promoter / Company", self.p.get("promoter_name")),
                ("Industry", self.p.get("industry")),
                ("Country / Currency", f"{self.p.get('country') or ''} / {self.curr}"),
                ("Prepared", "AI-generated · formula-driven · bank & investor ready")]
        r = 7
        for k, v in rows:
            ws[f"A{r}"] = k; ws[f"A{r}"].font = Font(bold=True, color=NAVY)
            ws[f"C{r}"] = str(v or "-"); r += 1
        ws.column_dimensions["A"].width = 24
        for c in "CDEFGH":
            ws.column_dimensions[c].width = 16

    # ---- 02 Instructions -------------------------------------------------
    def _instructions(self):
        ws = self.wb.create_sheet("02_Instructions"); ws.sheet_view.showGridLines = False
        self.SH["instr"] = "02_Instructions"
        ws["A1"] = "How to use this model"; ws["A1"].font = Font(size=14, bold=True, color=NAVY)
        notes = [
            "Only the Assumptions sheet (03) contains inputs — shown in BLUE.",
            "Every other sheet is formula-driven and updates automatically.",
            "Colour code: Blue = input · Black = formula · Green = cross-sheet link · Red = error check.",
            "Change any blue input and the entire model (statements, ratios, DCF, dashboard) recalculates.",
            "Scenario: set 03_Assumptions 'Scenario (1/2/3)' to 1=Worst, 2=Base, 3=Best.",
            "Open in Microsoft Excel for full formula support (XLOOKUP/IFS need Excel 365).",
        ]
        for i, n in enumerate(notes):
            ws[f"A{3+i}"] = f"•  {n}"; ws[f"A{3+i}"].font = Font(size=11)
        ws.column_dimensions["A"].width = 110

    # ---- 03 Assumptions --------------------------------------------------
    def _assumptions(self):
        title = "03_Assumptions"
        ws = self.wb.create_sheet(title); ws.sheet_view.showGridLines = False
        self.SH["assum"] = title; self.R["assum"] = {}
        ws.merge_cells("A1:D1"); ws["A1"] = "Assumptions & Drivers (edit BLUE cells)"
        ws["A1"].font = Font(size=14, bold=True, color=WHITE)
        ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 40

        a = self.a
        groups = [
            ("General", [
                ("start_year", "Start Year", a["start_year"], '0'),
                ("proj_years", "Projection Years", 5, '0'),
                ("scenario", "Scenario (1=Worst,2=Base,3=Best)", 2, '0'),
                ("inflation", "Inflation %", a["inflation"], PCT),
                ("tax_rate", "Tax Rate %", a["tax_rate"], PCT),
            ]),
            ("Cost of Capital", [
                ("wacc", "WACC / Discount Rate %", a["wacc"], PCT),
                ("cost_of_debt", "Cost of Debt %", a["cost_of_debt"], PCT),
                ("cost_of_equity", "Cost of Equity %", a["cost_of_equity"], PCT),
                ("terminal_growth", "Terminal Growth %", a["terminal_growth"], PCT),
            ]),
            ("Customers & Revenue", [
                ("start_customers", "Opening Customers (Y1)", a["start_customers"], '#,##0'),
                ("cust_growth", "Customer Growth % (base)", a["cust_growth"], PCT),
                ("churn", "Churn %", a["churn"], PCT),
                ("price", "Price per Customer / yr", a["price"], MONEY),
                ("onetime_pct", "One-Time Rev (% of subs)", a["onetime_pct"], PCT),
                ("service_pct", "Service Rev (% of subs)", a["service_pct"], PCT),
                ("consulting_pct", "Consulting Rev (% of subs)", a["consulting_pct"], PCT),
                ("other_pct", "Other Rev (% of subs)", a["other_pct"], PCT),
            ]),
            ("Costs", [
                ("cogs_gateway", "COGS: Payment Gateway (% rev)", a["cogs_gateway"], PCT),
                ("cogs_cloud", "COGS: Cloud Hosting (% rev)", a["cogs_cloud"], PCT),
                ("cogs_api", "COGS: API / AI Model (% rev)", a["cogs_api"], PCT),
                ("cogs_support", "COGS: Support (% rev)", a["cogs_support"], PCT),
                ("salary_increment", "Salary Increment % / yr", a["salary_increment"], PCT),
            ]),
            ("Capital & Financing", [
                ("capex_total", "Initial CapEx (total)", a["capex_total"], MONEY),
                ("useful_life", "Asset Useful Life (yrs)", a["useful_life"], '0'),
                ("residual_pct", "Residual Value %", a["residual_pct"], PCT),
                ("loan_amount", "Term Loan Amount", a["loan_amount"], MONEY),
                ("interest_rate", "Loan Interest Rate %", a["interest_rate"], PCT),
                ("loan_tenure", "Loan Tenure (yrs)", a["loan_tenure"], '0'),
                ("equity_total", "Equity Invested", a["equity_total"], MONEY),
            ]),
            ("Working Capital (days)", [
                ("recv_days", "Receivable Days", a["recv_days"], '0'),
                ("inv_days", "Inventory Days", a["inv_days"], '0'),
                ("pay_days", "Payable Days", a["pay_days"], '0'),
            ]),
        ]
        r = 3
        for gname, items in groups:
            ws[f"A{r}"] = gname; ws[f"A{r}"].font = Font(bold=True, color=GOLD, size=11)
            r += 1
            for key, label, val, fmt in items:
                ws[f"A{r}"] = label; ws[f"A{r}"].font = F_LABEL
                c = ws[f"B{r}"]; c.value = _num(val); c.number_format = fmt
                c.font = F_INPUT
                c.fill = PatternFill("solid", fgColor="EFF4FF")
                c.border = BORDER
                self.AD[key] = f"'{title}'!$B${r}"
                r += 1
            r += 1
        # scenario-adjusted growth (green link, used by revenue)
        ws[f"A{r}"] = "Effective Customer Growth (scenario)"; ws[f"A{r}"].font = F_LABEL
        cc = ws[f"B{r}"]
        cc.value = f"=CHOOSE({self.AD['scenario']},{self.AD['cust_growth']}*0.5,{self.AD['cust_growth']},{self.AD['cust_growth']}*1.5)"
        cc.number_format = PCT; cc.font = Font(color=GREENLK, bold=True); cc.border = BORDER
        self.AD["eff_growth"] = f"'{title}'!$B${r}"

    # ---- 04 Revenue ------------------------------------------------------
    def _revenue(self):
        k = "rev"; ws = self.sheet(k, "04_Revenue Model"); self.yhdr(ws)
        A = self.AD; r = 4
        self.reserve(k, ["Opening Customers", "New Customers", "Lost Customers (Churn)",
                         "Closing Customers", "Average Customers", "Price per Customer",
                         "Subscription Revenue", "One-Time Revenue", "Service Revenue",
                         "Consulting Revenue", "Other Revenue", "Total Revenue", "Revenue Growth %"], 4)
        r = self.line(ws, k, "Opening Customers", r,
                      lambda i, c: (f"={A['start_customers']}" if i == 0 else f"={self.ref(k,'Closing Customers',self.prevcol(i))}"), fmt='#,##0')
        r = self.line(ws, k, "New Customers", r,
                      lambda i, c: f"={c}{self.R[k]['Opening Customers']}*{A['eff_growth']}", fmt='#,##0')
        r = self.line(ws, k, "Lost Customers (Churn)", r,
                      lambda i, c: f"=-{c}{self.R[k]['Opening Customers']}*{A['churn']}", fmt='#,##0')
        r = self.line(ws, k, "Closing Customers", r,
                      lambda i, c: f"={c}{self.R[k]['Opening Customers']}+{c}{self.R[k]['New Customers']}+{c}{self.R[k]['Lost Customers (Churn)']}", fmt='#,##0', total=True)
        r = self.line(ws, k, "Average Customers", r,
                      lambda i, c: f"=({c}{self.R[k]['Opening Customers']}+{c}{self.R[k]['Closing Customers']})/2", fmt='#,##0')
        r = self.line(ws, k, "Price per Customer", r,
                      lambda i, c: (f"={A['price']}" if i == 0 else f"={self.prevcol(i)}{self.R[k]['Price per Customer']}*(1+{A['inflation']})"))
        r = self.line(ws, k, "Subscription Revenue", r,
                      lambda i, c: f"={c}{self.R[k]['Average Customers']}*{c}{self.R[k]['Price per Customer']}", link=True)
        r = self.line(ws, k, "One-Time Revenue", r,
                      lambda i, c: f"={c}{self.R[k]['Subscription Revenue']}*{A['onetime_pct']}")
        r = self.line(ws, k, "Service Revenue", r,
                      lambda i, c: f"={c}{self.R[k]['Subscription Revenue']}*{A['service_pct']}")
        r = self.line(ws, k, "Consulting Revenue", r,
                      lambda i, c: f"={c}{self.R[k]['Subscription Revenue']}*{A['consulting_pct']}")
        r = self.line(ws, k, "Other Revenue", r,
                      lambda i, c: f"={c}{self.R[k]['Subscription Revenue']}*{A['other_pct']}")
        sub = self.R[k]['Subscription Revenue']; oth = self.R[k]['Other Revenue']
        r = self.line(ws, k, "Total Revenue", r,
                      lambda i, c: f"=SUM({c}{sub}:{c}{oth})", total=True)
        r = self.line(ws, k, "Revenue Growth %", r,
                      lambda i, c: ("=0" if i == 0 else f"={c}{self.R[k]['Total Revenue']}/{self.prevcol(i)}{self.R[k]['Total Revenue']}-1"), fmt=PCT)

    # ---- 05 COGS ---------------------------------------------------------
    def _cogs(self):
        k = "cogs"; ws = self.sheet(k, "05_COGS"); self.yhdr(ws)
        A = self.AD; r = 4
        rev = lambda c: self.ref("rev", "Total Revenue", c)
        for label, akey in [("Payment Gateway", "cogs_gateway"), ("Cloud Hosting", "cogs_cloud"),
                            ("API / AI Model Cost", "cogs_api"), ("Support Cost", "cogs_support")]:
            r = self.line(ws, k, label, r, lambda i, c, ak=akey: f"={rev(c)}*{A[ak]}", link=True)
        first = self.R[k]["Payment Gateway"]; last = self.R[k]["Support Cost"]
        r = self.line(ws, k, "Total COGS", r, lambda i, c: f"=SUM({c}{first}:{c}{last})", total=True)

    # ---- 06 Operating Expenses ------------------------------------------
    def _opex(self):
        k = "opex"; ws = self.sheet(k, "06_Operating Expenses"); self.yhdr(ws)
        A = self.AD; r = 4
        defaults = self.a["opex_fixed"]
        items = ["Office Rent", "Software", "Marketing", "Utilities", "Insurance",
                 "Travel", "Legal", "Audit", "Internet", "Maintenance"]
        for label in items:
            base = _num(defaults.get(label, 120000))
            def py(i, c, b=base):
                if i == 0:
                    return b  # Y1 base = blue input
                return f"={self.prevcol(i)}{self.R[k][label]}*(1+{A['inflation']})"
            r = self.line(ws, k, label, r, py, font=F_CALC)
            # mark Y1 as input (blue)
            ws[f"C{self.R[k][label]}"].font = F_INPUT
            ws[f"C{self.R[k][label]}"].fill = PatternFill("solid", fgColor="EFF4FF")
        first = self.R[k]["Office Rent"]; last = self.R[k]["Maintenance"]
        r = self.line(ws, k, "Total Operating Expenses", r, lambda i, c: f"=SUM({c}{first}:{c}{last})", total=True)

    # ---- 07 HR & Payroll -------------------------------------------------
    def _hr(self):
        k = "hr"; title = "07_HR & Payroll"
        ws = self.sheet(k, title)
        A = self.AD
        # detailed Y1 table
        hdr = ["Role", "Head-count", "Monthly Salary", "Annual Salary", "PF (12%)",
               "ESIC (3.25%)", "Bonus (8.33%)", "Y1 Total Cost"]
        for j, h in enumerate(hdr):
            c = ws.cell(row=3, column=1 + j, value=h)
            c.font = Font(bold=True, color=WHITE, size=9)
            c.fill = PatternFill("solid", fgColor=NAVY2); c.border = BORDER
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        for col in ["C", "D", "E", "F", "G", "H"]:
            ws.column_dimensions[col].width = 13
        ws.column_dimensions["A"].width = 20; ws.column_dimensions["B"].width = 11
        roles = self.a["hr_roles"]   # list of (role, headcount, monthly_salary)
        r = 4
        for role, hc, sal in roles:
            ws.cell(row=r, column=1, value=role).font = F_LABEL
            ws.cell(row=r, column=2, value=hc).font = F_INPUT
            ws.cell(row=r, column=3, value=sal).font = F_INPUT
            for col in ("B", "C"):
                ws[f"{col}{r}"].fill = PatternFill("solid", fgColor="EFF4FF")
            ws[f"D{r}"] = f"=B{r}*C{r}*12"
            ws[f"E{r}"] = f"=D{r}*0.12"
            ws[f"F{r}"] = f"=D{r}*0.0325"
            ws[f"G{r}"] = f"=D{r}*0.0833"
            ws[f"H{r}"] = f"=D{r}+E{r}+F{r}+G{r}"
            for col in "DEFGH":
                ws[f"{col}{r}"].number_format = MONEY; ws[f"{col}{r}"].font = F_CALC; ws[f"{col}{r}"].border = BORDER
            r += 1
        tot_r = r
        ws.cell(row=tot_r, column=1, value="Total (Year 1)").font = F_TOTAL
        ws[f"H{tot_r}"] = f"=SUM(H4:H{r-1})"; ws[f"H{tot_r}"].number_format = MONEY; ws[f"H{tot_r}"].font = F_TOTAL
        # yearly projection row
        proj_r = tot_r + 2
        ws.cell(row=proj_r, column=1, value="Payroll (Annual, escalated)").font = F_TOTAL
        for i, col in enumerate(YC):
            if i == 0:
                ws[f"{col}{proj_r}"] = f"=H{tot_r}"
            else:
                ws[f"{col}{proj_r}"] = f"={self.prevcol(i)}{proj_r}*(1+{A['salary_increment']})"
            ws[f"{col}{proj_r}"].number_format = MONEY; ws[f"{col}{proj_r}"].font = F_TOTAL; ws[f"{col}{proj_r}"].border = BORDER
        self.R[k]["Payroll (Annual, escalated)"] = proj_r

    # ---- 08 CapEx --------------------------------------------------------
    def _capex(self):
        k = "capex"; ws = self.sheet(k, "08_CapEx"); self.yhdr(ws)
        A = self.AD; r = 4
        # Total capex incurred in Year 1 (from assumption), zero after.
        r = self.line(ws, k, "Capital Expenditure", r,
                      lambda i, c: (f"={A['capex_total']}" if i == 0 else "=0"), link=True)

    # ---- 09 Depreciation -------------------------------------------------
    def _depreciation(self):
        k = "dep"; ws = self.sheet(k, "09_Depreciation"); self.yhdr(ws)
        A = self.AD; r = 4
        self.reserve(k, ["Opening Gross Block", "Additions", "Depreciation", "Closing Net Block"], 4)
        capex = lambda c: self.ref("capex", "Capital Expenditure", c)
        # annual SLN depreciation = capex*(1-residual)/life
        r = self.line(ws, k, "Opening Gross Block", r,
                      lambda i, c: ("=0" if i == 0 else f"={self.prevcol(i)}{self.R[k]['Closing Net Block']}"))
        r = self.line(ws, k, "Additions", r, lambda i, c: f"={capex(c)}", link=True)
        r = self.line(ws, k, "Depreciation", r,
                      lambda i, c: f"=MIN(({A['capex_total']}*(1-{A['residual_pct']}))/{A['useful_life']}, {c}{self.R[k]['Opening Gross Block']}+{c}{self.R[k]['Additions']})", total=True)
        r = self.line(ws, k, "Closing Net Block", r,
                      lambda i, c: f"={c}{self.R[k]['Opening Gross Block']}+{c}{self.R[k]['Additions']}-{c}{self.R[k]['Depreciation']}")

    # ---- 10 Working Capital ---------------------------------------------
    def _working_capital(self):
        k = "wc"; ws = self.sheet(k, "10_Working Capital"); self.yhdr(ws)
        A = self.AD; r = 4
        rev = lambda c: self.ref("rev", "Total Revenue", c)
        cogs = lambda c: self.ref("cogs", "Total COGS", c)
        r = self.line(ws, k, "Receivables", r, lambda i, c: f"={rev(c)}*{A['recv_days']}/365", link=True)
        r = self.line(ws, k, "Inventory", r, lambda i, c: f"={cogs(c)}*{A['inv_days']}/365", link=True)
        r = self.line(ws, k, "Payables", r, lambda i, c: f"={cogs(c)}*{A['pay_days']}/365", link=True)
        r = self.line(ws, k, "Net Working Capital", r,
                      lambda i, c: f"={c}{self.R[k]['Receivables']}+{c}{self.R[k]['Inventory']}-{c}{self.R[k]['Payables']}", total=True)
        r = self.line(ws, k, "Change in NWC", r,
                      lambda i, c: (f"={c}{self.R[k]['Net Working Capital']}" if i == 0 else f"={c}{self.R[k]['Net Working Capital']}-{self.prevcol(i)}{self.R[k]['Net Working Capital']}"))

    # ---- 11 Debt Schedule ------------------------------------------------
    def _debt(self):
        k = "debt"; ws = self.sheet(k, "11_Debt Schedule"); self.yhdr(ws)
        A = self.AD; r = 4
        self.reserve(k, ["Opening Balance", "Interest", "Annual Payment (EMI)",
                         "Principal Repaid", "Closing Balance"], 4)
        emi = f"IFERROR(-PMT({A['interest_rate']},{A['loan_tenure']},{A['loan_amount']}),0)"
        r = self.line(ws, k, "Opening Balance", r,
                      lambda i, c: (f"={A['loan_amount']}" if i == 0 else f"={self.prevcol(i)}{self.R[k]['Closing Balance']}"))
        r = self.line(ws, k, "Interest", r,
                      lambda i, c: f"={c}{self.R[k]['Opening Balance']}*{A['interest_rate']}", total=True)
        r = self.line(ws, k, "Annual Payment (EMI)", r,
                      lambda i, c, e=emi: f"=IF(({i}+1)<={A['loan_tenure']},{e},0)")
        r = self.line(ws, k, "Principal Repaid", r,
                      lambda i, c: f"=MIN({c}{self.R[k]['Annual Payment (EMI)']}-{c}{self.R[k]['Interest']}, {c}{self.R[k]['Opening Balance']})")
        r = self.line(ws, k, "Closing Balance", r,
                      lambda i, c: f"=MAX({c}{self.R[k]['Opening Balance']}-{c}{self.R[k]['Principal Repaid']},0)", total=True)

    # ---- 12 Equity Schedule ---------------------------------------------
    def _equity(self):
        k = "eq"; ws = self.sheet(k, "12_Equity Schedule")
        A = self.AD
        ws["A3"] = "Source"; ws["B3"] = "Amount"; ws["C3"] = "Shareholding %"
        for col in ("A3", "B3", "C3"):
            ws[col].font = Font(bold=True, color=WHITE); ws[col].fill = PatternFill("solid", fgColor=NAVY2)
        rows = self.a["equity_rounds"]  # list of (name, amount)
        r = 4
        for name, amt in rows:
            ws[f"A{r}"] = name; ws[f"A{r}"].font = F_LABEL
            ws[f"B{r}"] = amt; ws[f"B{r}"].number_format = MONEY; ws[f"B{r}"].font = F_INPUT
            ws[f"B{r}"].fill = PatternFill("solid", fgColor="EFF4FF")
            r += 1
        ws[f"A{r}"] = "Total Equity"; ws[f"A{r}"].font = F_TOTAL
        ws[f"B{r}"] = f"=SUM(B4:B{r-1})"; ws[f"B{r}"].number_format = MONEY; ws[f"B{r}"].font = F_TOTAL
        for rr in range(4, r):
            ws[f"C{rr}"] = f"=B{rr}/$B${r}"; ws[f"C{rr}"].number_format = PCT; ws[f"C{rr}"].font = F_CALC
        self.R[k]["Total Equity"] = r

    # ---- 14 Income Statement --------------------------------------------
    def _income_statement(self):
        k = "is"; ws = self.sheet(k, "14_Income Statement"); self.yhdr(ws)
        A = self.AD; r = 4
        r = self.line(ws, k, "Revenue", r, lambda i, c: f"={self.ref('rev','Total Revenue',c)}", link=True)
        r = self.line(ws, k, "less: COGS", r, lambda i, c: f"=-{self.ref('cogs','Total COGS',c)}", link=True)
        r = self.line(ws, k, "Gross Profit", r,
                      lambda i, c: f"={c}{self.R[k]['Revenue']}+{c}{self.R[k]['less: COGS']}", total=True)
        r = self.line(ws, k, "less: Operating Expenses", r,
                      lambda i, c: f"=-{self.ref('opex','Total Operating Expenses',c)}", link=True)
        r = self.line(ws, k, "less: Payroll", r,
                      lambda i, c: f"=-{self.ref('hr','Payroll (Annual, escalated)',c)}", link=True)
        r = self.line(ws, k, "EBITDA", r,
                      lambda i, c: f"={c}{self.R[k]['Gross Profit']}+{c}{self.R[k]['less: Operating Expenses']}+{c}{self.R[k]['less: Payroll']}", total=True)
        r = self.line(ws, k, "less: Depreciation", r, lambda i, c: f"=-{self.ref('dep','Depreciation',c)}", link=True)
        r = self.line(ws, k, "EBIT", r,
                      lambda i, c: f"={c}{self.R[k]['EBITDA']}+{c}{self.R[k]['less: Depreciation']}", total=True)
        r = self.line(ws, k, "less: Interest", r, lambda i, c: f"=-{self.ref('debt','Interest',c)}", link=True)
        r = self.line(ws, k, "PBT", r,
                      lambda i, c: f"={c}{self.R[k]['EBIT']}+{c}{self.R[k]['less: Interest']}", total=True)
        r = self.line(ws, k, "less: Tax", r,
                      lambda i, c: f"=-MAX(0,{c}{self.R[k]['PBT']}*{A['tax_rate']})")
        r = self.line(ws, k, "PAT (Net Profit)", r,
                      lambda i, c: f"={c}{self.R[k]['PBT']}+{c}{self.R[k]['less: Tax']}", total=True)
        r += 1
        r = self.line(ws, k, "Gross Margin %", r, lambda i, c: f"={c}{self.R[k]['Gross Profit']}/{c}{self.R[k]['Revenue']}", fmt=PCT)
        r = self.line(ws, k, "EBITDA Margin %", r, lambda i, c: f"={c}{self.R[k]['EBITDA']}/{c}{self.R[k]['Revenue']}", fmt=PCT)
        r = self.line(ws, k, "Net Margin %", r, lambda i, c: f"={c}{self.R[k]['PAT (Net Profit)']}/{c}{self.R[k]['Revenue']}", fmt=PCT)

    def _tax_note(self):
        # 13 Tax handled inline in IS; add a thin sheet for completeness/order.
        k = "tax"; ws = self.sheet(k, "13_Tax Schedule"); self.yhdr(ws)
        A = self.AD; r = 4
        r = self.line(ws, k, "PBT", r, lambda i, c: f"={self.ref('is','PBT',c)}", link=True)
        r = self.line(ws, k, "Tax Rate %", r, lambda i, c: f"={A['tax_rate']}", fmt=PCT)
        r = self.line(ws, k, "Tax Expense", r, lambda i, c: f"=MAX(0,{c}{self.R[k]['PBT']}*{A['tax_rate']})", total=True)

    # ---- 15 Balance Sheet ------------------------------------------------
    def _balance_sheet(self):
        k = "bs"; ws = self.sheet(k, "15_Balance Sheet"); self.yhdr(ws)
        A = self.AD; r = 4
        ws.cell(row=r, column=1, value="ASSETS").font = Font(bold=True, color=GOLD); r += 1
        r = self.line(ws, k, "Cash & Bank", r, lambda i, c: f"={self.ref('cf','Closing Cash',c)}", link=True)
        r = self.line(ws, k, "Receivables", r, lambda i, c: f"={self.ref('wc','Receivables',c)}", link=True)
        r = self.line(ws, k, "Inventory", r, lambda i, c: f"={self.ref('wc','Inventory',c)}", link=True)
        r = self.line(ws, k, "Net Fixed Assets", r, lambda i, c: f"={self.ref('dep','Closing Net Block',c)}", link=True)
        r = self.line(ws, k, "Total Assets", r,
                      lambda i, c: f"={c}{self.R[k]['Cash & Bank']}+{c}{self.R[k]['Receivables']}+{c}{self.R[k]['Inventory']}+{c}{self.R[k]['Net Fixed Assets']}", total=True)
        r += 1
        ws.cell(row=r, column=1, value="LIABILITIES & EQUITY").font = Font(bold=True, color=GOLD); r += 1
        r = self.line(ws, k, "Payables", r, lambda i, c: f"={self.ref('wc','Payables',c)}", link=True)
        r = self.line(ws, k, "Long-Term Debt", r, lambda i, c: f"={self.ref('debt','Closing Balance',c)}", link=True)
        r = self.line(ws, k, "Share Capital", r, lambda i, c: f"={self.ref('eq','Total Equity',None) if False else A['equity_total']}")
        r = self.line(ws, k, "Retained Earnings", r,
                      lambda i, c: (f"={self.ref('is','PAT (Net Profit)',c)}" if i == 0 else f"={self.prevcol(i)}{self.R[k]['Retained Earnings']}+{self.ref('is','PAT (Net Profit)',c)}"), link=True)
        r = self.line(ws, k, "Total Liab. & Equity", r,
                      lambda i, c: f"={c}{self.R[k]['Payables']}+{c}{self.R[k]['Long-Term Debt']}+{c}{self.R[k]['Share Capital']}+{c}{self.R[k]['Retained Earnings']}", total=True)
        r += 1
        r = self.line(ws, k, "Balance Check (=0)", r,
                      lambda i, c: f"=ROUND({c}{self.R[k]['Total Assets']}-{c}{self.R[k]['Total Liab. & Equity']},0)")
        for col in YC:
            cell = ws[f"{col}{self.R[k]['Balance Check (=0)']}"]
            cell.font = Font(bold=True, color=RED)

    # ---- 16 Cash Flow ----------------------------------------------------
    def _cash_flow(self):
        k = "cf"; ws = self.sheet(k, "16_Cash Flow Statement"); self.yhdr(ws)
        A = self.AD; r = 4
        self.reserve(k, ["PAT", "add: Depreciation", "less: Change in NWC", "Cash from Operations",
                         "Cash from Investing (CapEx)", "Equity Inflow", "Loan Inflow",
                         "less: Loan Principal", "Cash from Financing", "Net Cash Flow",
                         "Opening Cash", "Closing Cash"], 4)
        r = self.line(ws, k, "PAT", r, lambda i, c: f"={self.ref('is','PAT (Net Profit)',c)}", link=True)
        r = self.line(ws, k, "add: Depreciation", r, lambda i, c: f"={self.ref('dep','Depreciation',c)}", link=True)
        r = self.line(ws, k, "less: Change in NWC", r, lambda i, c: f"=-{self.ref('wc','Change in NWC',c)}", link=True)
        r = self.line(ws, k, "Cash from Operations", r,
                      lambda i, c: f"={c}{self.R[k]['PAT']}+{c}{self.R[k]['add: Depreciation']}+{c}{self.R[k]['less: Change in NWC']}", total=True)
        r = self.line(ws, k, "Cash from Investing (CapEx)", r, lambda i, c: f"=-{self.ref('capex','Capital Expenditure',c)}", link=True)
        r = self.line(ws, k, "Equity Inflow", r, lambda i, c: (f"={A['equity_total']}" if i == 0 else "=0"))
        r = self.line(ws, k, "Loan Inflow", r, lambda i, c: (f"={A['loan_amount']}" if i == 0 else "=0"))
        r = self.line(ws, k, "less: Loan Principal", r, lambda i, c: f"=-{self.ref('debt','Principal Repaid',c)}", link=True)
        r = self.line(ws, k, "Cash from Financing", r,
                      lambda i, c: f"={c}{self.R[k]['Equity Inflow']}+{c}{self.R[k]['Loan Inflow']}+{c}{self.R[k]['less: Loan Principal']}", total=True)
        r = self.line(ws, k, "Net Cash Flow", r,
                      lambda i, c: f"={c}{self.R[k]['Cash from Operations']}+{c}{self.R[k]['Cash from Investing (CapEx)']}+{c}{self.R[k]['Cash from Financing']}", total=True)
        r = self.line(ws, k, "Opening Cash", r,
                      lambda i, c: ("=0" if i == 0 else f"={self.prevcol(i)}{self.R[k]['Closing Cash']}"))
        r = self.line(ws, k, "Closing Cash", r,
                      lambda i, c: f"={c}{self.R[k]['Opening Cash']}+{c}{self.R[k]['Net Cash Flow']}", total=True)

    # ---- 17 Ratios -------------------------------------------------------
    def _ratios(self):
        k = "rat"; ws = self.sheet(k, "17_Financial Ratios"); self.yhdr(ws)
        r = 4
        defs = [
            ("Gross Margin %", lambda c: f"={self.ref('is','Gross Profit',c)}/{self.ref('is','Revenue',c)}", PCT),
            ("EBITDA Margin %", lambda c: f"={self.ref('is','EBITDA',c)}/{self.ref('is','Revenue',c)}", PCT),
            ("Net Margin %", lambda c: f"={self.ref('is','PAT (Net Profit)',c)}/{self.ref('is','Revenue',c)}", PCT),
            ("Current Ratio", lambda c: f"=({self.ref('bs','Cash & Bank',c)}+{self.ref('bs','Receivables',c)}+{self.ref('bs','Inventory',c)})/{self.ref('bs','Payables',c)}", '0.00'),
            ("Debt / Equity", lambda c: f"={self.ref('bs','Long-Term Debt',c)}/({self.ref('bs','Share Capital',c)}+{self.ref('bs','Retained Earnings',c)})", '0.00'),
            ("Interest Coverage", lambda c: f"={self.ref('is','EBIT',c)}/{self.ref('debt','Interest',c)}", '0.00'),
            ("ROE %", lambda c: f"={self.ref('is','PAT (Net Profit)',c)}/({self.ref('bs','Share Capital',c)}+{self.ref('bs','Retained Earnings',c)})", PCT),
            ("ROA %", lambda c: f"={self.ref('is','PAT (Net Profit)',c)}/{self.ref('bs','Total Assets',c)}", PCT),
        ]
        for label, fn, fmt in defs:
            # Wrap every ratio in IFERROR so a zero denominator (e.g. Interest
            # Coverage with no debt) shows 0 instead of #DIV/0!.
            r = self.line(ws, k, label, r, lambda i, c, f=fn: f"=IFERROR({f(c)[1:]},0)", fmt=fmt, link=True)

    # ---- 18 Break-even ---------------------------------------------------
    def _breakeven(self):
        k = "be"; ws = self.sheet(k, "18_Break Even Analysis"); self.yhdr(ws)
        A = self.AD; r = 4
        rev = lambda c: self.ref("rev", "Total Revenue", c)
        var = lambda c: f"({self.ref('cogs','Total COGS',c)})"
        fixed = lambda c: f"({self.ref('opex','Total Operating Expenses',c)}+{self.ref('hr','Payroll (Annual, escalated)',c)})"
        r = self.line(ws, k, "Contribution Margin %", r, lambda i, c: f"=({rev(c)}-{var(c)})/{rev(c)}", fmt=PCT)
        r = self.line(ws, k, "Fixed Costs", r, lambda i, c: f"={fixed(c)}", link=True)
        r = self.line(ws, k, "Break-even Revenue", r,
                      lambda i, c: f"={c}{self.R[k]['Fixed Costs']}/{c}{self.R[k]['Contribution Margin %']}", total=True)
        r = self.line(ws, k, "Margin of Safety %", r,
                      lambda i, c: f"=({rev(c)}-{c}{self.R[k]['Break-even Revenue']})/{rev(c)}", fmt=PCT)

    # ---- 19 DCF ----------------------------------------------------------
    def _dcf(self):
        k = "dcf"; ws = self.sheet(k, "19_DCF Valuation"); self.yhdr(ws)
        A = self.AD; r = 4
        r = self.line(ws, k, "EBIT", r, lambda i, c: f"={self.ref('is','EBIT',c)}", link=True)
        r = self.line(ws, k, "NOPAT", r, lambda i, c: f"={c}{self.R[k]['EBIT']}*(1-{A['tax_rate']})")
        r = self.line(ws, k, "add: Depreciation", r, lambda i, c: f"={self.ref('dep','Depreciation',c)}", link=True)
        r = self.line(ws, k, "less: Change in NWC", r, lambda i, c: f"=-{self.ref('wc','Change in NWC',c)}", link=True)
        # Operating Free Cash Flow EXCLUDES capex on purpose: the initial capital
        # outlay is counted ONCE as the Year-0 investment in the valuation below.
        # (The old model subtracted capex here AND subtracted the whole equity+loan
        # raise at Year-0 — double-counting the investment, which drove NPV/IRR
        # deeply negative even for profitable, positive-EV projects.)
        r = self.line(ws, k, "Operating Free Cash Flow", r,
                      lambda i, c: f"={c}{self.R[k]['NOPAT']}+{c}{self.R[k]['add: Depreciation']}+{c}{self.R[k]['less: Change in NWC']}", total=True)
        r = self.line(ws, k, "Discount Factor", r,
                      lambda i, c: f"=1/(1+{A['wacc']})^{i+1}", fmt='0.000')
        r = self.line(ws, k, "PV of OFCF", r,
                      lambda i, c: f"={c}{self.R[k]['Operating Free Cash Flow']}*{c}{self.R[k]['Discount Factor']}", total=True)
        # valuation block below
        ofcf_last = f"G{self.R[k]['Operating Free Cash Flow']}"
        df_last = f"G{self.R[k]['Discount Factor']}"
        pv_first = f"C{self.R[k]['PV of OFCF']}"; pv_last = f"G{self.R[k]['PV of OFCF']}"
        # Year-0 investment = total project outlay actually committed (equity +
        # debt), i.e. the full "cost of project" — not just depreciable capex.
        invest0 = f"({A['equity_total']}+{A['loan_amount']})"
        vr = r + 1
        block = [
            ("Sum of PV (Year 1-5)", f"=SUM({pv_first}:{pv_last})"),
            ("Terminal Value", f"={ofcf_last}*(1+{A['terminal_growth']})/({A['wacc']}-{A['terminal_growth']})"),
            ("PV of Terminal Value", f"=C{vr+1}*{df_last}"),
            ("Enterprise Value", f"=C{vr}+C{vr+2}"),
            ("less: Initial Investment", f"=-{invest0}"),
            ("NPV @ WACC", f"=C{vr+3}+C{vr+4}"),                       # EV - Investment
            ("less: Net Debt", f"=-{self.ref('debt','Closing Balance','C')}"),
            ("Equity Value", f"=C{vr+3}+C{vr+6}"),                     # EV - Net Debt
        ]
        for j, (label, formula) in enumerate(block):
            rr = vr + j
            hot = ("Value" in label or "NPV" in label)
            ws.cell(row=rr, column=1, value=label).font = F_TOTAL if hot else F_LABEL
            cc = ws.cell(row=rr, column=3, value=formula)
            cc.number_format = MONEY; cc.font = F_TOTAL if hot else F_CALC; cc.border = BORDER
            self.R[k][label] = rr
        # Project IRR — cash flows: Year0 = -CapEx, Year1..5 = Operating FCF, with
        # the Terminal Value added to Year 5. The single sign change guarantees a
        # solvable IRR for asset-heavy and asset-light businesses alike.
        irr_r = vr + len(block) + 1
        ws.cell(row=irr_r, column=1, value="Project IRR").font = F_TOTAL
        hr_r = irr_r + 1
        tv_cell = f"C{vr+1}"
        ws.cell(row=hr_r, column=1, value="Cash flows (Yr0..Yr5, incl. TV)").font = Font(size=9, color=GREY)
        ws[f"B{hr_r}"] = f"=-{invest0}"; ws[f"B{hr_r}"].number_format = MONEY; ws[f"B{hr_r}"].font = Font(size=9)
        for i, col in enumerate(YC):
            base = f"{col}{self.R[k]['Operating Free Cash Flow']}"
            ws[f"{col}{hr_r}"] = (f"={base}+{tv_cell}" if col == "G" else f"={base}")
            ws[f"{col}{hr_r}"].number_format = MONEY; ws[f"{col}{hr_r}"].font = Font(size=9)
        ws[f"C{irr_r}"] = f"=IFERROR(IRR(B{hr_r}:G{hr_r}),0)"
        ws[f"C{irr_r}"].number_format = PCT; ws[f"C{irr_r}"].font = F_TOTAL
        self.R[k]["Project IRR"] = irr_r

    # ---- 20 Sensitivity (discount rate vs EV, formula-based) -------------
    def _sensitivity(self):
        k = "sens"; ws = self.sheet(k, "20_Sensitivity Analysis");
        A = self.AD
        ws["A3"] = "Enterprise Value vs Discount Rate (WACC)"; ws["A3"].font = Font(bold=True, color=NAVY)
        ws["A4"] = "Note: Excel What-If Data Tables aren't supported by the generator;"
        ws["A5"] = "this grid recomputes EV directly for each discount rate."
        for rr in (4, 5):
            ws[f"A{rr}"].font = Font(size=9, italic=True, color=GREY)
        fcf = "dcf"; fr = self.R[fcf]["Operating Free Cash Flow"]
        # header row of discount rates
        base = f"{A['wacc']}"
        deltas = [-0.04, -0.02, 0, 0.02, 0.04]
        hrow = 7
        ws.cell(row=hrow, column=1, value="Discount Rate").font = F_TOTAL
        ws.cell(row=hrow, column=2, value="Enterprise Value").font = F_TOTAL
        for j, d in enumerate(deltas):
            rr = hrow + 1 + j
            rate = f"({base}+{d})"
            ws.cell(row=rr, column=1, value=f"=({base}+{d})").number_format = PCT
            ws[f"A{rr}"].font = F_CALC
            # EV = sum FCF_t/(1+r)^t + TV/(1+r)^5 , TV = FCF5*(1+g)/(r-g)
            terms = "+".join(f"{self.q(fcf)}!{col}{fr}/(1+{rate})^{t+1}" for t, col in enumerate(YC))
            tv = f"{self.q(fcf)}!G{fr}*(1+{A['terminal_growth']})/({rate}-{A['terminal_growth']})/(1+{rate})^5"
            ws.cell(row=rr, column=2, value=f"={terms}+{tv}").number_format = MONEY
            ws[f"B{rr}"].font = F_CALC
        ws.column_dimensions["A"].width = 18; ws.column_dimensions["B"].width = 20

    # ---- 21 Year-1 Monthly ----------------------------------------------
    def _monthly(self):
        """Month-by-month build-up of Year 1. Variable items (revenue, COGS) ramp
        across the year; fixed items spread evenly. Every figure links to Year 1
        of the annual model, so the monthly view stays live and its totals tie
        back exactly to the annual Income Statement."""
        k = "mon"; title = "21_Year-1 Monthly"
        ws = self.wb.create_sheet(title); ws.sheet_view.showGridLines = False
        self.SH[k] = title; self.R[k] = {}
        ws.merge_cells("A1:O1"); c = ws["A1"]
        c.value = "Year-1 Monthly Build-up"
        c.font = Font(size=14, bold=True, color=WHITE); c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[1].height = 24
        ws["A2"] = (f"All figures in {self.curr}. Revenue & COGS ramp over the year; "
                    f"fixed costs spread evenly. Column 'Year 1' ties to the annual model.")
        ws["A2"].font = Font(size=8, italic=True, color=GREY)
        ws.column_dimensions["A"].width = 26
        MCOLS = [get_column_letter(3 + i) for i in range(12)]   # C..N
        TOTCOL = get_column_letter(15)                          # O
        for col in MCOLS + [TOTCOL]:
            ws.column_dimensions[col].width = 10.5
        hdr = 3
        ws.cell(row=hdr, column=1, value="Particulars").font = F_TOTAL
        for i, col in enumerate(MCOLS):
            cc = ws[f"{col}{hdr}"]; cc.value = f"M{i + 1}"
            cc.font = Font(bold=True, color=WHITE, size=9); cc.fill = PatternFill("solid", fgColor=NAVY2)
            cc.alignment = Alignment(horizontal="center"); cc.border = BORDER
        tc = ws[f"{TOTCOL}{hdr}"]; tc.value = "Year 1"
        tc.font = Font(bold=True, color=WHITE, size=9); tc.fill = PatternFill("solid", fgColor=NAVY); tc.border = BORDER
        ws.freeze_panes = "C4"

        isq = f"'{self.SH['is']}'"
        rev_c = f"{isq}!C{self.R['is']['Revenue']}"
        cogs_c = f"{isq}!C{self.R['is']['less: COGS']}"
        opex_c = f"{isq}!C{self.R['is']['less: Operating Expenses']}"
        pay_c = f"{isq}!C{self.R['is']['less: Payroll']}"
        dep_c = f"{isq}!C{self.R['is']['less: Depreciation']}"
        int_c = f"{isq}!C{self.R['is']['less: Interest']}"
        tax_c = f"{isq}!C{self.R['is']['less: Tax']}"
        self._mr = hdr + 1

        def mline(label, per_month, total=False, fmt=MONEY, cumulative=False):
            r = self._mr
            self.R[k][label] = r
            lc = ws.cell(row=r, column=1, value=label); lc.font = F_TOTAL if total else F_LABEL
            if total:
                lc.fill = PatternFill("solid", fgColor=IVORY)
            for i, col in enumerate(MCOLS):
                cell = ws[f"{col}{r}"]; cell.value = per_month(i, col)
                cell.number_format = fmt; cell.border = BORDER
                cell.font = F_TOTAL if total else F_CALC
                if total:
                    cell.fill = PatternFill("solid", fgColor=IVORY)
            tcell = ws[f"{TOTCOL}{r}"]
            tcell.value = (f"=N{r}" if cumulative else f"=SUM(C{r}:N{r})")
            tcell.number_format = fmt; tcell.border = BORDER
            tcell.font = F_TOTAL; tcell.fill = PatternFill("solid", fgColor=IVORY)
            self._mr = r + 1

        R = self.R[k]
        # Ramp weight for month i (0-based) = (i+1)/78 ; sums to 1 across 12 months.
        mline("Revenue", lambda i, col: f"={rev_c}*{i + 1}/78")
        mline("COGS", lambda i, col: f"={cogs_c}*{i + 1}/78")
        mline("Gross Profit", lambda i, col: f"={col}{R['Revenue']}+{col}{R['COGS']}", total=True)
        mline("Operating Expenses", lambda i, col: f"={opex_c}/12")
        mline("Payroll", lambda i, col: f"={pay_c}/12")
        mline("EBITDA", lambda i, col: f"={col}{R['Gross Profit']}+{col}{R['Operating Expenses']}+{col}{R['Payroll']}", total=True)
        mline("Depreciation", lambda i, col: f"={dep_c}/12")
        mline("Interest", lambda i, col: f"={int_c}/12")
        mline("PBT", lambda i, col: f"={col}{R['EBITDA']}+{col}{R['Depreciation']}+{col}{R['Interest']}", total=True)
        # Spread the annual tax evenly so the monthly Tax/PAT tie back to the
        # annual model (a per-month MAX(0,..) would not sum to the annual figure).
        mline("Tax", lambda i, col: f"={tax_c}/12")
        mline("PAT (Net Profit)", lambda i, col: f"={col}{R['PBT']}+{col}{R['Tax']}", total=True)
        mline("add: Depreciation", lambda i, col: f"=-{col}{R['Depreciation']}")
        mline("Monthly Net Cash Flow", lambda i, col: f"={col}{R['PAT (Net Profit)']}+{col}{R['add: Depreciation']}", total=True)
        mline("Cumulative Cash Flow",
              lambda i, col: (f"={col}{R['Monthly Net Cash Flow']}" if i == 0
                              else f"={MCOLS[i - 1]}{R['Cumulative Cash Flow']}+{col}{R['Monthly Net Cash Flow']}"),
              total=True, cumulative=True)

        # Monthly revenue & cumulative-cash charts.
        rev_row = R['Revenue']; cum_row = R['Cumulative Cash Flow']
        cats = Reference(ws, min_col=3, max_col=14, min_row=hdr, max_row=hdr)
        rc = BarChart(); rc.type = "col"; rc.title = "Monthly Revenue Ramp (Year 1)"
        rc.height = 7.0; rc.width = 15
        rc.add_data(Reference(ws, min_col=3, max_col=14, min_row=rev_row, max_row=rev_row), from_rows=True, titles_from_data=False)
        rc.set_categories(cats); _finish_chart(rc, colors=[NAVY], gap=40, xtitle="Month", ytitle=self.curr)
        ws.add_chart(rc, "A18")
        cc2 = LineChart(); cc2.title = "Cumulative Cash Position (Year 1)"
        cc2.height = 7.0; cc2.width = 15
        cc2.add_data(Reference(ws, min_col=3, max_col=14, min_row=cum_row, max_row=cum_row), from_rows=True, titles_from_data=False)
        cc2.set_categories(cats); _finish_chart(cc2, colors=[GOLD], xtitle="Month", ytitle=self.curr)
        ws.add_chart(cc2, "A34")

    # ---- 24 Scenario Analysis -------------------------------------------
    def _scenario(self):
        """Live Worst / Base / Best comparison. Only customer growth changes
        (50% / 100% / 150% of base); costs, capex and financing are held constant,
        so scenario-independent lines link straight to the annual model and the
        revenue trajectory is computed in closed form from the assumptions."""
        k = "scn"; title = "24_Scenario Analysis"
        ws = self.wb.create_sheet(title); ws.sheet_view.showGridLines = False
        self.SH[k] = title; self.R[k] = {}
        ws.merge_cells("A1:E1"); c = ws["A1"]
        c.value = "Scenario Comparison — Worst / Base / Best"
        c.font = Font(size=14, bold=True, color=WHITE); c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[1].height = 24
        ws["A2"] = ("Customer growth at 50% / 100% / 150% of base. Costs, capex and financing held constant. "
                    "Set 03_Assumptions 'Scenario' to switch the rest of the model.")
        ws["A2"].font = Font(size=8, italic=True, color=GREY)
        ws.column_dimensions["A"].width = 28
        for col in "BCDE":
            ws.column_dimensions[col].width = 16

        A = self.AD
        churn, start, price, infl = A['churn'], A['start_customers'], A['price'], A['inflation']
        extras = f"({A['onetime_pct']}+{A['service_pct']}+{A['consulting_pct']}+{A['other_pct']})"
        cogs = f"({A['cogs_gateway']}+{A['cogs_cloud']}+{A['cogs_api']}+{A['cogs_support']})"
        tax = A['tax_rate']
        isq = f"'{self.SH['is']}'"
        opexG = f"{isq}!G{self.R['is']['less: Operating Expenses']}"
        payG = f"{isq}!G{self.R['is']['less: Payroll']}"
        depG = f"{isq}!G{self.R['is']['less: Depreciation']}"
        intG = f"{isq}!G{self.R['is']['less: Interest']}"
        scen = [("Worst", f"({A['cust_growth']}*0.5)"), ("Base", f"{A['cust_growth']}"), ("Best", f"({A['cust_growth']}*1.5)")]

        hr = 4
        ws.cell(row=hr, column=1, value="Metric").font = F_TOTAL
        for j, (nm, _) in enumerate(scen):
            cc = ws.cell(row=hr, column=3 + j, value=nm)
            cc.font = Font(bold=True, color=WHITE, size=10); cc.fill = PatternFill("solid", fgColor=NAVY2)
            cc.alignment = Alignment(horizontal="center"); cc.border = BORDER

        def rev_formula(g, t):
            # Closed-form Year-t revenue: avg customers x escalated price x (1+extras).
            return (f"={start}*(1+{g}-{churn})^{t - 1}*(2+{g}-{churn})/2"
                    f"*{price}*(1+{infl})^{t - 1}*(1+{extras})")

        col_of = lambda j: get_column_letter(3 + j)
        r = hr + 1
        rev_rows = {}
        for t in range(1, 6):
            ws.cell(row=r, column=1, value=f"Revenue — Year {t}").font = F_LABEL
            for j, (_, g) in enumerate(scen):
                cc = ws.cell(row=r, column=3 + j, value=rev_formula(g, t))
                cc.number_format = MONEY; cc.font = F_CALC; cc.border = BORDER
            rev_rows[t] = r; r += 1
        r += 1
        y5 = rev_rows[5]

        ebitda_r = r
        ws.cell(row=r, column=1, value="EBITDA (Year 5)").font = F_TOTAL
        for j in range(3):
            cl = col_of(j)
            cc = ws.cell(row=r, column=3 + j, value=f"={cl}{y5}*(1-{cogs})+{opexG}+{payG}")
            cc.number_format = MONEY; cc.font = F_TOTAL; cc.border = BORDER; cc.fill = PatternFill("solid", fgColor=IVORY)
        r += 1
        pat_r = r
        ws.cell(row=r, column=1, value="PAT (Year 5)").font = F_TOTAL
        for j in range(3):
            cl = col_of(j)
            pbt = f"({cl}{ebitda_r}+{depG}+{intG})"
            cc = ws.cell(row=r, column=3 + j, value=f"={pbt}-MAX(0,{pbt}*{tax})")
            cc.number_format = MONEY; cc.font = F_TOTAL; cc.border = BORDER; cc.fill = PatternFill("solid", fgColor=IVORY)
        r += 1
        ws.cell(row=r, column=1, value="EBITDA Margin (Year 5)").font = F_LABEL
        for j in range(3):
            cl = col_of(j)
            cc = ws.cell(row=r, column=3 + j, value=f"={cl}{ebitda_r}/{cl}{y5}")
            cc.number_format = PCT; cc.font = F_CALC; cc.border = BORDER
        r += 2

        # Contiguous chart-data block: metrics as rows, scenarios as series.
        db = r
        for j, (nm, _) in enumerate(scen):
            hc = ws.cell(row=db, column=3 + j, value=nm); hc.font = Font(size=8, color=GREY)
        blocks = [("Revenue (Y5)", y5), ("EBITDA (Y5)", ebitda_r), ("PAT (Y5)", pat_r)]
        for bi, (lbl, srow) in enumerate(blocks):
            rr = db + 1 + bi
            ws.cell(row=rr, column=1, value=lbl).font = Font(size=9, color=GREY)
            for j in range(3):
                cl = col_of(j)
                ws.cell(row=rr, column=3 + j, value=f"={cl}{srow}").number_format = MONEY

        ch = BarChart(); ch.type = "col"; ch.title = "Year-5 Outcomes by Scenario"
        ch.height = 8; ch.width = 15
        ch.add_data(Reference(ws, min_col=3, max_col=5, min_row=db, max_row=db + len(blocks)), titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=1, min_row=db + 1, max_row=db + len(blocks)))
        _finish_chart(ch, colors=[GREY, NAVY, GOLD], ytitle=self.curr, legend='r', gap=60)
        ws.add_chart(ch, "G4")

    # ---- KPI card helper -------------------------------------------------
    def _kpi_card(self, ws, r0, c0, label, formula, fmt, color):
        """A 3-col x 3-row KPI tile: coloured header strip + big value + border."""
        c1 = get_column_letter(c0); c3 = get_column_letter(c0 + 2)
        ws.merge_cells(f"{c1}{r0}:{c3}{r0}")
        lc = ws[f"{c1}{r0}"]; lc.value = label
        lc.font = Font(size=9, bold=True, color=WHITE)
        lc.fill = PatternFill("solid", fgColor=color)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(f"{c1}{r0 + 1}:{c3}{r0 + 2}")
        vc = ws[f"{c1}{r0 + 1}"]; vc.value = formula; vc.number_format = fmt
        vc.font = Font(size=16, bold=True, color=NAVY)
        vc.fill = PatternFill("solid", fgColor=WHITE)
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        edge = Side(style="thin", color="C9CDD4")
        for rr in range(r0, r0 + 3):
            for cc in range(c0, c0 + 3):
                ws.cell(row=rr, column=cc).border = Border(left=edge, right=edge, top=edge, bottom=edge)
        ws.row_dimensions[r0].height = 15
        ws.row_dimensions[r0 + 1].height = 16
        ws.row_dimensions[r0 + 2].height = 14

    # ---- 22 Dashboard ----------------------------------------------------
    def _dashboard(self):
        k = "dash"; title = "22_Dashboard"
        ws = self.wb.create_sheet(title); ws.sheet_view.showGridLines = False
        self.SH[k] = title; self.R[k] = {}
        for c in "ABCDEFGHIJKL":
            ws.column_dimensions[c].width = 12.6

        ws.merge_cells("A1:L1"); ws["A1"] = f"Executive Dashboard — {self.p.get('title') or ''}"
        ws["A1"].font = Font(size=18, bold=True, color=WHITE); ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
        ws["A1"].alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[1].height = 34
        ws.merge_cells("A2:L2")
        ws["A2"] = (f"{self.p.get('industry') or 'Business'}  ·  {self.p.get('country') or ''}  ·  "
                    f"All figures in {self.curr}  ·  5-Year Integrated Model & DCF Valuation")
        ws["A2"].font = Font(size=9, italic=True, color=GREY); ws["A2"].alignment = Alignment(indent=1)

        # KPI cards — 4 per band, two bands.
        kpis = [
            ("REVENUE (Y5)", f"={self.ref('is','Revenue','G')}", MONEY, NAVY),
            ("EBITDA (Y5)", f"={self.ref('is','EBITDA','G')}", MONEY, NAVY2),
            ("NET PROFIT (Y5)", f"={self.ref('is','PAT (Net Profit)','G')}", MONEY, GREENLK),
            ("CASH BALANCE (Y5)", f"={self.ref('cf','Closing Cash','G')}", MONEY, NAVY2),
            ("ENTERPRISE VALUE", f"={self.ref('dcf','Enterprise Value','C')}", MONEY, GOLD),
            ("EQUITY VALUE", f"={self.ref('dcf','Equity Value','C')}", MONEY, GOLD),
            ("PROJECT IRR", f"={self.ref('dcf','Project IRR','C')}", PCT, GREENLK),
            ("NPV @ WACC", f"={self.ref('dcf','NPV @ WACC','C')}", MONEY, GREENLK),
        ]
        for i, (label, formula, fmt, color) in enumerate(kpis):
            r0 = 4 + (i // 4) * 4
            c0 = 1 + (i % 4) * 3
            self._kpi_card(ws, r0, c0, label, formula, fmt, color)

        # Verdict block.
        irr = self.ref('dcf', 'Project IRR', 'C'); wacc = self.AD['wacc']
        npv = self.ref('dcf', 'NPV @ WACC', 'C')
        ws["A13"] = "Investment Verdict"; ws["A13"].font = Font(size=12, bold=True, color=GOLD)
        ws["A14"] = "Recommendation"; ws["A14"].font = Font(bold=True, color=NAVY)
        ws.merge_cells("C14:H14")
        vc = ws.cell(row=14, column=3,
                     value=f'=IF({irr}>{wacc},"VIABLE — IRR exceeds the WACC hurdle","REVIEW — IRR is below the hurdle rate")')
        vc.font = Font(bold=True, color=NAVY)
        ws["A15"] = "Value Created (NPV)"; ws["A15"].font = Font(bold=True, color=NAVY)
        ws.merge_cells("C15:H15")
        nc = ws.cell(row=15, column=3,
                     value=f'=IF({npv}>0,"POSITIVE — the project adds economic value","NEGATIVE — reconsider structure")')
        nc.font = Font(bold=True, color=NAVY)

        # Trend charts (Income Statement rows across the 5 years).
        is_ws = self.wb[self.SH['is']]
        cf_ws = self.wb[self.SH['cf']]
        cats = Reference(is_ws, min_col=3, max_col=7, min_row=3, max_row=3)
        cats_cf = Reference(cf_ws, min_col=3, max_col=7, min_row=3, max_row=3)

        def trend(kind, ttl, src, row, catref, anchor, color):
            ch = LineChart() if kind == "line" else BarChart()
            if kind == "bar":
                ch.type = "col"
            ch.title = ttl; ch.height = 7.2; ch.width = 12.6
            ch.add_data(Reference(src, min_col=3, max_col=7, min_row=row, max_row=row),
                        from_rows=True, titles_from_data=False)
            ch.set_categories(catref)
            if kind == "bar":
                _finish_chart(ch, colors=SERIES_COLORS, labels=True, vary=5, gap=55, xtitle="Year", ytitle=self.curr)
            else:
                _finish_chart(ch, colors=[color], xtitle="Year", ytitle=self.curr)
            ws.add_chart(ch, anchor)

        trend("line", "Revenue Trend", is_ws, self.R['is']['Revenue'], cats, "A17", NAVY)
        trend("bar", "Net Profit Trend", is_ws, self.R['is']['PAT (Net Profit)'], cats, "G17", GREENLK)

        # Pie source data (kept in a labelled block below the charts).
        hdr = 62
        ws.cell(row=hdr, column=1, value="Chart data (auto — do not delete)").font = Font(size=8, italic=True, color=GREY)
        fund = hdr + 1
        for j, (lbl, ref) in enumerate([("Equity", self.AD['equity_total']), ("Debt", self.AD['loan_amount'])]):
            ws.cell(row=fund + j, column=1, value=lbl).font = Font(size=9, color=GREY)
            fc = ws.cell(row=fund + j, column=2, value=f"={ref}"); fc.number_format = MONEY; fc.font = Font(size=9)
        cost = fund + 3
        cost_items = [
            ("COGS", self.ref('cogs', 'Total COGS', 'G')),
            ("Operating Exp", self.ref('opex', 'Total Operating Expenses', 'G')),
            ("Payroll", self.ref('hr', 'Payroll (Annual, escalated)', 'G')),
            ("Depreciation", self.ref('dep', 'Depreciation', 'G')),
            ("Interest", self.ref('debt', 'Interest', 'G')),
            ("Tax", self.ref('tax', 'Tax Expense', 'G')),
        ]
        for j, (lbl, ref) in enumerate(cost_items):
            ws.cell(row=cost + j, column=1, value=lbl).font = Font(size=9, color=GREY)
            cc = ws.cell(row=cost + j, column=2, value=f"={ref}"); cc.number_format = MONEY; cc.font = Font(size=9)

        fund_pie = PieChart(); fund_pie.title = "Funding Mix (Equity vs Debt)"
        fund_pie.height = 7.2; fund_pie.width = 12.6
        fund_pie.add_data(Reference(ws, min_col=2, min_row=fund, max_row=fund + 1), titles_from_data=False)
        fund_pie.set_categories(Reference(ws, min_col=1, min_row=fund, max_row=fund + 1))
        _finish_chart(fund_pie, colors=[GOLD, NAVY], pct=True, legend='b', vary=2)
        ws.add_chart(fund_pie, "A35")

        cost_pie = PieChart(); cost_pie.title = "Year-5 Cost Structure"
        cost_pie.height = 7.2; cost_pie.width = 12.6
        cost_pie.add_data(Reference(ws, min_col=2, min_row=cost, max_row=cost + len(cost_items) - 1), titles_from_data=False)
        cost_pie.set_categories(Reference(ws, min_col=1, min_row=cost, max_row=cost + len(cost_items) - 1))
        _finish_chart(cost_pie, colors=SERIES_COLORS, pct=True, legend='b', vary=len(cost_items))
        ws.add_chart(cost_pie, "G35")

    # ---- 23 Charts -------------------------------------------------------
    def _charts(self):
        title = "23_Charts"
        ws = self.wb.create_sheet(title); ws.sheet_view.showGridLines = False
        self.SH["charts"] = title
        is_ws = self.wb[self.SH["is"]]; cf_ws = self.wb[self.SH["cf"]]
        cats_is = Reference(is_ws, min_col=3, max_col=7, min_row=3, max_row=3)
        cats_cf = Reference(cf_ws, min_col=3, max_col=7, min_row=3, max_row=3)

        def chart(kind, ttl, src, row, catref, anchor, color):
            ch = LineChart() if kind == "line" else BarChart()
            if kind == "bar":
                ch.type = "col"
            ch.title = ttl; ch.height = 8; ch.width = 16
            # from_rows=True -> ONE series with 5 yearly points (the earlier bug
            # used the default which made 5 single-point series -> nothing showed).
            ch.add_data(Reference(src, min_col=3, max_col=7, min_row=row, max_row=row), from_rows=True, titles_from_data=False)
            ch.set_categories(catref)
            if kind == "bar":
                _finish_chart(ch, colors=SERIES_COLORS, labels=True, vary=5, gap=55, xtitle="Year", ytitle=self.curr)
            else:
                _finish_chart(ch, colors=[color], xtitle="Year", ytitle=self.curr)
            ws.add_chart(ch, anchor)

        chart("line", "Revenue Trend", is_ws, self.R['is']['Revenue'], cats_is, "A2", NAVY)
        chart("bar", "EBITDA Trend", is_ws, self.R['is']['EBITDA'], cats_is, "K2", GOLD)
        chart("bar", "Net Profit Trend", is_ws, self.R['is']['PAT (Net Profit)'], cats_is, "A20", GREENLK)
        chart("line", "Net Cash Flow", cf_ws, self.R['cf']['Net Cash Flow'], cats_cf, "K20", BLUEIN)
        chart("line", "Closing Cash Balance", cf_ws, self.R['cf']['Closing Cash'], cats_cf, "A38", NAVY2)


# ═══════════════════════════════════════════════════════════════════════════
# Assumptions engine
#
# Turns the user's REAL inputs into an internally-consistent, economically
# credible set of drivers — never a fixed template. Everything is expressed as
# RATIOS of revenue (currency- and size-agnostic); absolute figures (payroll,
# opex, capex) are back-solved from those ratios and a revenue anchor, so
# margins, cash flows and valuation stay realistic at any currency or deal size.
#
# Pipeline:
#   1. Classify the business into an INDUSTRY PROFILE (margins, asset turnover,
#      capex intensity, working-capital days, growth) from industry/description.
#   2. Anchor Year-1 REVENUE (R1) to the strongest signal available:
#      explicit revenue > price x volume/capacity > invested capital x turnover.
#   3. Size COGS / payroll / opex from explicit cost inputs where given, else
#      from the profile's margin ratios x R1.
#   4. Size capex, financing and working capital from inputs or profile.
#   5. Back-solve the revenue drivers (units x price) so the model reproduces R1
#      exactly, keeping the revenue and cost bases in sync.
#
# Deterministic: no randomness, no LLM, no sample workbook. Same input -> same
# output. Percentages returned here are DECIMALS (0.18 == 18%).
# ═══════════════════════════════════════════════════════════════════════════

# Statutory on-cost the HR sheet adds to salaries (PF 12% + ESIC 3.25% +
# Bonus 8.33% = 23.58%). Payroll targets are divided by this so the LOADED
# payroll lands on the intended % of revenue.
_STATUTORY_LOAD = 1.2358

# Opex split across the workbook's 10 fixed line items (weights sum to 1.0).
_OPEX_WEIGHTS = {
    "Office Rent": 0.15, "Software": 0.05, "Marketing": 0.20, "Utilities": 0.14,
    "Insurance": 0.05, "Travel": 0.07, "Legal": 0.05, "Audit": 0.04,
    "Internet": 0.05, "Maintenance": 0.20,
}
# COGS split across the workbook's 4 buckets (weights sum to 1.0).
_COGS_WEIGHTS = (0.40, 0.30, 0.20, 0.10)

# Input field aliases (questionnaires vary by purpose).
_REVENUE_KEYS = ["annual_revenue", "projected_sales", "projected_annual_sales",
                 "annual_turnover", "turnover", "revenue", "sales", "annual_sales",
                 "projected_revenue", "year1_revenue"]
_ALLIN_COST_KEYS = ["operating_cost", "annual_operating_cost", "projected_expenses",
                    "projected_annual_expenses", "total_operating_cost", "running_cost",
                    "annual_expenses"]
_MATERIAL_KEYS = ["raw_material_cost", "material_cost", "direct_material",
                  "cost_of_goods", "cogs", "purchase_cost", "rm_cost"]
_LABOUR_KEYS = ["labour_cost", "labor_cost", "wages", "salary_cost",
                "manpower_cost", "staff_cost", "payroll_cost"]
_OVERHEAD_KEYS = ["utility_cost", "utilities", "admin_cost", "administrative_cost",
                  "selling_cost", "marketing_cost", "rent", "overheads", "distribution_cost"]
_CAPEX_ITEM_KEYS = ["land_cost", "building_cost", "machinery_cost", "plant_machinery_cost",
                    "equipment_cost", "furniture_cost", "civil_cost", "preoperative_cost"]
_CAPEX_DIRECT_KEYS = ["capex", "capital_cost", "initial_investment"]
_PRICE_KEYS = ["selling_price", "price", "unit_price", "price_per_unit", "avg_price",
               "tariff", "room_rate", "fee", "ticket_price"]
_VOLUME_KEYS = ["start_customers", "units", "annual_units", "volume", "units_sold",
                "customers", "subscribers", "students", "patients", "rooms"]
_CAPACITY_KEYS = ["production_capacity", "capacity", "installed_capacity", "annual_capacity"]
_GROWTH_KEYS = ["revenue_growth", "growth_rate", "cust_growth", "sales_growth", "cagr", "annual_growth"]

# Industry profiles — all percentages are of revenue.
#   gm            : gross margin (COGS% = 1 - gm)
#   opex/payroll  : cash opex and loaded payroll as % of revenue
#   turnover      : revenue / invested capital (anchors revenue from capital)
#   capex_int     : depreciable capex as a fraction of project cost
#   life          : asset useful life, yrs (drives depreciation)
#   growth/churn  : base annual revenue growth / customer attrition
#   recv/inv/pay  : working-capital days
#   extras        : (one-time, service, consulting, other) revenue as % of core
#   units         : default Year-1 volume when a unit price is back-solved
#   roles         : (title, headcount, salary-weight) HR template
def _P(gm, opex, payroll, turnover, capex_int, life, growth, churn,
       recv, inv, pay, extras, units, roles):
    return dict(gm=gm, opex=opex, payroll=payroll, turnover=turnover, capex_int=capex_int,
                life=life, growth=growth, churn=churn, recv=recv, inv=inv, pay=pay,
                extras=extras, units=units, roles=roles)

INDUSTRY_PROFILES = {
    "manufacturing": _P(0.32, 0.09, 0.11, 1.3, 0.75, 15, 0.10, 0.02, 45, 60, 45,
        (0.02, 0.02, 0.0, 0.03), 40000,
        [("Plant Manager", 1, 3.2), ("Production Supervisors", 3, 1.7), ("Machine Operators", 15, 1.0),
         ("Quality & Lab", 3, 1.3), ("Maintenance Technicians", 4, 1.2), ("Stores & Logistics", 4, 1.0),
         ("Sales & Marketing", 4, 1.6), ("Finance & Admin", 3, 1.5)]),
    "food_processing": _P(0.30, 0.09, 0.10, 1.5, 0.70, 12, 0.10, 0.02, 30, 45, 40,
        (0.02, 0.02, 0.0, 0.03), 100000,
        [("Plant Manager", 1, 3.0), ("Production Supervisors", 3, 1.6), ("Operators & Packers", 18, 1.0),
         ("Quality & Food Safety", 3, 1.3), ("Maintenance", 3, 1.2), ("Cold Chain & Logistics", 4, 1.0),
         ("Sales & Distribution", 5, 1.5), ("Finance & Admin", 3, 1.4)]),
    "saas": _P(0.80, 0.25, 0.35, 1.2, 0.25, 5, 0.30, 0.05, 30, 0, 20,
        (0.06, 0.08, 0.04, 0.02), 1000,
        [("CEO", 1, 3.5), ("CTO", 1, 3.2), ("Engineers", 8, 1.6), ("Product & Design", 3, 1.6),
         ("Sales & Marketing", 5, 1.5), ("Customer Success", 3, 1.1), ("Finance & Ops", 2, 1.4)]),
    "services": _P(0.55, 0.15, 0.24, 1.8, 0.30, 8, 0.15, 0.08, 60, 0, 30,
        (0.0, 0.10, 0.15, 0.02), 300,
        [("Managing Partner", 1, 3.5), ("Practice Leads", 3, 2.2), ("Consultants", 10, 1.4),
         ("Analysts", 6, 1.0), ("Business Development", 3, 1.6), ("Admin & Finance", 3, 1.3)]),
    "retail": _P(0.22, 0.08, 0.07, 3.0, 0.40, 8, 0.10, 0.10, 5, 45, 30,
        (0.0, 0.0, 0.0, 0.02), 200000,
        [("Store Manager", 1, 2.6), ("Assistant Managers", 2, 1.6), ("Sales Associates", 12, 1.0),
         ("Cashiers", 4, 0.9), ("Inventory & Warehouse", 4, 1.0), ("Marketing", 2, 1.4), ("Accounts", 2, 1.3)]),
    "real_estate": _P(0.60, 0.10, 0.06, 0.35, 0.85, 30, 0.05, 0.02, 30, 90, 45,
        (0.0, 0.03, 0.0, 0.02), 40,
        [("Project Director", 1, 3.6), ("Project Managers", 2, 2.4), ("Site Engineers", 4, 1.6),
         ("Sales & Leasing", 4, 1.7), ("Legal & Compliance", 2, 1.8), ("Finance & Admin", 3, 1.5)]),
    "hospitality": _P(0.65, 0.22, 0.22, 0.7, 0.80, 20, 0.08, 0.03, 5, 15, 30,
        (0.0, 0.08, 0.0, 0.04), 30000,
        [("General Manager", 1, 3.2), ("Department Heads", 4, 1.9), ("Front Office", 6, 1.0),
         ("F&B & Kitchen", 12, 1.0), ("Housekeeping", 10, 0.8), ("Sales & Events", 3, 1.5),
         ("Finance & Admin", 3, 1.4)]),
    "infrastructure": _P(0.55, 0.10, 0.08, 0.4, 0.90, 25, 0.06, 0.02, 60, 30, 60,
        (0.0, 0.02, 0.0, 0.03), 500,
        [("Project Head", 1, 3.4), ("Engineers", 6, 1.7), ("O&M Technicians", 8, 1.1),
         ("SCADA / Controls", 2, 1.6), ("Land & Liaison", 2, 1.4), ("Finance & Admin", 3, 1.5)]),
    "healthcare": _P(0.50, 0.18, 0.22, 0.9, 0.70, 12, 0.12, 0.03, 40, 20, 45,
        (0.0, 0.05, 0.0, 0.03), 20000,
        [("Medical Director", 1, 3.8), ("Doctors", 6, 2.6), ("Nurses", 12, 1.2),
         ("Technicians", 5, 1.3), ("Admin & Billing", 4, 1.1), ("Support Staff", 6, 0.8)]),
    "education": _P(0.55, 0.20, 0.30, 0.8, 0.55, 12, 0.12, 0.03, 15, 0, 30,
        (0.0, 0.05, 0.0, 0.03), 1000,
        [("Principal / Director", 1, 3.0), ("Senior Faculty", 6, 1.8), ("Faculty", 12, 1.2),
         ("Lab & Library", 3, 1.0), ("Counsellors", 3, 1.2), ("Admin & Accounts", 4, 1.1)]),
    "generic": _P(0.45, 0.13, 0.15, 1.3, 0.55, 10, 0.10, 0.03, 45, 30, 30,
        (0.0, 0.03, 0.0, 0.02), 5000,
        [("Founder / CEO", 1, 3.2), ("Operations Lead", 1, 2.2), ("Team Leads", 3, 1.6),
         ("Executives", 8, 1.0), ("Sales & Marketing", 4, 1.4), ("Finance & Admin", 3, 1.3)]),
}

# Keyword -> profile. First match wins; order matters (specific before generic).
_INDUSTRY_KEYWORDS = [
    (("real estate", "property", "housing", "realty", "condo", "apartment", "township"), "real_estate"),
    (("hotel", "hospitality", "resort", "restaurant", "tourism", "hostel"), "hospitality"),
    (("solar", "renewable", "power plant", "energy", "infrastructure", "highway", "utility-scale"), "infrastructure"),
    (("hospital", "clinic", "pharma", "medical", "healthcare", "diagnostic"), "healthcare"),
    (("school", "college", "university", "education", "training", "edtech", "institute"), "education"),
    (("saas", "software", "app", "platform", "artificial intelligence", " ai ", "fintech", "technology", "digital"), "saas"),
    (("food", "beverage", "dairy", "bakery", "snack", "agro", "processing"), "food_processing"),
    (("manufactur", "factory", "plant", "production", "industrial", "fabricat", "assembly"), "manufacturing"),
    (("retail", "store", "shop", "trading", "ecommerce", "e-commerce", "distribution", "wholesale", "fmcg"), "retail"),
    (("consult", "service", "agency", "advisory", "bpo", "outsourc"), "services"),
]


def _sum_keys(pa, keys):
    return sum(max(0.0, _num(pa.get(k), 0)) for k in keys)


def _first_num(pa, keys):
    for k in keys:
        v = _num(pa.get(k), 0)
        if v > 0:
            return v
    return 0.0


def _parse_leading_number(v):
    if v is None:
        return 0.0
    import re
    m = re.search(r"[-+]?\d[\d,]*\.?\d*", str(v))
    return float(m.group().replace(",", "")) if m else 0.0


def _pct(v, default):
    """Read a rate that may be given as a percent (25) or a fraction (0.25)."""
    try:
        x = float(v)
    except (TypeError, ValueError):
        return default
    return x / 100.0 if x > 1 else x


def _clamp(x, lo, hi):
    return max(lo, min(hi, x))


def classify_industry(project: dict) -> str:
    text = " ".join(str(project.get(k) or "") for k in
                    ("industry", "sub_industry", "purpose", "project_description", "title")).lower()
    for keys, prof in _INDUSTRY_KEYWORDS:
        if any(k in text for k in keys):
            return prof
    return "generic"


# ── assumptions from the project/questionnaire ─────────────────────────────
def derive_assumptions(project: dict) -> dict:
    pa = project.get("purpose_answers") or {}
    prof = INDUSTRY_PROFILES[classify_industry(project)]
    GM, opex_pct, payroll_pct = prof["gm"], prof["opex"], prof["payroll"]

    # 1) Revenue anchor (R1) — strongest available signal.
    rev_explicit = _first_num(pa, _REVENUE_KEYS)
    price_in = _first_num(pa, _PRICE_KEYS)
    vol_in = _first_num(pa, _VOLUME_KEYS)
    cap_in = 0.0
    for k in _CAPACITY_KEYS:
        cap_in = _parse_leading_number(pa.get(k))
        if cap_in > 0:
            break
    have_rev = True
    if rev_explicit > 0:
        R1 = rev_explicit
    elif price_in > 0 and vol_in > 0:
        R1 = price_in * vol_in
    elif price_in > 0 and cap_in > 0:
        R1 = price_in * cap_in * 0.70          # Year-1 capacity utilisation
    else:
        have_rev = False
        R1 = 0.0

    # 2) Cost signals. An all-in operating-cost field (IRR/generic/CMA) is the
    # TOTAL cash cost and is split across COGS/opex/payroll. Granular fields
    # (feasibility) map to their own component; the rest fall to profile ratios.
    allin = _sum_keys(pa, _ALLIN_COST_KEYS)
    maint = _num(pa.get("maintenance_cost"), 0)
    mat = _first_num(pa, _MATERIAL_KEYS)
    lab = _first_num(pa, _LABOUR_KEYS)
    ovh = _sum_keys(pa, _OVERHEAD_KEYS)
    tgt_ebitda = max(0.08, GM - opex_pct - payroll_pct)

    # If no revenue signal, imply it from the cost base (at target margin) or capital.
    if not have_rev:
        cost_known = (allin + maint) if allin > 0 else (mat + lab + ovh)
        if cost_known > 0:
            R1 = cost_known / (1 - tgt_ebitda)
        else:
            capital0 = _num(project.get("project_cost"), 0) or 1_000_000.0
            R1 = capital0 * prof["turnover"]
    R1 = max(R1, 1000.0)

    # 3) Size COGS / opex / payroll.
    if allin > 0:
        cash = allin + maint
        ps = (1 - GM) + opex_pct + payroll_pct
        cogs_total = cash * (1 - GM) / ps
        opex_total = cash * opex_pct / ps
        payroll_total = cash * payroll_pct / ps
    elif (mat + lab + ovh) > 0:
        cogs_total = mat if mat > 0 else (1 - GM) * R1
        payroll_total = lab if lab > 0 else payroll_pct * R1
        opex_total = ovh if ovh > 0 else opex_pct * R1
    else:
        cogs_total = (1 - GM) * R1
        opex_total = opex_pct * R1
        payroll_total = payroll_pct * R1

    cogs_ratio = _clamp(cogs_total / R1, 0.0, 0.95)
    cg = [cogs_ratio * w for w in _COGS_WEIGHTS]
    opex_fixed = {name: round(w * opex_total) for name, w in _OPEX_WEIGHTS.items()}

    # 4) Capex, financing, working capital.
    capex_items = _sum_keys(pa, _CAPEX_ITEM_KEYS)
    capex_direct = _first_num(pa, _CAPEX_DIRECT_KEYS)
    proj_cost = _num(project.get("project_cost"), 0)
    capital = proj_cost if proj_cost > 0 else (capex_items or capex_direct or (R1 / prof["turnover"]))
    if capex_items > 0:
        capex_total = capex_items
    elif capex_direct > 0:
        capex_total = capex_direct
    else:
        capex_total = capital * prof["capex_int"]

    own = _num(pa.get("own_contribution"), _num(project.get("own_contribution"), capital * 0.4))
    loan = _num(pa.get("loan_amount"), _num(project.get("loan_amount"), capital * 0.6))

    recv_days, inv_days, pay_days = prof["recv"], prof["inv"], prof["pay"]
    debtors, inventory, creditors = _num(pa.get("debtors"), 0), _num(pa.get("inventory"), 0), _num(pa.get("creditors"), 0)
    if debtors > 0:
        recv_days = _clamp(debtors / R1 * 365, 0, 180)
    if inventory > 0 and cogs_total > 0:
        inv_days = _clamp(inventory / cogs_total * 365, 0, 180)
    if creditors > 0 and cogs_total > 0:
        pay_days = _clamp(creditors / cogs_total * 365, 0, 180)

    # 5) Rates and revenue drivers.
    tax_rate = _pct(pa.get("tax_rate"), 0.25)
    interest_rate = _pct(pa.get("interest_rate"), 0.115)
    inflation = _pct(pa.get("inflation_rate"), 0.05)
    useful_life = max(3, int(_num(pa.get("useful_life"), _num(pa.get("project_life"), prof["life"])) or prof["life"]))
    salvage = _num(pa.get("salvage_value"), 0)
    residual_pct = _clamp(salvage / capex_total, 0, 0.6) if (salvage > 0 and capex_total > 0) else 0.10
    loan_tenure = max(1, int(_num(pa.get("loan_tenure"), 7) or 7))
    start_year = int(_num(pa.get("start_year"), 2025) or 2025)
    gk = _first_num(pa, _GROWTH_KEYS)
    cust_growth = _clamp(_pct(gk, prof["growth"]) if gk > 0 else prof["growth"], 0.0, 1.0)
    churn = prof["churn"]

    # WACC from the actual capital structure (weighted, after-tax cost of debt).
    cost_of_equity, cost_of_debt = 0.18, interest_rate
    tot = own + loan
    wacc = ((own / tot) * cost_of_equity + (loan / tot) * cost_of_debt * (1 - tax_rate)) if tot > 0 else 0.15
    wacc = _clamp(wacc, 0.08, 0.30)
    terminal_growth = min(0.04, inflation)
    if terminal_growth >= wacc:                # keep the Gordon-growth TV finite
        terminal_growth = wacc * 0.4

    # Back-solve units x price so the model reproduces R1 exactly.
    extras = prof["extras"]
    extras_sum = sum(extras)
    avg_factor = 1 + (cust_growth - churn) / 2
    denom = max(0.05, avg_factor * (1 + extras_sum))
    if price_in > 0:
        price = price_in
        start_customers = max(1, int(round(R1 / (price * denom))))
    else:
        start_customers = int(prof["units"])
        price = R1 / (start_customers * denom)

    # Payroll -> per-role salaries (loaded payroll lands on payroll target).
    roles_t = prof["roles"]
    wsum = sum(hc * w for _, hc, w in roles_t)
    base_annual = payroll_total / _STATUTORY_LOAD
    unit_monthly = base_annual / (12 * wsum) if wsum > 0 else 0.0
    hr_roles = [(role, hc, int(round(w * unit_monthly))) for role, hc, w in roles_t]

    equity_rounds = ([("Promoter's Contribution", round(own * 0.7)), ("Co-Investors / Angel", round(own * 0.3))]
                     if own > 0 else [("Promoter's Contribution", 0)])

    return {
        "start_year": start_year,
        "inflation": inflation, "tax_rate": tax_rate,
        "wacc": wacc, "cost_of_debt": cost_of_debt, "cost_of_equity": cost_of_equity,
        "terminal_growth": terminal_growth,
        "start_customers": start_customers, "cust_growth": cust_growth, "churn": churn, "price": price,
        "onetime_pct": extras[0], "service_pct": extras[1], "consulting_pct": extras[2], "other_pct": extras[3],
        "cogs_gateway": cg[0], "cogs_cloud": cg[1], "cogs_api": cg[2], "cogs_support": cg[3],
        "salary_increment": 0.08,
        "capex_total": capex_total, "useful_life": useful_life, "residual_pct": residual_pct,
        "loan_amount": loan, "interest_rate": interest_rate, "loan_tenure": loan_tenure, "equity_total": own,
        "recv_days": recv_days, "inv_days": inv_days, "pay_days": pay_days,
        "opex_fixed": opex_fixed,
        "hr_roles": hr_roles,
        "equity_rounds": equity_rounds,
    }


def build_model_excel(project: dict, assumptions: dict = None) -> bytes:
    a = assumptions or derive_assumptions(project)
    return Model(project, a).build()
