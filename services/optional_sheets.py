"""
optional_sheets.py

Sections the report should only carry when they mean something.

The workbook templates hold every section, because a template cannot know the borrower.
The Existing Loan sheet is the clearest case: a borrower with no prior debt should not
receive a schedule of zeros and a balance sheet line for a loan that does not exist — the
client's instruction was "varna banane ki zaroorat hi nahi hai".

Deleting a sheet is not enough on its own: five other sheets reference it, and Excel would
render #REF! across the model. So every reference is rewritten to drop the term FIRST, and
only then is the sheet removed — leaving a workbook that reads as if the feature had never
been part of it.
"""

from __future__ import annotations

import logging
import re
from io import BytesIO

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

EXISTING_LOAN_SHEET = "Existing Loan"
_TREATMENT = "Assumptions!C72"
_OUTSTANDING = "Assumptions!C73"
_ASSUMPTION_ROWS = range(71, 76)          # the input block, header included

# "+'Existing Loan'!C14"  or  "+'Existing Loan'!C14/12"
_TERM = re.compile(r"\+\s*'" + re.escape(EXISTING_LOAN_SHEET) + r"'!\$?[A-Z]{1,3}\$?\d+(?:/12)?")
# the Form VI bring-forward: "+IF(Assumptions!$C$72=2,Assumptions!$C$73,0)"
_BRING_FORWARD = re.compile(r"\+\s*IF\(Assumptions!\$C\$72=2,\s*Assumptions!\$C\$73,\s*0\)")


def has_existing_loan(answers: dict) -> bool:
    """True only when a real outstanding amount was given."""
    if not isinstance(answers, dict):
        return False
    v = answers.get(_OUTSTANDING)
    return isinstance(v, (int, float)) and not isinstance(v, bool) and v > 0


def strip_existing_loan(xlsx_bytes: bytes) -> bytes:
    """Remove the Existing Loan sheet and every trace of it. Returns the new workbook.

    Non-fatal by design: if anything about the workbook is unexpected the original bytes
    come back, because shipping the extra sheet is far better than shipping a broken file.
    """
    try:
        wb = load_workbook(BytesIO(xlsx_bytes))
        if EXISTING_LOAN_SHEET not in wb.sheetnames:
            return xlsx_bytes

        rewritten = 0
        for ws in wb.worksheets:
            if ws.title == EXISTING_LOAN_SHEET:
                continue
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if not isinstance(v, str) or not v.startswith("="):
                        continue
                    new = _TERM.sub("", v)
                    new = _BRING_FORWARD.sub("", new)
                    if new != v:
                        cell.value = new
                        rewritten += 1

        # the label was widened for the bring-forward that no longer applies
        if "Form_VI_FundFlow" in wb.sheetnames:
            fv = wb["Form_VI_FundFlow"]
            if isinstance(fv.cell(8, 2).value, str) and "brought forward" in fv.cell(8, 2).value:
                fv.cell(8, 2, "Term loan raised")

        a = wb["Assumptions"] if "Assumptions" in wb.sheetnames else None
        if a is not None:
            for r in _ASSUMPTION_ROWS:
                for c in (2, 3, 4):
                    a.cell(r, c).value = None

        # the Cover lists the sheet by name in its contents table — plain text, not a
        # reference, but it would advertise a tab that is about to disappear
        if "Cover" in wb.sheetnames:
            cv = wb["Cover"]
            for row in cv.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.strip() == EXISTING_LOAN_SHEET:
                        cv.cell(c.row, c.column).value = None
                        nxt = cv.cell(c.row, c.column + 1)
                        if not isinstance(nxt.value, str) or not nxt.value.startswith("="):
                            nxt.value = None

        del wb[EXISTING_LOAN_SHEET]

        # No FORMULA may still point at the sheet we just removed. Plain text mentioning
        # it is not a reference and must not veto the removal.
        left = [f"{ws.title}!{c.coordinate}" for ws in wb.worksheets
                for row in ws.iter_rows() for c in row
                if isinstance(c.value, str) and c.value.startswith("=")
                and EXISTING_LOAN_SHEET in c.value]
        if left:
            logger.warning("optional_sheets: %d references to %r survived; keeping the "
                           "sheet rather than shipping #REF!", len(left), EXISTING_LOAN_SHEET)
            return xlsx_bytes

        buf = BytesIO()
        wb.save(buf)
        logger.info("optional_sheets: no prior debt — dropped %r and cleaned %d formulas",
                    EXISTING_LOAN_SHEET, rewritten)
        return buf.getvalue()
    except Exception:
        logger.warning("optional_sheets: could not drop %r; shipping it as-is",
                       EXISTING_LOAN_SHEET, exc_info=True)
        return xlsx_bytes


def apply(xlsx_bytes: bytes, answers: dict) -> bytes:
    """Drop every optional section this borrower has no use for."""
    if not has_existing_loan(answers):
        return strip_existing_loan(xlsx_bytes)
    return xlsx_bytes
