"""
recalc_service.py

Server-side recalculation of a filled .xlsx using LibreOffice (headless).

`template_fill_service.fill_template` writes the input cells and flips the
workbook's fullCalcOnLoad flag, but the formula RESULTS are only computed when a
spreadsheet application actually opens the file. So the derived figures the report
cares about — DSCR, IRR, NPV, the ratio grid — do not exist as numbers until then.

To make the Word (and any PDF) report quote the SAME numbers the Excel shows, we
recalculate on the server: open the workbook in headless LibreOffice configured to
"always recalculate on load", let it compute every formula, and save. We then read
the cached results back with openpyxl(data_only=True) and feed them to the report.

Reliability first — everything degrades gracefully:
  * If LibreOffice is not installed, `libreoffice_available()` is False and callers
    fall back to the input-derived headline figures. Generation never breaks.
  * If a recalc subprocess fails/times out, the exception is surfaced to the caller,
    which logs and falls back.

The soffice binary is discovered from $LIBREOFFICE_PATH / $SOFFICE_BIN, then PATH,
then the usual per-OS install locations.
"""

import io
import logging
import os
import shutil
import subprocess
import tempfile

from openpyxl import load_workbook

logger = logging.getLogger("recalc")

# A private LibreOffice user profile carrying exactly one setting: recalculate every
# formula on load (0 = Always) for both OOXML (.xlsx) and ODF. Combined with the
# workbook's fullCalcOnLoad flag, a single --convert-to pass recomputes the model.
_RECALC_XCU = (
    '<?xml version="1.0" encoding="UTF-8"?>\n'
    '<oor:items xmlns:oor="http://openoffice.org/2001/registry" '
    'xmlns:xs="http://www.w3.org/2001/XMLSchema" '
    'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">\n'
    ' <item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
    '<prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>\n'
    ' <item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
    '<prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop></item>\n'
    '</oor:items>\n'
)

_SOFFICE_CANDIDATES = (
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    "/usr/bin/soffice",
    "/usr/bin/libreoffice",
    "/opt/libreoffice/program/soffice",
    "/snap/bin/libreoffice",
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
)


def _soffice_bin():
    """Path to the LibreOffice `soffice` binary, or None if not found."""
    for env in ("LIBREOFFICE_PATH", "SOFFICE_BIN"):
        p = os.environ.get(env)
        if p and os.path.isfile(p):
            return p
    for name in ("soffice", "libreoffice"):
        p = shutil.which(name)
        if p:
            return p
    for p in _SOFFICE_CANDIDATES:
        if os.path.isfile(p):
            return p
    return None


def libreoffice_available() -> bool:
    return _soffice_bin() is not None


def _to_file_url(path: str) -> str:
    p = os.path.abspath(path).replace("\\", "/")
    return "file:///" + p.lstrip("/")


def recalculate_xlsx(xlsx_bytes: bytes, timeout: int = 180) -> bytes:
    """Return the workbook bytes with every formula recomputed by LibreOffice.

    Raises RuntimeError if LibreOffice is unavailable or the recalc produced no
    output; propagates subprocess.TimeoutExpired on timeout. Callers treat any
    failure as "recalc unavailable" and fall back to input-derived figures.
    """
    soffice = _soffice_bin()
    if not soffice:
        raise RuntimeError("LibreOffice not found (set LIBREOFFICE_PATH or SOFFICE_BIN)")

    workdir = tempfile.mkdtemp(prefix="recalc_")
    try:
        profile = os.path.join(workdir, "profile")
        os.makedirs(os.path.join(profile, "user"), exist_ok=True)
        with open(os.path.join(profile, "user", "registrymodifications.xcu"), "w", encoding="utf-8") as fh:
            fh.write(_RECALC_XCU)

        src = os.path.join(workdir, "model.xlsx")
        outdir = os.path.join(workdir, "out")
        os.makedirs(outdir, exist_ok=True)
        with open(src, "wb") as fh:
            fh.write(xlsx_bytes)

        cmd = [
            soffice, "--headless", "--norestore", "--nologo",
            "--nofirststartwizard", "--nolockcheck",
            "-env:UserInstallation=" + _to_file_url(profile),
            "--convert-to", "xlsx:Calc MS Excel 2007 XML",
            "--outdir", outdir, src,
        ]
        proc = subprocess.run(cmd, capture_output=True, timeout=timeout)
        out_file = os.path.join(outdir, "model.xlsx")
        if not os.path.isfile(out_file):
            raise RuntimeError(
                f"LibreOffice recalc produced no output (rc={proc.returncode}): "
                f"{(proc.stderr or b'')[:300]!r}")
        with open(out_file, "rb") as fh:
            data = fh.read()
        logger.info("recalc: recomputed workbook (%d -> %d bytes)", len(xlsx_bytes), len(data))
        return data
    finally:
        shutil.rmtree(workdir, ignore_errors=True)


def to_pdf(data: bytes, suffix: str, timeout: int = 180) -> bytes:
    """Convert an office document (xlsx/xlsm/docx) to PDF with LibreOffice, so a PDF
    report is rendered from the SAME file as the Excel/Word — identical values, no
    re-computation. `suffix` is the source extension without the dot (e.g. 'docx').
    Raises RuntimeError if LibreOffice is unavailable."""
    soffice = _soffice_bin()
    if not soffice:
        raise RuntimeError("LibreOffice not found (set LIBREOFFICE_PATH or SOFFICE_BIN)")
    workdir = tempfile.mkdtemp(prefix="topdf_")
    try:
        profile = os.path.join(workdir, "profile")
        os.makedirs(os.path.join(profile, "user"), exist_ok=True)
        with open(os.path.join(profile, "user", "registrymodifications.xcu"), "w", encoding="utf-8") as fh:
            fh.write(_RECALC_XCU)   # recalc-on-load too, so xlsx PDFs show computed values
        src = os.path.join(workdir, f"doc.{suffix}")
        outdir = os.path.join(workdir, "out")
        os.makedirs(outdir, exist_ok=True)
        with open(src, "wb") as fh:
            fh.write(data)
        cmd = [soffice, "--headless", "--norestore", "--nologo", "--nofirststartwizard",
               "--nolockcheck", "-env:UserInstallation=" + _to_file_url(profile),
               "--convert-to", "pdf", "--outdir", outdir, src]
        proc = subprocess.run(cmd, capture_output=True, timeout=timeout)
        out_file = os.path.join(outdir, "doc.pdf")
        if not os.path.isfile(out_file):
            raise RuntimeError(
                f"PDF conversion produced no output (rc={proc.returncode}): "
                f"{(proc.stderr or b'')[:300]!r}")
        with open(out_file, "rb") as fh:
            return fh.read()
    finally:
        shutil.rmtree(workdir, ignore_errors=True)


def _fmt(v, ftype: str) -> str:
    if not isinstance(v, (int, float)) or isinstance(v, bool):
        return str(v)
    if ftype == "percent":
        return f"{v * 100:.1f}%"
    if ftype == "ratio_x":
        return f"{v:.2f}x"
    if ftype == "ratio":
        return f"{v:.2f} : 1"
    if ftype == "years":
        return f"{v:.1f} yrs"
    # crore / plain number
    if float(v).is_integer():
        return f"{v:,.0f}"
    return f"{v:,.2f}"


def read_computed_kpis(schema: dict, xlsx_bytes: bytes) -> list:
    """Read the recalculated values for the schema's `computed_kpis` spec —
    a list of {label, cell:"Sheet!Cell", type} — from a workbook LibreOffice has
    recomputed. Returns [{label, value}], skipping any cell that resolves to None.
    """
    specs = schema.get("computed_kpis") or []
    if not specs:
        return []
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    try:
        out = []
        for sp in specs:
            ref = sp.get("cell", "")
            if "!" not in ref:
                continue
            sheet, cell = ref.rsplit("!", 1)
            if sheet not in wb.sheetnames:
                continue
            v = wb[sheet][cell].value
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            out.append({"label": sp["label"], "value": _fmt(v, sp.get("type", "number"))})
        return out
    finally:
        wb.close()
