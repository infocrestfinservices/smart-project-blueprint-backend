"""
verify_templates.py

Every fix, asserted against every template.  Run it after ANY change to a workbook:

    python verify_templates.py

Why this file exists
--------------------
Fixes in this project are migration scripts run against eleven .xlsx binaries. There is no
compiler and no import graph to catch a template that missed one, so for a long time each
feature carried only its OWN check. That let a fix silently disappear:

  * The monthly-moratorium fix was applied to the ten industry templates and deliberately
    NOT to bank_loan, to keep Bank Loan byte-identical. That reason expired the day
    bank_loan was rebuilt at the client's request — but nobody went back for it, and a
    manufacturing report kept repaying principal straight through its holiday period
    (Year-1 principal doubled, DSCR Yr1 read 1.75 instead of 2.61).

  * Several migrations restore from their own .bak before applying, so they can be re-run
    safely. That is only true in isolation: if migration A's backup predates migration B,
    re-running A silently reverts B.

Neither failure is visible in a single report. Both are obvious here.

Add a check whenever you add a fix. A check returns None when the invariant holds, or a
short string describing what is wrong.
"""

from __future__ import annotations

import glob
import os
import re
import sys
import zipfile

import openpyxl

BACKEND = os.path.dirname(os.path.abspath(__file__))
TEMPLATES = os.path.join(BACKEND, "templates")
if BACKEND not in sys.path:
    sys.path.insert(0, BACKEND)

FOLDERS = ["bank_loan", "hotel_cma", "retail_cma", "software_cma", "hospital_cma",
           "education_cma", "restaurant_cma", "trading_cma", "transport_cma",
           "media_cma", "other_cma"]

CHECKS: dict[str, callable] = {}


def check(name):
    def deco(fn):
        CHECKS[name] = fn
        return fn
    return deco


@check("moratorium honoured in the monthly repayment schedule")
def _(folder, path, wb):
    if "Repayment" not in wb.sheetnames:
        return "no Repayment sheet"
    rp = wb["Repayment"]
    flat, guarded = [], 0
    for r in range(20, min(rp.max_row, 60) + 1):
        for c in range(3, 16):
            v = rp.cell(r, c).value
            if not isinstance(v, str) or not v.startswith("="):
                continue
            if re.fullmatch(r"=\$[A-Z]\$\d+/12", v.strip()):
                flat.append(f"{rp.cell(r, c).coordinate}={v}")
            elif "$C$12" in v:
                guarded += 1
    if flat and not guarded:
        return f"flat /12 principal repays through the moratorium (e.g. {flat[0]})"
    if not guarded:
        return "no moratorium term ($C$12) in the monthly schedule"
    return None


@check("revenue streams have both a volume and a price cell")
def _(folder, path, wb):
    a = wb["Assumptions"]
    missing = [f"{col}{r}" for r in range(66, 70) for col in ("C", "D")
               if a.cell(r, 3 if col == "C" else 4).value is None]
    return f"missing driver cells {missing}" if missing else None


@check("target-market segment block present")
def _(folder, path, wb):
    return None if wb["Assumptions"].cell(58, 2).value else "no segment rows at B58"


@check("existing-loan inputs and sheet present")
def _(folder, path, wb):
    if "Existing Loan" not in wb.sheetnames:
        return "no 'Existing Loan' sheet"
    a = wb["Assumptions"]
    for r, what in ((72, "treatment"), (73, "outstanding"), (74, "rate"), (75, "tenure")):
        if a.cell(r, 3).value is None:
            return f"Assumptions C{r} ({what}) missing"
    return None


@check("existing loan wired into P&L, balance sheet, DSCR and fund flow")
def _(folder, path, wb):
    for sheet, row in (("Expenses", 14), ("Form_III_BalanceSheet", 9),
                       ("DSCR", 11), ("Form_VI_FundFlow", 14)):
        if sheet not in wb.sheetnames:
            return f"{sheet} missing"
        ws = wb[sheet]
        if not any(isinstance(ws.cell(row, c).value, str)
                   and "Existing Loan" in ws.cell(row, c).value
                   for c in range(2, min(ws.max_column, 70) + 1)):
            return f"{sheet} row {row} does not read 'Existing Loan'"
    return None


@check("WC sheet: labels in column B, linked to the model, loan split present")
def _(folder, path, wb):
    if "WC & CC-OD Limit" not in wb.sheetnames:
        return "sheet missing"
    ws = wb["WC & CC-OD Limit"]
    if ws.cell(6, 2).value is None and ws.cell(6, 1).value is not None:
        return "labels are in column A — the wide-gap layout is back"
    if "TERM LOAN" not in " ".join(str(ws.cell(r, 2).value or "")
                                   for r in range(1, 40)).upper():
        return "no term-loan vs working-capital split"
    if not isinstance(ws.cell(6, 3).value, str) or "Form_IV" not in ws.cell(6, 3).value:
        return "not linked to the model — dead blue inputs again"
    return None


@check("inventory schedule exactly where the industry carries stock")
def _(folder, path, wb):
    from financial_engine.industry_calc.operating_models import get_operating_model
    m = get_operating_model({"bank_loan": "manufacturing"}.get(
        folder, folder.replace("_cma", "")))
    has = "Inventory Schedule" in wb.sheetnames
    if m and m.holds_inventory and not has:
        return "industry holds stock but has no Inventory Schedule"
    if m and not m.holds_inventory and has:
        return "industry holds no stock but has an Inventory Schedule"
    return None


@check("SWOT sheet stays removed")
def _(folder, path, wb):
    return "SWOT sheet is back" if "SWOT" in wb.sheetnames else None


@check("one revenue sheet, no mirror")
def _(folder, path, wb):
    if folder == "bank_loan":
        return None if "Sales" in wb.sheetnames else "no Sales sheet"
    if "Revenue Plan" in wb.sheetnames:
        return "the Revenue Plan mirror is back"
    return None if "Revenue Build-Up" in wb.sheetnames else "no Revenue Build-Up"


@check("no formula points at a sheet that does not exist")
def _(folder, path, wb):
    names = set(wb.sheetnames)
    bad = {ref for ws in wb.worksheets for row in ws.iter_rows() for c in row
           if isinstance(c.value, str) and c.value.startswith("=")
           for ref in re.findall(r"'([^']+)'!", c.value) if ref not in names}
    return f"dangling references to {sorted(bad)}" if bad else None


@check("every chart series resolves")
def _(folder, path, wb):
    names = set(wb.sheetnames)
    bad = set()
    with zipfile.ZipFile(path) as z:
        for n in z.namelist():
            if "/charts/chart" in n:
                for m in re.findall(r"<f>([^<]+)</f>", z.read(n).decode("utf8", "ignore")):
                    if "!" in m and m.split("!")[0].strip("'") not in names:
                        bad.add(m.split("!")[0].strip("'"))
    return f"charts point at {sorted(bad)}" if bad else None


# ── model-level invariants (per operating model, not per workbook) ─────────────
# Every workbook check above is per template folder. This one is per INDUSTRY, and it
# exists because of a fault the folder checks could never have caught: the streams block
# was structurally perfect in every template, while six industries had no stream profile
# to fill it with, so their reports shipped with the block at zero. The template was fine;
# the data behind it was missing. Agriculture, textile, automobile, mining, renewable
# energy and construction were left at (0,0,0,0) when the capacity workbook had no streams
# section, and nobody went back when it grew one. A new industry added tomorrow would fall
# into exactly the same hole silently, so it is asserted here instead of remembered.
def _model_invariants() -> list[str]:
    from financial_engine.industry_calc.operating_models import _MODELS
    problems = []
    for key, m in _MODELS.items():
        if not any(m.stream_mix):
            problems.append(f"{key}: stream_mix is all zero — its reports ship an empty "
                            f"revenue-streams block")
        elif not any(m.stream_vol_per_core):
            problems.append(f"{key}: has a stream_mix but no stream_vol_per_core — the "
                            f"guard needs both and stands down without it")
        if m.margin_hint and not (0 < m.margin_hint[0] < m.margin_hint[1] < 1):
            problems.append(f"{key}: margin_hint {m.margin_hint} is not a sane band")
    return problems


def _chain_order_invariant() -> list[str]:
    """Assumptions!C18 is read by two guards with two meanings — year-1 capacity
    UTILISATION in reconcile_scale, which divides by it, and the year-1 growth INDEX in
    reconcile_working_capital, which resets it to 1.0. Order decides which is left stale.
    With scale first, a 25% ramp in that cell quadrupled the solar plant's output and the
    reset that followed never went back to it: a 5.76 MW generation on a 2.12 MW capex,
    shipped. This drives a real project through the chain and fails if the volume comes out
    divided by C18 again, so a reorder cannot quietly reintroduce it.
    """
    from database import SessionLocal
    from models.project_model import Project
    from routers.generation_router import (_stored_answers, _resolve_template,
                                           resolve_purpose, _reconcile_all)
    db = SessionLocal()
    p = next((x for x in db.query(Project).order_by(Project.id.desc())
              if x.report and _stored_answers(db, x).get("Assumptions!C16")), None)
    if p is None:
        return ["no project with a filled workbook to test the chain order against"]
    a = _stored_answers(db, p)
    pk, t = _resolve_template(resolve_purpose(p.purpose, p.financial_format), a,
                              p.industry, p.purpose)
    ramp, base = dict(a), dict(a)
    ramp["Assumptions!C16"] = base["Assumptions!C16"] = 1_000_000
    ramp["Assumptions!C18"], base["Assumptions!C18"] = 0.25, 1.0
    for d in (ramp, base):
        d.pop("_scale_capacity_before_raise", None)
    rv = _reconcile_all(ramp, p, t).get("Assumptions!C16")
    bv = _reconcile_all(base, p, t).get("Assumptions!C16")
    if rv and bv and abs(rv - bv) > max(bv * 0.01, 1):
        return [f"a year-1 ramp in C18 still changes the volume ({bv:,.0f} -> {rv:,.0f}): "
                f"reconcile_working_capital must run BEFORE reconcile_scale"]
    return []


def main() -> int:
    results = []
    for folder in FOLDERS:
        matches = glob.glob(os.path.join(TEMPLATES, folder, "*.xlsx"))
        if not matches:
            results += [(folder, n, "no workbook on disk") for n in CHECKS]
            continue
        path = matches[0]
        wb = openpyxl.load_workbook(path)
        for name, fn in CHECKS.items():
            try:
                results.append((folder, name, fn(folder, path, wb)))
            except Exception as exc:                       # a broken check is a failure
                results.append((folder, name, f"check errored: {type(exc).__name__}: {exc}"))

    width = max(len(n) for n in CHECKS) + 2
    print(f"{'invariant':<{width}}" + "".join(f"{f.replace('_cma', ''):>12}" for f in FOLDERS))
    print("-" * (width + 12 * len(FOLDERS)))
    failures = 0
    for name in CHECKS:
        line = f"{name:<{width}}"
        for folder in FOLDERS:
            problem = next(p for (f, n, p) in results if f == folder and n == name)
            line += f"{('ok' if problem is None else 'FAIL'):>12}"
            failures += problem is not None
        print(line)

    seen = set()
    detail = [(n, p) for (_, n, p) in results if p and (n, p) not in seen
              and not seen.add((n, p))]
    if detail:
        print("\ndetail:")
        for name, problem in detail:
            who = sorted({f for (f, n, p) in results if n == name and p == problem})
            print(f"  [{name}]\n     {problem}\n     affects: {who}")

    model_problems = _model_invariants()
    print(f"\nevery industry has a revenue-stream profile: "
          f"{'ok' if not model_problems else 'FAIL'}")
    for p in model_problems:
        print(f"     {p}")

    try:
        order_problems = _chain_order_invariant()
    except Exception as exc:
        order_problems = [f"check errored: {type(exc).__name__}: {exc}"]
    print(f"a year-1 ramp in C18 does not inflate the volume: "
          f"{'ok' if not order_problems else 'FAIL'}")
    for p in order_problems:
        print(f"     {p}")

    failures += len(model_problems) + len(order_problems)
    print(f"\n{'ALL INVARIANTS HOLD' if not failures else f'{failures} FAILURES'}")
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
