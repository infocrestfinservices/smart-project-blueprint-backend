"""
excel_writer.py

Writes mapped values into an EXISTING Excel workbook, in place, using openpyxl only.
It is the last step of the pipeline: given the worksheet payloads (from
financial_model_mapper.build_excel_mapping) and a cell map that says which A1 cell each
field belongs in, it opens the template, overwrites ONLY those mapped cells, and saves.

It performs NO financial calculation. It does not touch formulas, formatting, charts,
rows, columns, merges, or worksheet dimensions — it only overwrites the values of the
cells named in cell_mapping. Every other cell (including every formula and all styling)
is left exactly as loaded, so on the next spreadsheet recalculation the workbook's own
formulas recompute from the freshly-written inputs.

Mapping shapes
--------------
    worksheet_mapping : {sheet_name: {field: value}}       (from the model mapper)
    cell_mapping      : {sheet_name: {field: "A1"}}         scalar  -> one cell
                        {sheet_name: {field: ["A1","B1"]}}  series  -> one cell per item

A mapped cell that currently holds a formula IS overwritten (mapping it is the explicit
intent to replace it). Unmapped cells are never read or written, so their formulas and
formatting are untouched.

Validation is done fully BEFORE any write, so a malformed mapping raises ValueError
without leaving a partially-written output file.
"""

from __future__ import annotations

import os
import re

from openpyxl import load_workbook

# Plain A1 notation: 1-3 column letters + a positive row number. No ranges, no '$'.
_A1_RE = re.compile(r"^[A-Za-z]{1,3}[1-9][0-9]*$")


def _is_valid_a1(ref) -> bool:
    return isinstance(ref, str) and bool(_A1_RE.match(ref))


def write_excel_mapping(
    workbook_path: str,
    output_path: str,
    worksheet_mapping: dict,
    cell_mapping: dict,
) -> None:
    """Write worksheet_mapping values into workbook_path's mapped cells and save to
    output_path. Overwrites only the cells named in cell_mapping; preserves formulas,
    formatting, charts, and workbook structure. Raises ValueError on malformed input.
    """
    fn = "write_excel_mapping"

    # ── input validation ────────────────────────────────────────────────────────
    if not isinstance(workbook_path, str) or not os.path.isfile(workbook_path):
        raise ValueError(f"{fn}: workbook_path does not exist: {workbook_path!r}")
    if not isinstance(worksheet_mapping, dict):
        raise ValueError(f"{fn}: worksheet_mapping must be a dict, "
                         f"got {type(worksheet_mapping).__name__}")
    if not isinstance(cell_mapping, dict):
        raise ValueError(f"{fn}: cell_mapping must be a dict, "
                         f"got {type(cell_mapping).__name__}")

    # data_only=False keeps formulas as formulas (never replace them with cached values).
    workbook = load_workbook(workbook_path)
    sheet_names = set(workbook.sheetnames)

    # ── pass 1: validate the whole cell_mapping before writing anything ──────────
    # (build a flat list of concrete (sheet, cell, value) writes as we go)
    writes = []  # list of (sheet_name, "A1", value)
    for sheet_name, field_map in cell_mapping.items():
        if sheet_name not in sheet_names:
            raise ValueError(f"{fn}: worksheet {sheet_name!r} referenced in cell_mapping "
                             f"does not exist in the workbook (sheets: {sorted(sheet_names)}).")
        if not isinstance(field_map, dict):
            raise ValueError(f"{fn}: cell_mapping[{sheet_name!r}] must be a dict, "
                             f"got {type(field_map).__name__}")
        if sheet_name not in worksheet_mapping or not isinstance(worksheet_mapping[sheet_name], dict):
            raise ValueError(f"{fn}: worksheet_mapping is missing a dict for sheet "
                             f"{sheet_name!r} referenced by cell_mapping.")
        values_for_sheet = worksheet_mapping[sheet_name]

        for field, target in field_map.items():
            if field not in values_for_sheet:
                raise ValueError(f"{fn}: field {field!r} (cell_mapping[{sheet_name!r}]) is "
                                 f"not present in worksheet_mapping[{sheet_name!r}].")
            value = values_for_sheet[field]

            if isinstance(target, str):                       # scalar -> one cell
                if not _is_valid_a1(target):
                    raise ValueError(f"{fn}: invalid A1 cell reference {target!r} for "
                                     f"{sheet_name!r}.{field}")
                writes.append((sheet_name, target, value))

            elif isinstance(target, (list, tuple)):           # series -> one cell each
                if not isinstance(value, (list, tuple)):
                    raise ValueError(f"{fn}: {sheet_name!r}.{field} maps to {len(target)} cells "
                                     f"but its value is not a list.")
                if len(target) != len(value):
                    raise ValueError(f"{fn}: {sheet_name!r}.{field} has {len(target)} cells but "
                                     f"{len(value)} values.")
                for ref in target:
                    if not _is_valid_a1(ref):
                        raise ValueError(f"{fn}: invalid A1 cell reference {ref!r} for "
                                         f"{sheet_name!r}.{field}")
                for ref, item in zip(target, value):
                    writes.append((sheet_name, ref, item))

            else:
                raise ValueError(f"{fn}: cell_mapping[{sheet_name!r}][{field!r}] must be an A1 "
                                 f"string or a list of A1 strings, got {type(target).__name__}")

    # ── pass 2: write values only (styles/formulas of unmapped cells untouched) ──
    for sheet_name, ref, value in writes:
        # Assigning .value overwrites only the value; openpyxl leaves the cell's
        # number_format, font, fill, border, etc. in place.
        workbook[sheet_name][ref] = value

    workbook.save(output_path)
