"""
test_excel_writer_manual.py

Standalone test for excel_writer.write_excel_mapping. Builds a temporary workbook with
values, a formula, and cell formatting; writes a mapping into it; reloads; and verifies
mapped values changed, formulas and formatting are preserved, sheet names are intact,
and the workbook reloads cleanly. Also checks malformed mappings raise ValueError.

Run from backend/:
    python excel_writer/test_excel_writer_manual.py
"""

import importlib.util
import os
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

# excel_writer.py lives inside a directory also named 'excel_writer', so a normal
# `import excel_writer.excel_writer` is ambiguous (the module shadows the package).
# Load it directly by file path — unambiguous regardless of sys.path.
_HERE = Path(__file__).resolve().parent
_spec = importlib.util.spec_from_file_location("_excel_writer_under_test", _HERE / "excel_writer.py")
_mod = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_mod)
write_excel_mapping = _mod.write_excel_mapping

fails = []


def ok(cond, label):
    print(f"   {'OK ' if cond else 'FAIL'}  {label}")
    if not cond:
        fails.append(label)


def build_template(path):
    """A 2-sheet workbook with values, a formula, and formatting to preserve."""
    wb = Workbook()
    s1 = wb.active
    s1.title = "Sheet1"
    s1["A1"] = 10
    s1["A2"] = 20
    s1["A3"] = "=A1+A2"          # formula, NOT mapped -> must survive untouched
    s1["D1"] = "=A1*2"           # formula, IS mapped -> must be overwritten with a value
    # B1 is a mapped input cell that carries formatting we must preserve
    s1["B1"] = 0
    s1["B1"].font = Font(bold=True, italic=True)
    s1["B1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    s1["B1"].number_format = "#,##0.00"
    # B2:B4 will receive a 3-value series
    for r in (2, 3, 4):
        s1[f"B{r}"] = 0
    s2 = wb.create_sheet("Sheet2")
    s2["C1"] = 0
    wb.save(path)


def main():
    tmp = tempfile.mkdtemp(prefix="xlwriter_")
    src = os.path.join(tmp, "template.xlsx")
    out = os.path.join(tmp, "output.xlsx")
    build_template(src)

    worksheet_mapping = {
        "Sheet1": {"revenue": 999.5, "series": [111, 222, 333], "override": 500},
        "Sheet2": {"irr": 0.1523},
    }
    cell_mapping = {
        "Sheet1": {"revenue": "B1", "series": ["B2", "B3", "B4"], "override": "D1"},
        "Sheet2": {"irr": "C1"},
    }

    print("=" * 74)
    print("EXCEL WRITER")
    print("=" * 74)

    # 4. apply
    write_excel_mapping(src, out, worksheet_mapping, cell_mapping)

    # 5. reload (data_only=False so formula strings are visible)
    wb = load_workbook(out)
    s1, s2 = wb["Sheet1"], wb["Sheet2"]

    print("\n6a. Mapped values updated correctly:")
    ok(s1["B1"].value == 999.5, "Sheet1!B1 (scalar) == 999.5")
    ok([s1["B2"].value, s1["B3"].value, s1["B4"].value] == [111, 222, 333],
       "Sheet1!B2:B4 (series) == [111,222,333]")
    ok(s2["C1"].value == 0.1523, "Sheet2!C1 == 0.1523")
    ok(s1["D1"].value == 500, "Sheet1!D1 (mapped formula) overwritten with 500")

    print("\n6b. Formulas unchanged (unmapped formula preserved):")
    ok(s1["A3"].value == "=A1+A2", "Sheet1!A3 still '=A1+A2'")
    ok(s1["A1"].value == 10 and s1["A2"].value == 20, "Sheet1!A1,A2 untouched (10,20)")

    print("\n6c. Worksheet names unchanged:")
    ok(wb.sheetnames == ["Sheet1", "Sheet2"], f"sheetnames == {wb.sheetnames}")

    print("\n6d. Workbook loads successfully:")
    ok(wb is not None, "reload returned a workbook")

    print("\n6e. Formatting preserved on the mapped cell B1:")
    ok(s1["B1"].font.bold is True, "B1 font.bold preserved")
    ok(s1["B1"].font.italic is True, "B1 font.italic preserved")
    ok(s1["B1"].fill.fill_type == "solid", "B1 fill.fill_type == 'solid'")
    ok(str(s1["B1"].fill.fgColor.rgb).endswith("FFFF00"), "B1 fill colour preserved (…FFFF00)")
    ok(s1["B1"].number_format == "#,##0.00", "B1 number_format == '#,##0.00'")

    print("\n7. Malformed inputs raise ValueError:")
    cases = [
        ("missing workbook file",
         lambda: write_excel_mapping(os.path.join(tmp, "nope.xlsx"), out,
                                     worksheet_mapping, cell_mapping)),
        ("worksheet_mapping not a dict",
         lambda: write_excel_mapping(src, out, [], cell_mapping)),
        ("cell_mapping not a dict",
         lambda: write_excel_mapping(src, out, worksheet_mapping, [])),
        ("sheet not in workbook",
         lambda: write_excel_mapping(src, out, {"Ghost": {"x": 1}}, {"Ghost": {"x": "A1"}})),
        ("invalid A1 address",
         lambda: write_excel_mapping(src, out, {"Sheet1": {"x": 1}}, {"Sheet1": {"x": "A1:B2"}})),
        ("field missing in worksheet_mapping",
         lambda: write_excel_mapping(src, out, {"Sheet1": {}}, {"Sheet1": {"x": "A1"}})),
        ("series length mismatch",
         lambda: write_excel_mapping(src, out, {"Sheet1": {"s": [1, 2]}},
                                     {"Sheet1": {"s": ["A1", "B1", "C1"]}})),
    ]
    for label, fn in cases:
        try:
            fn()
            ok(False, f"{label} -> should have raised")
        except ValueError as e:
            ok(True, f"{label} -> ValueError: {str(e).split(':',1)[1].strip()[:34]}")

    print("\n" + "=" * 74)
    if fails:
        print(f"FAILED — {len(fails)} check(s): {fails}")
        raise AssertionError("excel_writer did not behave as specified")
    print("PASSED — mapped values written, formulas & formatting preserved, sheet")
    print("structure intact, workbook reloads, malformed mappings rejected.")
    print("=" * 74)
    return 0


if __name__ == "__main__":
    sys.exit(main())
