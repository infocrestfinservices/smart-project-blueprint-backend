"""
engine_test/template_definition.py

TEST-ONLY scaffolding for validating the new financial engine end to end.

Nothing here touches production. It provides three things the validation route needs:

  1. TEMPLATE_DEFINITION — the {sheet: {field: cell}} map from the engine's output
     fields to cells in a VALIDATION workbook.
  2. ensure_template() — builds that validation workbook on demand under
     templates/engine_test/ (separate from every production template).
  3. DEMO_ASSUMPTIONS — a complete, validated 44-field assumptions set so the route
     is callable with no body.

WHY A SEPARATE TEMPLATE
-----------------------
The production workbooks (e.g. CMA_Dashboard_Premium.xlsx) are FORMULA-DRIVEN: you
write ~59 input cells and Excel computes everything else (Profit!N21 = "=N19-N20",
Ratios!C15 = "=DSCR!C14", ...). The new pipeline writes COMPUTED OUTPUTS, so pointing
it at a production template would overwrite those formulas with static values and
destroy the live model.

This validation workbook is therefore a deliberate VALUE SINK: labels and headers
only, with every target cell empty and formula-free. Writing engine outputs into it
clobbers nothing, which makes it a clean harness for proving the engine produces a
populated workbook end to end.

Assumption translation is intentionally NOT implemented here: the route accepts the
engine's native assumptions vocabulary directly (that IS the engine's input contract),
so no translator is required to validate the engine. Translating a real
project/BusinessProfile into these fields is a separate concern, and part of the
later migration decision.
"""

from __future__ import annotations

import os

BACKEND_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_DIR = os.path.join(BACKEND_DIR, "templates", "engine_test")
TEMPLATE_PATH = os.path.join(TEMPLATE_DIR, "engine_test_model.xlsx")

YEARS = ["B", "C", "D", "E", "F"]            # Year 1..5
CF_COLS = ["B", "C", "D", "E", "F", "G"]     # Year 0..5 (cash flow carries Year 0)


def _row(cols, row):
    return [f"{c}{row}" for c in cols]


# field -> cell(s), per sheet. Field names are exactly the keys the engine's
# worksheet payloads (build_excel_mapping) expose.
TEMPLATE_DEFINITION = {
    "Dashboard": {
        "Revenue": _row(YEARS, 5),
        "EBITDA": _row(YEARS, 6),
        "PAT": _row(YEARS, 7),
        "DSCR": _row(YEARS, 8),
        "IRR": "B10",
        "NPV": "B11",
    },
    "ProfitLoss": {
        "ebitda": _row(YEARS, 5),
        "ebit": _row(YEARS, 6),
        "pbt": _row(YEARS, 7),
        "income_tax": _row(YEARS, 8),
        "pat": _row(YEARS, 9),
        "cash_accrual": _row(YEARS, 10),
    },
    "BalanceSheet": {
        "net_worth": _row(YEARS, 5),
        "term_loan_closing": _row(YEARS, 6),
        "total_liabilities": _row(YEARS, 7),
        "net_fixed_assets": _row(YEARS, 8),
        "inventory": _row(YEARS, 9),
        "debtors": _row(YEARS, 10),
        "cash_balancing_figure": _row(YEARS, 11),
        "total_assets": _row(YEARS, 12),
    },
    "CashFlow": {
        "operating_cash_flow": _row(CF_COLS, 5),
        "investing_cash_flow": _row(CF_COLS, 6),
        "financing_cash_flow": _row(CF_COLS, 7),
        "net_cash_flow": _row(CF_COLS, 8),
        "cash_flow_series": _row(CF_COLS, 9),
    },
    "Ratios": {
        "current_ratio": _row(YEARS, 5),
        "debt_equity": _row(YEARS, 6),
        "interest_coverage": _row(YEARS, 7),
        "net_profit_margin": _row(YEARS, 8),
        "dscr": _row(YEARS, 9),
        "average_dscr": "B11",
    },
    "LoanSchedule": {
        "opening_balance": _row(YEARS, 5),
        "interest": _row(YEARS, 6),
        "principal": _row(YEARS, 7),
        "closing_balance": _row(YEARS, 8),
    },
    "WorkingCapital": {
        "net_sales": _row(YEARS, 5),
        "total_current_assets": _row(YEARS, 6),
        "total_current_liabilities": _row(YEARS, 7),
        "working_capital_gap": _row(YEARS, 8),
        "mpbf": _row(YEARS, 9),
        "wc_interest_annual": _row(YEARS, 10),
    },
    "Depreciation": {
        "annual_depreciation_series": _row(YEARS, 5),
        "monthly_depreciation": "B7",
        "annual_depreciation": "B8",
    },
}

# Row labels for the validation workbook (column A), so the output is readable.
_LABELS = {
    "Dashboard": {4: "Year", 5: "Revenue", 6: "EBITDA", 7: "PAT", 8: "DSCR",
                  10: "IRR", 11: "NPV"},
    "ProfitLoss": {4: "Year", 5: "EBITDA", 6: "EBIT", 7: "PBT", 8: "Income Tax",
                   9: "PAT", 10: "Cash Accrual"},
    "BalanceSheet": {4: "Year", 5: "Net Worth", 6: "Term Loan (closing)",
                     7: "Total Liabilities", 8: "Net Fixed Assets", 9: "Inventory",
                     10: "Debtors", 11: "Cash", 12: "TOTAL ASSETS"},
    "CashFlow": {4: "Year", 5: "Operating", 6: "Investing", 7: "Financing",
                 8: "Net Cash Flow", 9: "Cash Flow Series"},
    "Ratios": {4: "Year", 5: "Current Ratio", 6: "Debt-Equity", 7: "Interest Coverage",
               8: "Net Profit Margin", 9: "DSCR", 11: "Average DSCR"},
    "LoanSchedule": {4: "Year", 5: "Opening Balance", 6: "Interest", 7: "Principal",
                     8: "Closing Balance"},
    "WorkingCapital": {4: "Year", 5: "Net Sales", 6: "Current Assets",
                       7: "Current Liabilities", 8: "Working Capital Gap", 9: "MPBF",
                       10: "WC Interest"},
    "Depreciation": {4: "Year", 5: "Annual Depreciation", 7: "Monthly Depreciation",
                     8: "Annual Depreciation (total)"},
}


def ensure_template(force: bool = False) -> str:
    """Create the validation workbook if it does not exist and return its path.

    Deliberately a VALUE SINK: labels + year headers only, every engine target cell
    left empty and formula-free, so the writer has nothing to clobber. Uses openpyxl
    only to author this test asset — the pipeline itself never touches openpyxl
    outside the writer module.
    """
    if os.path.isfile(TEMPLATE_PATH) and not force:
        return TEMPLATE_PATH

    from openpyxl import Workbook
    from openpyxl.styles import Font

    os.makedirs(TEMPLATE_DIR, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    for sheet in TEMPLATE_DEFINITION:
        ws = wb.create_sheet(title=sheet)
        ws["A1"] = f"ENGINE VALIDATION — {sheet}"
        ws["A1"].font = Font(bold=True, size=13)
        ws["A2"] = "Test-only workbook. Values written by the Python financial engine."
        cols = CF_COLS if sheet == "CashFlow" else YEARS
        start = 0 if sheet == "CashFlow" else 1
        for i, col in enumerate(cols):
            ws[f"{col}4"] = f"Year {i + start}"
            ws[f"{col}4"].font = Font(bold=True)
            ws.column_dimensions[col].width = 16
        for row, label in _LABELS.get(sheet, {}).items():
            if row == 4:
                ws["A4"] = label
                ws["A4"].font = Font(bold=True)
                continue
            ws[f"A{row}"] = label
        ws.column_dimensions["A"].width = 28
    wb.save(TEMPLATE_PATH)
    return TEMPLATE_PATH


# A complete, engine-valid 44-field assumptions set (5-year loan, zero moratorium),
# so the validation route is callable with no request body.
DEMO_ASSUMPTIONS = {
    "name_of_unit": "Engine Validation Demo Pvt Ltd", "constitution": "Private Limited",
    "line_of_activity": "Precision components", "industry_type": "Manufacturing",
    "term_loan_amount": 4_000_000, "promoters_capital": 2_000_000,
    "interest_rate_term_loan": 0.11, "term_loan_tenure_months": 60,
    "moratorium_months": 0, "interest_rate_wc": 0.12, "income_tax_rate": 0.25,
    "installed_capacity": 300_000,
    "capacity_utilisation_y1_y5": [0.6, 0.7, 0.75, 0.8, 0.85],
    "monthly_seasonality_weights": [1] * 12,
    "selling_price_y1": 100, "selling_price_escalation": 0.05,
    "cost1_per_unit_y1": 40, "cost1_escalation": 0.05,
    "cost2_per_unit_y1": 15, "cost2_escalation": 0.05,
    "other_variable_cost_y1": 5, "other_variable_escalation": 0.05,
    "wages_monthly_y1": 150_000, "wages_escalation": 0.08,
    "factory_overheads_monthly_y1": 80_000, "factory_oh_escalation": 0.06,
    "repairs_maintenance_monthly_y1": 20_000, "rm_escalation": 0.06,
    "admin_expenses_monthly_y1": 40_000, "admin_escalation": 0.06,
    "selling_distribution": 0.02,
    "land_cost": 500_000, "building_cost": 2_000_000, "building_dep_rate": 0.10,
    "plant_machinery_cost": 4_000_000, "plant_machinery_dep_rate": 0.15,
    "furniture_other_cost": 500_000, "furniture_dep_rate": 0.15,
    "raw_material_holding_days": 30, "finished_goods_holding_days": 15,
    "receivables_days": 45, "payables_days": 30,
    "min_cash_balance": 100_000, "wc_margin_pct": 0.25,
    "discount_rate": 0.12,
}
