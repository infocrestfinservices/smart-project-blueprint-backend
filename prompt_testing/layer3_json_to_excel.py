"""
Layer 3 — feed an Assumption Architect JSON output into the CMA Excel
template, then force LibreOffice to recalculate all formulas.

This is STILL isolated to prompt_testing/ — it does not touch backend/ or
frontend/. Once this is proven to work reliably, the same mapping logic
gets ported into the real backend service.

Usage:
    python layer3_json_to_excel.py <path_to_assumption_json> <output_xlsx_name>

Example:
    python layer3_json_to_excel.py outputs/manufacturing_bank_loan_03.json manufacturing_report.xlsx

Requires:
    pip install openpyxl --break-system-packages
    LibreOffice installed, with soffice.exe reachable (adjust SOFFICE_PATH
    below if it's not on PATH).
"""

import json
import os
import platform
import re
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

# ---- CONFIG — adjust these two paths for your machine ----
TEMPLATE_PATH = Path(__file__).parent.parent / "templates" / "bank_loan" / "CMA_Dashboard_Premium.xlsx"
SOFFICE_PATH = r"C:\Program Files\LibreOffice\program\soffice.exe"
# r"C:\Program Files\LibreOffice\program\soffice.exe"
# ------------------------------------------------------------

MACRO_FILENAME = "Module1.xba"
RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>
"""


def get_macro_dir() -> str:
    system = platform.system()
    if system == "Windows":
        return os.path.expandvars(r"%APPDATA%\LibreOffice\4\user\basic\Standard")
    if system == "Darwin":
        return os.path.expanduser("~/Library/Application Support/LibreOffice/4/user/basic/Standard")
    return os.path.expanduser("~/.config/libreoffice/4/user/basic/Standard")


def setup_libreoffice_macro() -> bool:
    """Install a small Basic macro into the user's LibreOffice profile that
    forces a full recalculation and saves the file. Only needs to run once
    per machine, but it's cheap to re-check every run."""
    macro_dir = get_macro_dir()
    macro_file = os.path.join(macro_dir, MACRO_FILENAME)

    if os.path.exists(macro_file) and "RecalculateAndSave" in Path(macro_file).read_text(encoding="utf-8", errors="ignore"):
        return True

    if not os.path.exists(macro_dir):
        # touching soffice once creates the user profile directories
        subprocess.run([SOFFICE_PATH, "--headless", "--terminate_after_init"],
                        capture_output=True, timeout=20)
        os.makedirs(macro_dir, exist_ok=True)

    try:
        Path(macro_file).write_text(RECALCULATE_MACRO, encoding="utf-8")
        return True
    except Exception as e:
        print(f"[error] Could not install LibreOffice macro: {e}")
        return False

SCHEMA_PATH = Path(__file__).parent / "schemas" / "assumption_schema.json"
OUTPUT_DIR = Path(__file__).parent / "generated_reports"


def load_schema() -> dict:
    with open(SCHEMA_PATH, encoding="utf-8") as f:
        return json.load(f)


def parse_cell_ref(cell_ref: str):
    """'Assumptions!C25' -> ('Assumptions', 'C25', None)
       'Assumptions!C18:G18' -> ('Assumptions', 'C18:G18', (start_col,start_row,end_col,end_row))"""
    sheet, ref = cell_ref.split("!", 1)
    if ":" in ref:
        bounds = range_boundaries(ref)  # (min_col, min_row, max_col, max_row)
        return sheet, ref, bounds
    return sheet, ref, None


def fill_assumptions(wb, assumptions: dict, schema: dict) -> list:
    """Write every field from the JSON into its mapped cell(s).
    Returns a list of warnings (fields present in JSON but not in schema,
    or vice versa)."""
    warnings = []
    schema_keys = {k for k in schema if not k.startswith("_")}
    json_keys = {k for k in assumptions if not k.startswith("_")}

    missing_in_json = schema_keys - json_keys
    if missing_in_json:
        warnings.append(f"Fields in schema but missing from JSON (will be skipped): {sorted(missing_in_json)}")

    extra_in_json = json_keys - schema_keys
    if extra_in_json:
        warnings.append(f"Fields in JSON but not in schema (ignored): {sorted(extra_in_json)}")

    for key, value in assumptions.items():
        if key.startswith("_"):
            continue  # e.g. _assumptions_notes — not written to a cell
        if key not in schema:
            continue  # already warned above

        cell_ref = schema[key]["cell"]
        sheet_name, ref, bounds = parse_cell_ref(cell_ref)
        ws = wb[sheet_name]

        if bounds is None:
            # single cell
            ws[ref] = value
        else:
            # array spread across a range, e.g. C18:G18 (row) or a column range
            min_col, min_row, max_col, max_row = bounds
            if not isinstance(value, list):
                warnings.append(f"{key}: schema expects an array for {ref} but got {type(value).__name__}; skipped")
                continue

            if min_row == max_row:
                # horizontal spread
                width = max_col - min_col + 1
                if len(value) != width:
                    warnings.append(f"{key}: expected {width} values for {ref}, got {len(value)}; writing what's given")
                for i, v in enumerate(value):
                    if min_col + i > max_col:
                        break
                    ws.cell(row=min_row, column=min_col + i, value=v)
            else:
                # vertical spread
                height = max_row - min_row + 1
                if len(value) != height:
                    warnings.append(f"{key}: expected {height} values for {ref}, got {len(value)}; writing what's given")
                for i, v in enumerate(value):
                    if min_row + i > max_row:
                        break
                    ws.cell(row=min_row + i, column=min_col, value=v)

    return warnings


def recalculate_with_libreoffice(filepath: Path) -> bool:
    """Force LibreOffice to open, recalculate ALL formulas (not just the
    ones convert-to happens to touch), and re-save the file in place.

    NOTE: a plain `soffice --convert-to xlsx` does NOT reliably force a
    full recalculation — LibreOffice may just re-export cached/empty
    values. This uses an installed macro (ThisComponent.calculateAll())
    which is the approach proven to actually recompute every formula."""
    if shutil.which(SOFFICE_PATH) is None and not Path(SOFFICE_PATH).exists():
        print(f"[warning] Could not find '{SOFFICE_PATH}' — skipping recalculation.")
        print("          Install LibreOffice, or set SOFFICE_PATH at the top of this script.")
        return False

    if not setup_libreoffice_macro():
        return False

    abs_path = str(filepath.resolve())
    cmd = [
        SOFFICE_PATH,
        "--headless",
        "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        abs_path,
    ]
    print("Running:", " ".join(cmd))
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
    if result.returncode != 0:
        print("[error] LibreOffice recalculation failed:")
        print(result.stdout)
        print(result.stderr)
        return False
    return True


def check_for_errors(filepath: Path) -> list:
    """Scan every cell for Excel error strings after recalculation."""
    ERROR_TOKENS = {"#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"}
    wb = load_workbook(filepath, data_only=True)
    found = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in ERROR_TOKENS:
                    found.append(f"{sheet}!{cell.coordinate} = {cell.value}")
    return found


def main():
    if len(sys.argv) < 3:
        print("Usage: python layer3_json_to_excel.py <assumption_json_path> <output_xlsx_name>")
        sys.exit(1)

    json_path = Path(sys.argv[1])
    output_name = sys.argv[2]

    if not TEMPLATE_PATH.exists():
        print(f"[error] Template not found at {TEMPLATE_PATH}")
        print("        Copy CMA_Report_Template_MultiIndustry.xlsx into prompt_testing/")
        print("        or update TEMPLATE_PATH at the top of this script.")
        sys.exit(1)

    with open(json_path, encoding="utf-8") as f:
        assumptions = json.load(f)

    schema = load_schema()
    wb = load_workbook(TEMPLATE_PATH)

    warnings = fill_assumptions(wb, assumptions, schema)

    OUTPUT_DIR.mkdir(exist_ok=True)
    output_path = OUTPUT_DIR / output_name
    wb.save(output_path)
    print(f"Saved (pre-recalc) -> {output_path}")

    if warnings:
        print("\nWARNINGS:")
        for w in warnings:
            print(" -", w)

    print("\nRecalculating with LibreOffice...")
    ok = recalculate_with_libreoffice(output_path)

    if ok:
        print("Recalculation done. Checking for formula errors...")
        errors = check_for_errors(output_path)
        if errors:
            print(f"\n[!!] {len(errors)} ERROR CELLS FOUND:")
            for e in errors[:20]:
                print(" -", e)
            if len(errors) > 20:
                print(f"   ...and {len(errors) - 20} more")
        else:
            print("No error cells found. Report looks structurally sound.")
    else:
        print("Skipped error-check since recalculation did not run.")

    print(f"\nFinal file: {output_path}")


if __name__ == "__main__":
    main()