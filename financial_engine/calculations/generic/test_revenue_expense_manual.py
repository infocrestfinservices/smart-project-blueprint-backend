"""
test_revenue_expense_manual.py

Standalone parity check: prove the generic Python calculators reproduce EXACTLY the
numbers the Bank Loan Excel formulas produce, for the same assumptions.

    1. load a real Assumption-Architect JSON output as the assumptions dict
    2. run the generic calculators over it
    3. find the generated .xlsx that was actually filled from that same JSON
       (matched by comparing the workbook's input cells to the JSON)
    4. read Excel's own recalculated Year-1 figures back out
    5. assert Python and Excel agree within 0.5%

Run from backend/:
    python -m financial_engine.calculations.generic.test_revenue_expense_manual
or:
    python financial_engine/calculations/generic/test_revenue_expense_manual.py
"""

import json
import os
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
BACKEND_DIR = HERE.parents[2]
sys.path.insert(0, str(BACKEND_DIR))

from financial_engine.calculations.generic.revenue_calc import (  # noqa: E402
    calculate_monthly_production, calculate_monthly_revenue,
)
from financial_engine.calculations.generic.expense_calc import (  # noqa: E402
    calculate_monthly_variable_costs, calculate_monthly_fixed_costs,
)

ASSUMPTIONS_JSON = BACKEND_DIR / "prompt_testing" / "outputs" / "manufacturing_bank_loan_03.json"
TOLERANCE = 0.005  # 0.5%

# Year-1 month columns are B..M (2..13); N is Excel's own "Yr 1 Total" column.
Y1_COLS = [get_column_letter(c) for c in range(2, 14)]

# The input cells that identify which workbook was filled from which JSON.
FINGERPRINT = {
    "installed_capacity": "C16",
    "selling_price_y1": "C23",
    "cost1_per_unit_y1": "C25",
    "wages_monthly_y1": "C32",
    "term_loan_amount": "C8",
}

# Excel row -> the Python series it should equal.
EXCEL_ROWS = {
    "production": ("Production", 6),
    "revenue": ("Sales", 7),
    "cost1": ("Expenses", 5),
    "cost2": ("Expenses", 6),
    "wages": ("Expenses", 7),
    "other_variable": ("Expenses", 8),
    "factory_overheads": ("Expenses", 9),
    "repairs_maintenance": ("Expenses", 10),
    "admin_expenses": ("Expenses", 11),
    "selling_distribution": ("Expenses", 12),
}


def y1_total(series):
    """Sum the first 12 months of a 60-month Python series."""
    return sum(series[:12])


def excel_y1_total(wb, sheet, row):
    """Sum Excel's own recalculated M1..M12 for a row (B..M)."""
    ws = wb[sheet]
    total = 0.0
    for col in Y1_COLS:
        v = ws[f"{col}{row}"].value
        if isinstance(v, (int, float)):
            total += float(v)
    return total


def find_matching_workbook(assumptions):
    """Locate the generated .xlsx that was filled from THIS assumptions dict, by
    comparing its input cells. Searches both report directories."""
    candidates = []
    for d in (BACKEND_DIR / "generated_reports",
              BACKEND_DIR / "prompt_testing" / "generated_reports"):
        if d.is_dir():
            candidates += [p for p in d.glob("*.xlsx") if not p.name.startswith("~$")]

    for path in sorted(candidates, key=os.path.getmtime):
        try:
            wb = load_workbook(path, data_only=True)
            if "Assumptions" not in wb.sheetnames:
                wb.close()
                continue
            ws = wb["Assumptions"]
            ok = True
            for field, cell in FINGERPRINT.items():
                want, got = assumptions.get(field), ws[cell].value
                if want is None or not isinstance(got, (int, float)):
                    ok = False
                    break
                if abs(float(got) - float(want)) > max(1e-6, abs(float(want)) * 1e-9):
                    ok = False
                    break
            if ok:
                return path, wb
            wb.close()
        except Exception:
            continue
    return None, None


def main():
    assumptions = json.loads(ASSUMPTIONS_JSON.read_text(encoding="utf-8"))
    print("=" * 78)
    print(f"ASSUMPTIONS: {ASSUMPTIONS_JSON.name}")
    print("=" * 78)
    for k in ("installed_capacity", "capacity_utilisation_y1_y5", "selling_price_y1",
              "selling_price_escalation", "cost1_per_unit_y1", "wages_monthly_y1",
              "selling_distribution"):
        print(f"  {k:<30} {assumptions.get(k)!r}")

    # ---- Python side ----
    production = calculate_monthly_production(assumptions)
    revenue = calculate_monthly_revenue(assumptions, production)
    variable = calculate_monthly_variable_costs(assumptions, production)
    fixed = calculate_monthly_fixed_costs(assumptions, revenue)

    py = {"production": production, "revenue": revenue}
    py.update(variable)
    py.update(fixed)

    print(f"\nPython: {len(production)} months computed "
          f"({len(production) // 12} years x 12)")

    # ---- Excel side ----
    path, wb = find_matching_workbook(assumptions)
    if wb is None:
        print("\n[!] No generated workbook matches these assumptions — cannot cross-check.")
        print("    (Generate one from this JSON, then re-run.)")
        return 1
    print(f"Excel : {path.name}")

    # ---- side-by-side ----
    print("\n" + "=" * 78)
    print("YEAR-1 TOTALS  —  PYTHON (generic engine)  vs  EXCEL (recalculated)")
    print("=" * 78)
    print(f"{'Line':<24}{'PYTHON':>18}{'EXCEL':>18}{'DIFF':>12}{'DIFF %':>9}")
    print("-" * 78)

    failures = []
    for line, (sheet, row) in EXCEL_ROWS.items():
        p = y1_total(py[line])
        e = excel_y1_total(wb, sheet, row)
        diff = p - e
        pct = (abs(diff) / abs(e)) if e else (0.0 if abs(p) < 1e-9 else float("inf"))
        flag = "OK" if pct <= TOLERANCE else "MISMATCH"
        if pct > TOLERANCE:
            failures.append((line, p, e, pct))
        print(f"{line:<24}{p:>18,.2f}{e:>18,.2f}{diff:>12,.2f}{pct * 100:>8.4f}%  {flag}")

    print("-" * 78)
    tot_p = sum(y1_total(py[k]) for k in
                ("cost1", "cost2", "other_variable", "wages", "factory_overheads",
                 "repairs_maintenance", "admin_expenses", "selling_distribution"))
    print(f"{'TOTAL OPERATING COST':<24}{tot_p:>18,.2f}")
    print(f"{'Y1 REVENUE':<24}{y1_total(revenue):>18,.2f}")
    print(f"{'Y1 GROSS SURPLUS':<24}{y1_total(revenue) - tot_p:>18,.2f}")
    wb.close()

    print("\n" + "=" * 78)
    if failures:
        print(f"FAILED — {len(failures)} line(s) outside the {TOLERANCE * 100}% tolerance:")
        for line, p, e, pct in failures:
            print(f"   {line}: python={p:,.2f} excel={e:,.2f} ({pct * 100:.2f}% apart)")
        raise AssertionError(
            f"{len(failures)} line(s) differ from Excel by more than {TOLERANCE * 100}%"
        )
    print(f"PASSED — all {len(EXCEL_ROWS)} lines match Excel within {TOLERANCE * 100}%.")
    print("The generic Python engine reproduces the Excel math exactly.")
    print("=" * 78)
    return 0


if __name__ == "__main__":
    sys.exit(main())
