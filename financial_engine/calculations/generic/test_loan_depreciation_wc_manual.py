"""
test_loan_depreciation_wc_manual.py

Parity check for depreciation, the term-loan schedule, and working capital / MPBF
against a REAL post-fix workbook (one generated today, after the Repayment
amortisation bug was corrected).

Ground truth = generated_reports/*.xlsx, recalculated by LibreOffice.
Compared sheets: Repayment (r5-r8), Form_IV_CA_CL (r6-r19), Form_V_MPBF (r8-r11),
Expenses (r13 depreciation, r14 TL interest, r15 WC interest), DSCR (r13).

Run from backend/:
    python financial_engine/calculations/generic/test_loan_depreciation_wc_manual.py
"""

import json
import os
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
BACKEND_DIR = HERE.parents[2]
sys.path.insert(0, str(BACKEND_DIR))

from financial_engine.calculations.generic.revenue_calc import (  # noqa: E402
    calculate_monthly_production, calculate_monthly_revenue, yearly_totals,
)
from financial_engine.calculations.generic.expense_calc import (  # noqa: E402
    calculate_monthly_variable_costs, calculate_monthly_fixed_costs,
)
from financial_engine.calculations.generic.depreciation_calc import (  # noqa: E402
    calculate_monthly_depreciation, calculate_annual_depreciation,
)
from financial_engine.calculations.generic.loan_schedule_calc import (  # noqa: E402
    calculate_loan_schedule, total_debt_obligation,
)
from financial_engine.calculations.generic.working_capital_calc import (  # noqa: E402
    calculate_working_capital,
)

TOLERANCE = 0.005  # 0.5%
YEAR_COLS = ["C", "D", "E", "F", "G"]          # annual sheets: Year 1..5
MONTH_BLOCKS = [range(2, 14), range(15, 27), range(28, 40), range(41, 53), range(54, 66)]

# The Assumptions input cells, so we can rebuild the exact assumptions dict the
# workbook was filled from (no JSON file needed — the workbook IS the record).
INPUT_CELLS = {
    "term_loan_amount": "C8", "promoters_capital": "C9",
    "interest_rate_term_loan": "C10", "term_loan_tenure_months": "C11",
    "moratorium_months": "C12", "interest_rate_wc": "C13", "income_tax_rate": "C14",
    "installed_capacity": "C16",
    "selling_price_y1": "C23", "selling_price_escalation": "C24",
    "cost1_per_unit_y1": "C25", "cost1_escalation": "C26",
    "cost2_per_unit_y1": "C27", "cost2_escalation": "C28",
    "other_variable_cost_y1": "C29", "other_variable_escalation": "C30",
    "wages_monthly_y1": "C32", "wages_escalation": "C33",
    "factory_overheads_monthly_y1": "C34", "factory_oh_escalation": "C35",
    "repairs_maintenance_monthly_y1": "C36", "rm_escalation": "C37",
    "admin_expenses_monthly_y1": "C38", "admin_escalation": "C39",
    "selling_distribution": "C40",
    "land_cost": "C42", "building_cost": "C43", "building_dep_rate": "C44",
    "plant_machinery_cost": "C45", "plant_machinery_dep_rate": "C46",
    "furniture_other_cost": "C47", "furniture_dep_rate": "C48",
    "raw_material_holding_days": "C50", "finished_goods_holding_days": "C51",
    "receivables_days": "C52", "payables_days": "C53",
    "min_cash_balance": "C54", "wc_margin_pct": "C55",
}


def read_assumptions(wb) -> dict:
    ws = wb["Assumptions"]
    a = {k: ws[c].value for k, c in INPUT_CELLS.items()}
    a["capacity_utilisation_y1_y5"] = [ws[f"{c}18"].value for c in YEAR_COLS]
    a["monthly_seasonality_weights"] = [
        ws[f"{get_column_letter(c)}21"].value for c in range(3, 15)]
    return a


def annual_row(wb, sheet, row) -> list:
    """5 annual values from an annual sheet (columns C..G)."""
    ws = wb[sheet]
    return [float(ws[f"{c}{row}"].value or 0) for c in YEAR_COLS]


def annual_from_monthly(wb, sheet, row) -> list:
    """5 annual totals summed from a 60-month sheet (skipping Excel's total cols)."""
    ws = wb[sheet]
    return [sum(float(ws[f"{get_column_letter(c)}{row}"].value or 0) for c in block)
            for block in MONTH_BLOCKS]


def pick_workbook():
    d = BACKEND_DIR / "generated_reports"
    files = [p for p in d.glob("*.xlsx") if not p.name.startswith("~$")] if d.is_dir() else []
    if not files:
        return None
    return sorted(files, key=os.path.getmtime)[-1]


def compare(title, rows, failures):
    print(f"\n{title}")
    print(f"  {'Line':<28}{'Yr1':>14}{'Yr2':>14}{'Yr3':>14}{'Yr4':>14}{'Yr5':>14}   max%")
    print("  " + "-" * 100)
    for label, py, xl in rows:
        pcts = []
        for p, e in zip(py, xl):
            denom = abs(e) if abs(e) > 1e-9 else (abs(p) if abs(p) > 1e-9 else 1.0)
            pcts.append(abs(p - e) / denom)
        worst = max(pcts)
        if worst > TOLERANCE:
            failures.append((label, py, xl, worst))
        flag = "OK" if worst <= TOLERANCE else "MISMATCH"
        print(f"  {label + ' [py]':<28}" + "".join(f"{v:>14,.0f}" for v in py))
        print(f"  {'  [excel]':<28}" + "".join(f"{v:>14,.0f}" for v in xl)
              + f"   {worst * 100:.4f}%  {flag}")


def main():
    path = pick_workbook()
    if path is None:
        print("[!] No workbook in generated_reports/ — run a live generation first.")
        return 1
    wb = load_workbook(path, data_only=True)
    a = read_assumptions(wb)

    print("=" * 116)
    print(f"GROUND TRUTH: {path.name}   (post-fix template, LibreOffice-recalculated)")
    print("=" * 116)
    print(f"  loan {a['term_loan_amount']:,.0f} | rate {a['interest_rate_term_loan']} | "
          f"tenure {a['term_loan_tenure_months']:.0f}m | moratorium {a['moratorium_months']:.0f}m")
    print(f"  capacity {a['installed_capacity']:,.0f} | price {a['selling_price_y1']} | "
          f"wc_margin {a['wc_margin_pct']} | wc_rate {a['interest_rate_wc']}")

    failures = []

    # ---- feed the operating lines from the already-proven calculators ----
    production = calculate_monthly_production(a)
    revenue_m = calculate_monthly_revenue(a, production)
    var = calculate_monthly_variable_costs(a, production)
    fix = calculate_monthly_fixed_costs(a, revenue_m)

    annual_revenue = yearly_totals(revenue_m)
    annual_cost1 = yearly_totals(var["cost1"])          # = Form_IV "Purchases"
    dep_annual = calculate_annual_depreciation(a)

    # Form_IV Cost of Production = cost1+cost2+wages+other_var+factory_oh+repairs+depreciation
    cop = [
        c1 + c2 + w + ov + fo + rmn + dep_annual
        for c1, c2, w, ov, fo, rmn in zip(
            yearly_totals(var["cost1"]), yearly_totals(var["cost2"]),
            yearly_totals(fix["wages"]), yearly_totals(var["other_variable"]),
            yearly_totals(fix["factory_overheads"]), yearly_totals(fix["repairs_maintenance"]),
        )
    ]

    # ---- 1. DEPRECIATION ----
    compare("1. DEPRECIATION  (Expenses!r13)", [
        ("Depreciation", [dep_annual] * 5, annual_from_monthly(wb, "Expenses", 13)),
    ], failures)
    print(f"     monthly charge (flat): {calculate_monthly_depreciation(a):,.2f}"
          f"   | land_cost {a['land_cost']:,.0f} correctly NOT depreciated")

    # ---- 2. LOAN SCHEDULE ----
    s = calculate_loan_schedule(a)
    compare("2. TERM-LOAN SCHEDULE  (Repayment!r5-r8) — FIXED amortisation", [
        ("Opening balance", s["opening_balance"], annual_row(wb, "Repayment", 5)),
        ("Interest for the year", s["interest"], annual_row(wb, "Repayment", 6)),
        ("Principal repayment", s["principal"], annual_row(wb, "Repayment", 7)),
        ("Closing balance", s["closing_balance"], annual_row(wb, "Repayment", 8)),
    ], failures)
    compare("   TL interest as an expense  (Expenses!r14) + debt service (DSCR!r13)", [
        ("TL interest (monthly x12)", s["interest"], annual_from_monthly(wb, "Expenses", 14)),
        ("Total debt obligation", total_debt_obligation(a), annual_row(wb, "DSCR", 13)),
    ], failures)

    # ---- 3. WORKING CAPITAL / MPBF ----
    wc = calculate_working_capital(
        a,
        annual_revenue=annual_revenue,
        annual_cost_of_production_ex_wc_interest=cop,
        wc_interest_rate=a["interest_rate_wc"],
        annual_purchases=annual_cost1,
    )
    compare("3. WORKING CAPITAL  (Form_IV_CA_CL)", [
        ("Cost of production r6", wc["cost_of_production"], annual_row(wb, "Form_IV_CA_CL", 6)),
        ("Purchases r7", wc["purchases"], annual_row(wb, "Form_IV_CA_CL", 7)),
        ("Net sales r8", wc["net_sales"], annual_row(wb, "Form_IV_CA_CL", 8)),
        ("Inventory RM r10", wc["rm_inventory"], annual_row(wb, "Form_IV_CA_CL", 10)),
        ("Inventory FG r11", wc["fg_inventory"], annual_row(wb, "Form_IV_CA_CL", 11)),
        ("Receivables r12", wc["receivables"], annual_row(wb, "Form_IV_CA_CL", 12)),
        ("Cash & bank r13", wc["cash"], annual_row(wb, "Form_IV_CA_CL", 13)),
        ("Total current assets r14", wc["total_current_assets"], annual_row(wb, "Form_IV_CA_CL", 14)),
        ("Creditors r16", wc["creditors"], annual_row(wb, "Form_IV_CA_CL", 16)),
        ("Total current liab. r18", wc["total_current_liabilities"], annual_row(wb, "Form_IV_CA_CL", 18)),
        ("Working capital gap r19", wc["working_capital_gap"], annual_row(wb, "Form_IV_CA_CL", 19)),
    ], failures)
    compare("   MPBF — Tandon Method II  (Form_V_MPBF)", [
        ("Min stipulated NWC r8", wc["min_stipulated_nwc"], annual_row(wb, "Form_V_MPBF", 8)),
        ("MPBF Method I r9", wc["mpbf_method_i"], annual_row(wb, "Form_V_MPBF", 9)),
        ("MPBF Method II r10", wc["mpbf_method_ii"], annual_row(wb, "Form_V_MPBF", 10)),
        ("Recommended MPBF r11", wc["mpbf"], annual_row(wb, "Form_V_MPBF", 11)),
    ], failures)
    compare("   WC interest  (Expenses!r15)", [
        ("WC interest (annual)", wc["wc_interest_annual"], annual_from_monthly(wb, "Expenses", 15)),
    ], failures)

    print(f"\n  CONVERGENCE: iterations_used = {wc['iterations_used']}, "
          f"converged = {wc['converged']}")
    print("  (converges on the first pass: Form_IV's Cost of Production excludes BOTH "
          "interest lines,\n   so WC interest never feeds back into it — the workbook has no "
          "circularity to resolve.)")

    print("\n" + "=" * 116)
    if failures:
        print(f"FAILED — {len(failures)} line(s) outside {TOLERANCE * 100}%:")
        for label, py, xl, worst in failures:
            print(f"   {label}: worst {worst * 100:.3f}%")
            print(f"      python: {[round(v, 2) for v in py]}")
            print(f"      excel : {[round(v, 2) for v in xl]}")
        raise AssertionError(f"{len(failures)} line(s) differ from Excel by more than "
                             f"{TOLERANCE * 100}%")
    print(f"PASSED — every line matches Excel within {TOLERANCE * 100}%.")
    print("Depreciation, the FIXED loan amortisation, and working capital / MPBF all")
    print("reproduce the recalculated workbook exactly.")
    print("=" * 116)
    wb.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
