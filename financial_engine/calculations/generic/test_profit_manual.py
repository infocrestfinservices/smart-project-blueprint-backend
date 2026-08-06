"""
test_profit_manual.py

Parity check for the P&L (EBITDA -> EBIT -> PBT -> Tax -> PAT -> Cash Accrual)
against the real post-fix workbooks in generated_reports/, recalculated by
LibreOffice.

Chains every previously-proven calculator (revenue, expenses, depreciation, loan
schedule, working capital) into profit_calc, so this also proves the modules compose
correctly — not just that each works alone.

Run from backend/:
    python financial_engine/calculations/generic/test_profit_manual.py
"""

import glob
import os
import sys
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
BACKEND_DIR = HERE.parents[2]
sys.path.insert(0, str(BACKEND_DIR))

from financial_engine.calculations.generic.test_loan_depreciation_wc_manual import (  # noqa: E402
    read_assumptions, annual_row, annual_from_monthly,
)
from financial_engine.calculations.generic.revenue_calc import (  # noqa: E402
    calculate_monthly_production, calculate_monthly_revenue, yearly_totals,
)
from financial_engine.calculations.generic.expense_calc import (  # noqa: E402
    calculate_monthly_variable_costs, calculate_monthly_fixed_costs,
)
from financial_engine.calculations.generic.depreciation_calc import (  # noqa: E402
    calculate_monthly_depreciation, calculate_annual_depreciation,
)
from financial_engine.calculations.generic.loan_schedule_calc import (  # noqa: E402
    calculate_loan_schedule,
)
from financial_engine.calculations.generic.working_capital_calc import (  # noqa: E402
    calculate_working_capital,
)
from financial_engine.calculations.generic.profit_calc import (  # noqa: E402
    calculate_profit_and_loss,
)

TOLERANCE = 0.005  # 0.5%

# The five annual-total columns of the 60-month sheets (each caps a 12-month block).
ANNUAL_TOTAL_COLS = ["N", "AA", "AN", "BA", "BN"]

# Rows 20/21/22 read the annual-total column directly, not summed from monthly cells
# — the monthly cells intentionally still hold the old per-month tax logic for
# month-level detail views, while the annual-total columns carry the corrected
# annual-PBT-based tax calculation (see the Repayment/tax-fix history). Summing
# monthly cells here would silently re-introduce the bug the fix corrected.
_ANNUAL_TOTAL_ROWS = {20, 21, 22}


def profit_annual(wb, row):
    """Ground truth for a Profit-sheet row, read the way the fixed workbook means it:
    rows 20/21/22 from the annual-total column (N/AA/AN/BA/BN), all other rows summed
    from their 12 monthly cells (which for tax-free rows equals the total anyway)."""
    if row in _ANNUAL_TOTAL_ROWS:
        ws = wb["Profit"]
        return [float(ws[f"{c}{row}"].value or 0) for c in ANNUAL_TOTAL_COLS]
    return annual_from_monthly(wb, "Profit", row)


# Profit sheet rows -> the key returned by calculate_profit_and_loss
PROFIT_ROWS = [
    ("EBITDA", 14, "ebitda"),
    ("EBIT (Operating Profit)", 16, "ebit"),
    ("Profit Before Tax", 19, "pbt"),
    ("Income Tax", 20, "income_tax"),
    ("Profit After Tax", 21, "pat"),
    ("Cash Accrual", 22, "cash_accrual"),
]


def build_pl(wb, a):
    """Chain all the calculators to produce the P&L for this workbook's assumptions."""
    production = calculate_monthly_production(a)
    revenue = calculate_monthly_revenue(a, production)
    var = calculate_monthly_variable_costs(a, production)
    fix = calculate_monthly_fixed_costs(a, revenue)
    dep_m = calculate_monthly_depreciation(a)
    dep_a = calculate_annual_depreciation(a)

    schedule = calculate_loan_schedule(a)

    cop = [c1 + c2 + w + ov + fo + rmn + dep_a for c1, c2, w, ov, fo, rmn in zip(
        yearly_totals(var["cost1"]), yearly_totals(var["cost2"]),
        yearly_totals(fix["wages"]), yearly_totals(var["other_variable"]),
        yearly_totals(fix["factory_overheads"]), yearly_totals(fix["repairs_maintenance"]))]
    wc = calculate_working_capital(
        a,
        annual_revenue=yearly_totals(revenue),
        annual_cost_of_production_ex_wc_interest=cop,
        wc_interest_rate=a["interest_rate_wc"],
        annual_purchases=yearly_totals(var["cost1"]),
    )

    return calculate_profit_and_loss(
        a,
        monthly_revenue=revenue,
        monthly_variable_costs=var,
        monthly_fixed_costs=fix,
        monthly_depreciation=dep_m,
        loan_interest_annual=schedule["interest"],
        wc_interest_annual=wc["wc_interest_annual"],
    )


def main():
    files = sorted(p for p in glob.glob(str(BACKEND_DIR / "generated_reports" / "*.xlsx"))
                   if not os.path.basename(p).startswith("~$"))
    if not files:
        print("[!] No workbooks in generated_reports/.")
        return 1

    failures = []
    overall = 0.0

    for path in files:
        wb = load_workbook(path, data_only=True)
        a = read_assumptions(wb)
        pl = build_pl(wb, a)

        print("=" * 112)
        print(f"{os.path.basename(path)}")
        print(f"  tax rate {a['income_tax_rate']} | seasonality flat? "
              f"{len(set(a['monthly_seasonality_weights'])) == 1}")
        print("=" * 112)
        print(f"  {'Line':<26}{'Yr1':>16}{'Yr2':>16}{'Yr3':>16}{'Yr4':>16}{'Yr5':>16}   max%")
        print("  " + "-" * 108)

        for label, row, key in PROFIT_ROWS:
            py = pl[key]
            xl = profit_annual(wb, row)
            pcts = []
            for p, e in zip(py, xl):
                denom = abs(e) if abs(e) > 1e-9 else (abs(p) if abs(p) > 1e-9 else 1.0)
                pcts.append(abs(p - e) / denom)
            worst = max(pcts)
            overall = max(overall, worst)
            if worst > TOLERANCE:
                failures.append((os.path.basename(path), label, py, xl, worst))
            flag = "OK" if worst <= TOLERANCE else "MISMATCH"
            print(f"  {label + ' [py]':<26}" + "".join(f"{v:>16,.0f}" for v in py))
            print(f"  {'  [excel]':<26}" + "".join(f"{v:>16,.0f}" for v in xl)
                  + f"   {worst * 100:.4f}%  {flag}")

        # cross-check against Annual_Summary's PAT (r26), which the DSCR sheet consumes
        xl_pat = annual_row(wb, "Annual_Summary", 26)
        d = max(abs(p - e) / (abs(e) if abs(e) > 1e-9 else 1.0)
                for p, e in zip(pl["pat"], xl_pat))
        overall = max(overall, d)
        print(f"\n  PAT vs Annual_Summary!r26 (the figure DSCR consumes): {d * 100:.4f}%  "
              f"{'OK' if d <= TOLERANCE else 'MISMATCH'}")

        # surface the monthly-tax quirk where it actually bites
        neg_years = [i + 1 for i, v in enumerate(pl["pbt"]) if v < 0]
        if neg_years:
            print(f"  loss-making year(s): {neg_years} -> tax charged anyway: "
                  + ", ".join(f"Y{y}={pl['income_tax'][y - 1]:,.0f}" for y in neg_years))
            print("     (Excel taxes each MONTH independently with no loss relief — mirrored.)")
        wb.close()
        print()

    print("=" * 112)
    print(f"WORST DEVIATION ACROSS {len(files)} WORKBOOKS: {overall * 100:.8f}%")
    if failures:
        print(f"FAILED — {len(failures)} line(s) outside {TOLERANCE * 100}%:")
        for wbname, label, py, xl, worst in failures:
            print(f"  {wbname} / {label}: {worst * 100:.3f}%")
            print(f"    python: {[round(v, 2) for v in py]}")
            print(f"    excel : {[round(v, 2) for v in xl]}")
        raise AssertionError(f"{len(failures)} line(s) differ from Excel by more than "
                             f"{TOLERANCE * 100}%")
    print(f"PASSED — EBITDA, EBIT, PBT, Income Tax, PAT and Cash Accrual all match Excel")
    print(f"within {TOLERANCE * 100}% across every year of every workbook.")
    print("=" * 112)
    return 0


if __name__ == "__main__":
    sys.exit(main())
